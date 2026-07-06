"""FastAPI Main Application for OLT Manager"""
import asyncio
import logging
import requests
import bcrypt
import json
from datetime import datetime, timedelta
from zoneinfo import ZoneInfo
from contextlib import asynccontextmanager
from typing import Optional, List
from concurrent.futures import ThreadPoolExecutor

import os
import sys
import uuid
import shutil
from pathlib import Path
from fastapi import FastAPI, Depends, HTTPException, Query, BackgroundTasks, UploadFile, File, WebSocket, WebSocketDisconnect, Body
from fastapi.middleware.cors import CORSMiddleware
from fastapi.staticfiles import StaticFiles
from typing import Dict, Set
from sqlalchemy.orm import Session
from sqlalchemy import func, or_
from pydantic import BaseModel

from models import (
    init_db, get_db, OLT, ONU, PollLog, Region, User, user_olts, Settings,
    TrafficSnapshot, TrafficHistory, Diagram, OLTPort, EventLog, ScheduledTask,
    ConfigBackup, AlertRule, SentAlert, SystemBackup, BackupSettings,
    Tenant, Workspace, set_session_tenant, AgentKey,
)
from tenancy import tenant_session
from schemas import (
    OLTCreate, OLTUpdate, OLTResponse, OLTListResponse,
    ONUResponse, ONUListResponse, DashboardStats, PollResult,
    RegionCreate, RegionUpdate, RegionResponse, RegionListResponse,
    UserLogin, UserCreate, UserUpdate, UserResponse, UserListResponse, LoginResponse,
    DiagramCreate, DiagramUpdate, DiagramResponse, DiagramListResponse
)
from olt_connector import poll_olt_snmp, get_traffic_counters_snmp, get_olt_health_snmp, ONUData, OLTConnector
from olt_web_scraper import get_onu_opm_data_web, get_onu_models_web, get_onu_list_web, get_onu_offline_reason_web, get_onu_status_info_web
from olt_drivers import (
    get_driver,
    get_driver_class,
    list_supported_models,
    check_model_support,
    DriverPollResult,
)
from trap_receiver import SimpleTrapReceiver, TrapEvent
from traffic_rate import compute_traffic_rate, RateInput
from config import POLL_INTERVAL, encrypt_sensitive, decrypt_sensitive
from auth import (
    authenticate_user, create_access_token, get_password_hash,
    require_auth, require_admin, get_current_user, create_default_admin
)

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

# Thread pool for blocking I/O operations (SNMP, web scraping, etc.)
thread_executor = ThreadPoolExecutor(max_workers=5)

# Helper function to get current time in user's timezone
def get_user_timezone(db: Session) -> str:
    """Get the user's configured timezone from settings"""
    try:
        tz_setting = db.query(Settings).filter(Settings.key == 'timezone').first()
        if tz_setting and tz_setting.value:
            return tz_setting.value
    except:
        pass
    return 'UTC'

def get_current_time_in_timezone(db: Session) -> datetime:
    """Get current time converted to user's timezone"""
    try:
        tz_name = get_user_timezone(db)
        tz = ZoneInfo(tz_name)
        return datetime.now(tz)
    except:
        return datetime.now()

def format_timestamp_for_filename(db: Session) -> str:
    """Get timestamp string for filenames in user's timezone"""
    try:
        tz_name = get_user_timezone(db)
        tz = ZoneInfo(tz_name)
        return datetime.now(tz).strftime('%Y%m%d_%H%M%S')
    except:
        return datetime.now().strftime('%Y%m%d_%H%M%S')

# Cleanup counter - run cleanup every N poll cycles
cleanup_counter = 0
CLEANUP_INTERVAL_CYCLES = 60  # Run cleanup every 60 poll cycles (once per hour if polling every minute)

# Background polling task handle
polling_task: Optional[asyncio.Task] = None

# SNMP Trap receiver
trap_receiver: Optional[SimpleTrapReceiver] = None
trap_task: Optional[asyncio.Task] = None
fallback_task: Optional[asyncio.Task] = None

# Weak signal alert tracking to prevent notification spam
# Format: {onu_id: datetime_of_last_alert}
# Alerts are suppressed for 1 hour after being sent
weak_signal_alert_cache: Dict[int, datetime] = {}
# Per-OLT cooldown for high-temperature alerts (else a hot OLT spams every poll).
high_temp_alert_cache: Dict[int, datetime] = {}
HIGH_TEMP_COOLDOWN_S = 3600  # 1 hour
WEAK_SIGNAL_ALERT_COOLDOWN_HOURS = 1  # Don't re-alert for the same ONU within this time

# WebSocket connection manager for live traffic
class TrafficConnectionManager:
    """Manages WebSocket connections for live traffic updates"""
    def __init__(self):
        # olt_id -> set of websocket connections
        self.active_connections: Dict[int, Set[WebSocket]] = {}
        self.traffic_tasks: Dict[int, asyncio.Task] = {}

    async def connect(self, websocket: WebSocket, olt_id: int):
        await websocket.accept()
        if olt_id not in self.active_connections:
            self.active_connections[olt_id] = set()
        self.active_connections[olt_id].add(websocket)
        logger.info(f"WebSocket connected for OLT {olt_id}. Total connections: {len(self.active_connections[olt_id])}")

    def disconnect(self, websocket: WebSocket, olt_id: int):
        if olt_id in self.active_connections:
            self.active_connections[olt_id].discard(websocket)
            if not self.active_connections[olt_id]:
                del self.active_connections[olt_id]
                # Stop traffic polling task if no connections
                if olt_id in self.traffic_tasks:
                    self.traffic_tasks[olt_id].cancel()
                    del self.traffic_tasks[olt_id]
        logger.info(f"WebSocket disconnected for OLT {olt_id}")

    async def broadcast(self, olt_id: int, message: dict):
        if olt_id in self.active_connections:
            connections = list(self.active_connections[olt_id])
            if connections:
                logger.info(f"Broadcasting to {len(connections)} clients for OLT {olt_id}: {message.get('onu_count', 0)} ONUs, poll_ms={message.get('poll_ms', 'N/A')}")
            dead_connections = set()
            for connection in connections:
                try:
                    await connection.send_json(message)
                except Exception as e:
                    logger.warning(f"Failed to send to WebSocket for OLT {olt_id}: {e}")
                    dead_connections.add(connection)
            # Clean up dead connections
            for conn in dead_connections:
                self.active_connections[olt_id].discard(conn)

traffic_manager = TrafficConnectionManager()


def parse_image_urls(image_urls_json: str) -> Optional[List[str]]:
    """Parse JSON image_urls field to list"""
    if not image_urls_json:
        return None
    try:
        return json.loads(image_urls_json)
    except:
        return None


def get_whatsapp_settings(db: Session) -> dict:
    """Get WhatsApp notification settings from database"""
    settings = {}
    for key in ['whatsapp_enabled', 'whatsapp_api_url', 'whatsapp_secret',
                'whatsapp_account', 'whatsapp_recipients']:
        setting = db.query(Settings).filter(Settings.key == key).first()
        if setting:
            settings[key] = setting.value
    return settings


def parse_whatsapp_recipients(recipients_json: str) -> list:
    """Parse JSON recipients field to list of {name, phone}"""
    if not recipients_json:
        return []
    try:
        recipients = json.loads(recipients_json)
        if isinstance(recipients, list):
            return recipients
        return []
    except:
        return []


def send_whatsapp_notification_batch(db: Session, online_onus: list, offline_onus: list, olt_name: str):
    """Send a single WhatsApp notification for all ONU status changes in a poll cycle"""
    try:
        # Log entry for debugging
        logger.info(f"send_whatsapp_notification_batch called: {len(online_onus)} online, {len(offline_onus)} offline for {olt_name}")

        # Skip if no changes
        if not online_onus and not offline_onus:
            logger.debug("No ONUs to notify about (both lists empty)")
            return

        # Check alarm settings
        alarm_settings = get_alarm_settings(db)

        # Check quiet hours
        if is_in_quiet_hours(alarm_settings, db):
            logger.info("Skipping ONU status notifications - quiet hours")
            return

        # Filter based on alarm settings
        online_alarm_enabled = is_alarm_enabled(alarm_settings, "onu_back_online")
        offline_alarm_enabled = is_alarm_enabled(alarm_settings, "onu_offline")
        logger.info(f"Alarm settings - onu_back_online: {online_alarm_enabled}, onu_offline: {offline_alarm_enabled}")

        filtered_online = online_onus if online_alarm_enabled else []
        filtered_offline = offline_onus if offline_alarm_enabled else []

        # Apply ONU/Region filtering based on alarm settings
        filtered_online = filter_onus_by_selection(filtered_online, alarm_settings)
        filtered_offline = filter_onus_by_selection(filtered_offline, alarm_settings)

        logger.info(f"After filtering - online: {len(filtered_online)}, offline: {len(filtered_offline)}")

        # Skip if nothing to notify after filtering
        if not filtered_online and not filtered_offline:
            logger.info("Skipping ONU status notifications - alarms disabled for these types or no ONUs to notify")
            return

        settings = get_whatsapp_settings(db)

        # Check if WhatsApp notifications are enabled (case-insensitive)
        if str(settings.get('whatsapp_enabled', '')).lower() != 'true':
            return

        api_url = settings.get('whatsapp_api_url', '').strip()
        secret = decrypt_sensitive(settings.get('whatsapp_secret', '')).strip()
        account = settings.get('whatsapp_account', '').strip()

        # Parse recipients (supports both old single recipient and new multiple format)
        recipients = parse_whatsapp_recipients(settings.get('whatsapp_recipients', ''))

        if not recipients:
            logger.warning("No WhatsApp recipients configured, skipping notification")
            return

        if not all([api_url, secret, account]):
            logger.warning("WhatsApp settings incomplete, skipping notification")
            return

        # Build consolidated message with detailed format
        message_parts = []

        # Add offline ONUs section
        for onu in filtered_offline:
            onu_name = onu.description if onu.description and onu.description.upper() != "NULL" else "No Name"
            region_name = onu.region.name if onu.region else "No Region"

            message_parts.append("🔴 *ONU OFFLINE*")
            message_parts.append("")
            message_parts.append(f"Description: {onu_name}")
            message_parts.append(f"OLT: {olt_name}")
            message_parts.append(f"Region: {region_name}")
            message_parts.append(f"MAC: {onu.mac_address}")
            message_parts.append(f"Port: {onu.pon_port}/{onu.onu_id}")
            # Add last signal before disconnect (prefer onu_rx_power from ONU, fallback to rx_power from OLT)
            last_signal = onu.onu_rx_power if onu.onu_rx_power is not None else onu.rx_power
            if last_signal is not None:
                message_parts.append(f"Last Signal: {last_signal} dBm")
            # Add last distance before disconnect
            if onu.distance is not None:
                distance_km = onu.distance / 1000 if onu.distance >= 1000 else None
                if distance_km:
                    message_parts.append(f"Last Distance: {distance_km:.2f} km ({onu.distance} m)")
                else:
                    message_parts.append(f"Last Distance: {onu.distance} m")
            if onu.latitude and onu.longitude:
                message_parts.append(f"Location: https://maps.google.com/?q={onu.latitude},{onu.longitude}")
            if onu.address:
                message_parts.append(f"Address: {onu.address}")
            message_parts.append("")
            message_parts.append(f"Time: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
            message_parts.append("")
            message_parts.append("─" * 20)
            message_parts.append("")

        # Add online ONUs section
        for onu in filtered_online:
            onu_name = onu.description if onu.description and onu.description.upper() != "NULL" else "No Name"
            region_name = onu.region.name if onu.region else "No Region"

            message_parts.append("🟢 *ONU ONLINE*")
            message_parts.append("")
            message_parts.append(f"Description: {onu_name}")
            message_parts.append(f"OLT: {olt_name}")
            message_parts.append(f"Region: {region_name}")
            message_parts.append(f"MAC: {onu.mac_address}")
            message_parts.append(f"Port: {onu.pon_port}/{onu.onu_id}")
            if onu.latitude and onu.longitude:
                message_parts.append(f"Location: https://maps.google.com/?q={onu.latitude},{onu.longitude}")
            if onu.address:
                message_parts.append(f"Address: {onu.address}")
            message_parts.append("")
            message_parts.append(f"Time: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
            message_parts.append("")
            message_parts.append("─" * 20)
            message_parts.append("")

        # Remove trailing separator
        if message_parts and message_parts[-1] == "":
            message_parts.pop()
        if message_parts and message_parts[-1] == "─" * 20:
            message_parts.pop()
        if message_parts and message_parts[-1] == "":
            message_parts.pop()

        message = "\n".join(message_parts)

        # Send WhatsApp message to all recipients
        success_count = 0
        for recipient in recipients:
            phone = recipient.get('phone', '').strip()
            name = recipient.get('name', 'Unknown')

            if not phone:
                continue

            try:
                response = requests.post(
                    api_url,
                    data={
                        'secret': secret,
                        'account': account,
                        'recipient': phone,
                        'type': 'text',
                        'message': message,
                        'priority': 1
                    },
                    timeout=30
                )

                if _wa_ok(response):
                    success_count += 1
                    logger.info(f"WhatsApp notification sent to {name} ({phone})")
                else:
                    logger.error(f"WhatsApp notification failed for {name} ({phone}): {response.text}")
            except Exception as e:
                logger.error(f"Failed to send to {name} ({phone}): {e}")

        logger.info(f"WhatsApp batch notification: {success_count}/{len(recipients)} recipients, {len(filtered_offline)} offline, {len(filtered_online)} online")

    except Exception as e:
        logger.error(f"Failed to send WhatsApp notification: {e}")


def get_alarm_settings(db: Session) -> dict:
    """Get alarm settings from database"""
    settings = db.query(Settings).filter(Settings.key.like('alarm_%')).all()
    result = {}
    for s in settings:
        key = s.key.replace('alarm_', '')
        result[key] = s.value

    # Apply defaults
    defaults = {
        "new_onu_registration": "true",
        "onu_offline": "true",
        "onu_back_online": "true",
        "olt_offline": "true",
        "olt_back_online": "true",
        "weak_signal": "false",
        "weak_signal_threshold": "-25",
        "weak_signal_lower_threshold": "-30",
        "high_temperature": "false",
        "high_temperature_threshold": "60",
        "selected_onus": "[]",
        "selected_regions": "[]",
        "quiet_hours_enabled": "false",
        "quiet_hours_start": "22:00",
        "quiet_hours_end": "07:00"
    }
    for key, value in defaults.items():
        if key not in result:
            result[key] = value

    return result


def is_alarm_enabled(alarm_settings: dict, alarm_type: str) -> bool:
    """Check if a specific alarm type is enabled"""
    value = alarm_settings.get(alarm_type, "false")
    return str(value).lower() == "true"


def _wa_ok(response) -> bool:
    """Best-effort delivery check for the WhatsApp gateway.

    Many such gateways return HTTP 200 with a JSON error body, so a bare
    status_code==200 counts failures as successes. Treat as failure only when
    the body clearly signals an error (avoids false negatives on unknown shapes).
    """
    if response.status_code != 200:
        return False
    try:
        j = response.json()
        if isinstance(j, dict):
            st = str(j.get("status", j.get("result", j.get("success", "")))).strip().lower()
            if st in ("error", "failed", "failure", "false", "0"):
                return False
            if j.get("error") or j.get("errors"):
                return False
    except Exception:
        pass
    return True


def is_in_quiet_hours(alarm_settings: dict, db: Session = None) -> bool:
    """Check if current time is within quiet hours"""
    if not is_alarm_enabled(alarm_settings, "quiet_hours_enabled"):
        return False

    try:
        # Use the configured (tenant) timezone, not naive server-local time,
        # so the quiet window lands on the right wall-clock hours.
        now = (get_current_time_in_timezone(db).time() if db is not None
               else datetime.now().time())
        start_str = alarm_settings.get("quiet_hours_start", "22:00")
        end_str = alarm_settings.get("quiet_hours_end", "07:00")

        start_parts = start_str.split(":")
        end_parts = end_str.split(":")

        start_time = datetime.strptime(f"{start_parts[0]}:{start_parts[1]}", "%H:%M").time()
        end_time = datetime.strptime(f"{end_parts[0]}:{end_parts[1]}", "%H:%M").time()

        # Handle overnight quiet hours (e.g., 22:00 to 07:00)
        if start_time > end_time:
            return now >= start_time or now <= end_time
        else:
            return start_time <= now <= end_time
    except Exception as e:
        logger.error(f"Error checking quiet hours: {e}")
        return False


def filter_onus_by_selection(onus: list, alarm_settings: dict) -> list:
    """Filter ONUs based on selected_onus and selected_regions in alarm settings.

    If no ONUs or regions are selected, returns all ONUs (no filtering).
    If specific ONUs are selected, only those ONUs will pass the filter.
    If specific regions are selected, only ONUs in those regions will pass the filter.
    If both are selected, ONUs matching either criterion will pass.
    """
    selected_onus_raw = alarm_settings.get("selected_onus", "[]")
    selected_regions_raw = alarm_settings.get("selected_regions", "[]")

    logger.info(f"filter_onus_by_selection - raw selected_onus: {repr(selected_onus_raw)}, raw selected_regions: {repr(selected_regions_raw)}")

    # Handle selected_onus - could be a list (already parsed) or a JSON string
    if isinstance(selected_onus_raw, list):
        selected_onus = selected_onus_raw
    elif isinstance(selected_onus_raw, str) and selected_onus_raw:
        try:
            selected_onus = json.loads(selected_onus_raw)
        except Exception as e:
            logger.error(f"Failed to parse selected_onus: {e}")
            selected_onus = []
    else:
        selected_onus = []

    # Handle selected_regions - could be a list (already parsed) or a JSON string
    if isinstance(selected_regions_raw, list):
        selected_regions = selected_regions_raw
    elif isinstance(selected_regions_raw, str) and selected_regions_raw:
        try:
            selected_regions = json.loads(selected_regions_raw)
        except Exception as e:
            logger.error(f"Failed to parse selected_regions: {e}")
            selected_regions = []
    else:
        selected_regions = []

    logger.info(f"filter_onus_by_selection - parsed selected_onus: {selected_onus}, parsed selected_regions: {selected_regions}")

    # If no specific ONUs or regions are selected, return all ONUs
    if not selected_onus and not selected_regions:
        logger.info("No ONU/Region filtering configured - returning all ONUs")
        return onus

    logger.info(f"Applying ONU/Region filter - selected_onus: {selected_onus}, selected_regions: {selected_regions}")

    def onu_matches_filter(onu):
        # Check if ONU ID is in selected_onus list
        if selected_onus and onu.id in selected_onus:
            return True
        # Check if ONU's region is in selected_regions list
        if selected_regions and onu.region_id and onu.region_id in selected_regions:
            return True
        # Didn't match any selection
        return False

    filtered = [onu for onu in onus if onu_matches_filter(onu)]
    logger.info(f"After ONU/Region filter: {len(filtered)} of {len(onus)} ONUs passed")
    return filtered


def send_new_onu_notification(db: Session, onu, olt_name: str):
    """Send WhatsApp notification for new ONU registration"""
    try:
        alarm_settings = get_alarm_settings(db)

        # Check if new ONU registration alarm is enabled
        if not is_alarm_enabled(alarm_settings, "new_onu_registration"):
            logger.debug("New ONU registration alarm is disabled")
            return

        # Check quiet hours
        if is_in_quiet_hours(alarm_settings, db):
            logger.debug("Skipping new ONU notification - quiet hours")
            return

        settings = get_whatsapp_settings(db)

        # Check if WhatsApp notifications are enabled
        if str(settings.get('whatsapp_enabled', '')).lower() != 'true':
            return

        api_url = settings.get('whatsapp_api_url', '').strip()
        secret = decrypt_sensitive(settings.get('whatsapp_secret', '')).strip()
        account = settings.get('whatsapp_account', '').strip()
        recipients = parse_whatsapp_recipients(settings.get('whatsapp_recipients', ''))

        if not recipients or not all([api_url, secret, account]):
            return

        # Build message with signal info
        message_parts = [
            "🆕 *NEW ONU REGISTERED*",
            "",
            f"MAC: {onu.mac_address}",
            f"OLT: {olt_name}",
            f"Port: {onu.pon_port}/{onu.onu_id}",
        ]

        # Add model if available
        if onu.model:
            message_parts.append(f"Model: {onu.model}")

        # Add signal information section
        signal_info = []
        if onu.distance is not None:
            signal_info.append(f"Distance: {onu.distance}m")
        if onu.onu_rx_power is not None:
            signal_info.append(f"RX Power: {onu.onu_rx_power} dBm")
        if onu.onu_tx_power is not None:
            signal_info.append(f"TX Power: {onu.onu_tx_power} dBm")
        if onu.onu_temperature is not None:
            signal_info.append(f"Temperature: {onu.onu_temperature}°C")
        if onu.onu_voltage is not None:
            signal_info.append(f"Voltage: {onu.onu_voltage}V")

        if signal_info:
            message_parts.append("")
            message_parts.append("📊 *Signal Info:*")
            message_parts.extend(signal_info)

        message_parts.append("")
        message_parts.append(f"Time: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")

        message = "\n".join(message_parts)

        # Send to all recipients
        for recipient in recipients:
            phone = recipient.get('phone', '').strip()
            if not phone:
                continue

            try:
                response = requests.post(
                    api_url,
                    data={
                        'secret': secret,
                        'account': account,
                        'recipient': phone,
                        'type': 'text',
                        'message': message
                    },
                    timeout=10
                )
                if _wa_ok(response):
                    logger.info(f"New ONU notification sent to {phone}")
                else:
                    logger.warning(f"Failed to send new ONU notification to {phone}: {response.text}")
            except Exception as e:
                logger.error(f"Error sending new ONU notification to {phone}: {e}")

    except Exception as e:
        logger.error(f"Failed to send new ONU notification: {e}")


def send_olt_status_notification(db: Session, olt, is_online: bool):
    """Send WhatsApp notification for OLT status change"""
    try:
        alarm_settings = get_alarm_settings(db)

        # Check if appropriate alarm is enabled
        alarm_type = "olt_back_online" if is_online else "olt_offline"
        if not is_alarm_enabled(alarm_settings, alarm_type):
            logger.debug(f"OLT {alarm_type} alarm is disabled")
            return

        # Check quiet hours
        if is_in_quiet_hours(alarm_settings, db):
            logger.debug("Skipping OLT status notification - quiet hours")
            return

        settings = get_whatsapp_settings(db)

        # Check if WhatsApp notifications are enabled
        if str(settings.get('whatsapp_enabled', '')).lower() != 'true':
            return

        api_url = settings.get('whatsapp_api_url', '').strip()
        secret = decrypt_sensitive(settings.get('whatsapp_secret', '')).strip()
        account = settings.get('whatsapp_account', '').strip()
        recipients = parse_whatsapp_recipients(settings.get('whatsapp_recipients', ''))

        if not recipients or not all([api_url, secret, account]):
            return

        # Build message
        if is_online:
            emoji = "✅"
            status = "BACK ONLINE"
        else:
            emoji = "🔴"
            status = "OFFLINE"

        message_parts = [
            f"{emoji} *OLT {status}*",
            "",
            f"Name: {olt.name}",
            f"IP: {olt.ip_address}",
            "",
            f"Time: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        ]

        message = "\n".join(message_parts)

        # Send to all recipients
        for recipient in recipients:
            phone = recipient.get('phone', '').strip()
            if not phone:
                continue

            try:
                response = requests.post(
                    api_url,
                    data={
                        'secret': secret,
                        'account': account,
                        'recipient': phone,
                        'type': 'text',
                        'message': message
                    },
                    timeout=10
                )
                if _wa_ok(response):
                    logger.info(f"OLT status notification sent to {phone}")
                else:
                    logger.warning(f"Failed to send OLT status notification to {phone}: {response.text}")
            except Exception as e:
                logger.error(f"Error sending OLT status notification to {phone}: {e}")

    except Exception as e:
        logger.error(f"Failed to send OLT status notification: {e}")


def send_weak_signal_notification(db: Session, onus_with_weak_signal: list, olt_name: str, upper_threshold: float, lower_threshold: float):
    """Send WhatsApp notification for ONUs in the 'danger zone' (weak signal between thresholds)"""
    global weak_signal_alert_cache

    try:
        if not onus_with_weak_signal:
            return

        alarm_settings = get_alarm_settings(db)

        # Check if weak signal alarm is enabled
        if not is_alarm_enabled(alarm_settings, "weak_signal"):
            logger.debug("Weak signal alarm is disabled")
            return

        # Check quiet hours
        if is_in_quiet_hours(alarm_settings, db):
            logger.debug("Skipping weak signal notification - quiet hours")
            return

        # Apply ONU/Region filtering
        onus_with_weak_signal = filter_onus_by_selection(onus_with_weak_signal, alarm_settings)
        if not onus_with_weak_signal:
            logger.debug("No weak signal ONUs left after ONU/Region filtering")
            return

        # Filter out ONUs that were recently alerted (within cooldown period)
        now = datetime.utcnow()
        cooldown_threshold = now - timedelta(hours=WEAK_SIGNAL_ALERT_COOLDOWN_HOURS)

        # Clean up old entries from cache
        expired_ids = [onu_id for onu_id, alert_time in weak_signal_alert_cache.items()
                       if alert_time < cooldown_threshold]
        for onu_id in expired_ids:
            del weak_signal_alert_cache[onu_id]

        # Filter to only ONUs that haven't been alerted recently
        onus_to_alert = []
        for onu in onus_with_weak_signal:
            if onu.id not in weak_signal_alert_cache:
                onus_to_alert.append(onu)

        if not onus_to_alert:
            logger.debug("All weak signal ONUs were already alerted recently, skipping")
            return

        # Log events to EventLog for ALL weak signal ONUs (so they show in dashboard)
        for onu in onus_to_alert:
            onu_name = onu.description if onu.description and onu.description.upper() != "NULL" else onu.mac_address
            signal = onu.rx_power

            # Calculate risk level
            if signal is not None:
                range_size = upper_threshold - lower_threshold
                position = (signal - lower_threshold) / range_size if range_size != 0 else 0.5
                if position < 0.33:
                    risk_level = "CRITICAL"
                elif position < 0.66:
                    risk_level = "HIGH"
                else:
                    risk_level = "WARNING"
            else:
                risk_level = "WARNING"

            # Log to EventLog table
            event = EventLog(
                event_type="weak_signal",
                entity_type="onu",
                entity_id=onu.id,
                olt_id=onu.olt_id,
                description=f"Weak Signal [{risk_level}]: {onu_name} ({onu.mac_address}) on {olt_name} - Signal: {signal} dBm (threshold: {upper_threshold} dBm)",
                details=json.dumps({
                    "onu_name": onu_name,
                    "mac_address": onu.mac_address,
                    "signal": signal,
                    "threshold": upper_threshold,
                    "lower_threshold": lower_threshold,
                    "risk_level": risk_level,
                    "pon_port": onu.pon_port,
                    "onu_id": onu.onu_id
                })
            )
            db.add(event)

            # Mark as alerted in cache
            weak_signal_alert_cache[onu.id] = now

        db.commit()
        logger.info(f"Logged {len(onus_to_alert)} weak signal events to EventLog")

        # Now check if WhatsApp is configured for notifications
        settings = get_whatsapp_settings(db)

        # Check if WhatsApp notifications are enabled
        if str(settings.get('whatsapp_enabled', '')).lower() != 'true':
            return

        api_url = settings.get('whatsapp_api_url', '').strip()
        secret = decrypt_sensitive(settings.get('whatsapp_secret', '')).strip()
        account = settings.get('whatsapp_account', '').strip()
        recipients = parse_whatsapp_recipients(settings.get('whatsapp_recipients', ''))

        if not recipients or not all([api_url, secret, account]):
            return

        # Build detailed message for each ONU (onus_to_alert already filtered above)
        for onu in onus_to_alert[:5]:  # Limit to 5 ONUs per batch to avoid long messages
            onu_name = onu.description if onu.description and onu.description.upper() != "NULL" else "No Name"
            region_name = onu.region.name if onu.region else "No Region"

            # Get signal value (prefer ONU-reported rx_power, fallback to OLT-measured)
            signal = onu.rx_power

            # Calculate risk level based on proximity to lower threshold
            if signal is not None:
                range_size = upper_threshold - lower_threshold
                position = (signal - lower_threshold) / range_size if range_size != 0 else 0.5
                if position < 0.33:
                    risk_level = "CRITICAL - Disconnect imminent!"
                elif position < 0.66:
                    risk_level = "HIGH - May disconnect soon"
                else:
                    risk_level = "WARNING - Signal degrading"
            else:
                risk_level = "UNKNOWN"

            message_parts = [
                "⚠️ *WEAK SIGNAL ALERT*",
                "",
                f"Description: {onu_name}",
                f"OLT: {olt_name}",
                f"Region: {region_name}",
                f"MAC: {onu.mac_address}",
                f"Port: {onu.pon_port}/{onu.onu_id}",
                "",
                f"📶 Signal: {signal} dBm" if signal else "📶 Signal: Unknown",
                f"🎯 Danger Zone: {upper_threshold} to {lower_threshold} dBm",
                f"⚡ Risk: {risk_level}",
            ]

            # Add location if available
            if onu.latitude and onu.longitude:
                message_parts.append(f"📍 Location: https://maps.google.com/?q={onu.latitude},{onu.longitude}")
            if onu.address:
                message_parts.append(f"🏠 Address: {onu.address}")

            message_parts.append("")
            message_parts.append(f"Time: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
            message_parts.append("")
            message_parts.append("Action: Check fiber connection")

            message = "\n".join(message_parts)

            # Send to all recipients
            for recipient in recipients:
                phone = recipient.get('phone', '').strip()
                if not phone:
                    continue

                try:
                    response = requests.post(
                        api_url,
                        data={
                            'secret': secret,
                            'account': account,
                            'recipient': phone,
                            'type': 'text',
                            'message': message
                        },
                        timeout=10
                    )
                    if _wa_ok(response):
                        logger.info(f"Weak signal notification sent to {phone} for ONU {onu.mac_address}")
                    else:
                        logger.warning(f"Failed to send weak signal notification to {phone}: {response.text}")
                except Exception as e:
                    logger.error(f"Error sending weak signal notification to {phone}: {e}")

        if len(onus_to_alert) > 5:
            logger.info(f"Weak signal alert: Sent notifications for 5 ONUs, {len(onus_to_alert) - 5} more have weak signal (not yet alerted)")

    except Exception as e:
        logger.error(f"Failed to send weak signal notification: {e}")


def send_high_temperature_notification(db: Session, olt, temperature: float):
    """Send WhatsApp notification for OLT high temperature"""
    try:
        alarm_settings = get_alarm_settings(db)

        # Check if high temperature alarm is enabled
        if not is_alarm_enabled(alarm_settings, "high_temperature"):
            logger.debug("High temperature alarm is disabled")
            return

        # Check quiet hours
        if is_in_quiet_hours(alarm_settings, db):
            logger.debug("Skipping high temperature notification - quiet hours")
            return

        settings = get_whatsapp_settings(db)

        # Check if WhatsApp notifications are enabled
        if str(settings.get('whatsapp_enabled', '')).lower() != 'true':
            return

        api_url = settings.get('whatsapp_api_url', '').strip()
        secret = decrypt_sensitive(settings.get('whatsapp_secret', '')).strip()
        account = settings.get('whatsapp_account', '').strip()
        recipients = parse_whatsapp_recipients(settings.get('whatsapp_recipients', ''))

        if not recipients or not all([api_url, secret, account]):
            return

        threshold = alarm_settings.get("high_temperature_threshold", 60)

        # Build message
        message_parts = [
            f"🌡️ *HIGH TEMPERATURE ALERT*",
            "",
            f"OLT: {olt.name}",
            f"IP: {olt.ip_address}",
            f"Temperature: {temperature}°C",
            f"Threshold: {threshold}°C",
            "",
            f"Time: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        ]

        message = "\n".join(message_parts)

        # Send to all recipients
        for recipient in recipients:
            phone = recipient.get('phone', '').strip()
            if not phone:
                continue

            try:
                response = requests.post(
                    api_url,
                    data={
                        'secret': secret,
                        'account': account,
                        'recipient': phone,
                        'type': 'text',
                        'message': message
                    },
                    timeout=10
                )
                if _wa_ok(response):
                    logger.info(f"High temperature notification sent to {phone}")
                else:
                    logger.warning(f"Failed to send high temperature notification to {phone}: {response.text}")
            except Exception as e:
                logger.error(f"Error sending high temperature notification to {phone}: {e}")

    except Exception as e:
        logger.error(f"Failed to send high temperature notification: {e}")


async def collect_traffic_history(olt, db):
    """Collect traffic data and save to history for an OLT."""
    try:
        loop = asyncio.get_event_loop()
        current_counters = await loop.run_in_executor(
            thread_executor,
            get_traffic_counters_snmp,
            olt.ip_address,
            "public"
        )

        if not current_counters:
            return

        current_time = datetime.utcnow()

        # Get Mikrotik traffic rates if configured (replaces SNMP rates)
        mk_rates = {}
        if getattr(olt, 'mk_enabled', False) and getattr(olt, 'mk_ip', None):
            try:
                from mikrotik_traffic import get_mikrotik_traffic
                onu_db_map = {}
                for onu in db.query(ONU).filter(ONU.olt_id == olt.id).all():
                    onu_db_map[(onu.pon_port, onu.onu_id)] = onu.mac_address
                mk_rates = get_mikrotik_traffic(
                    mk_ip=olt.mk_ip,
                    mk_user=olt.mk_username or 'admin',
                    mk_pass=olt.mk_password or '',
                    mk_port=olt.mk_port or 8728,
                    olt_ip=olt.ip_address,
                    snmp_community=olt.snmp_community or 'public',
                    onu_db_map=onu_db_map,
                )
                if mk_rates:
                    logger.info(f"Mikrotik traffic overlay: {len(mk_rates)} ONUs from {olt.mk_ip}")
            except Exception as exc:
                logger.warning(f"Mikrotik traffic failed for {olt.name}: {exc}")

        # Get previous snapshots for this OLT
        prev_snapshots = {
            s.mac_address: s
            for s in db.query(TrafficSnapshot).filter(TrafficSnapshot.olt_id == olt.id).all()
        }

        traffic_data = []

        for key, counters in current_counters.items():
            rx_bytes = counters['rx_bytes']
            tx_bytes = counters['tx_bytes']
            pon_port = counters.get('pon_port', 0)
            onu_id = counters.get('onu_id', 0)

            # Handle both MAC and pon:onu key formats
            # V1600G2-B returns "pon:onu" keys (e.g., "1:5"), V1600D8 returns MAC keys
            if ':' in key and len(key) < 10:  # pon:onu format (short like "1:5")
                # Look up ONU by pon:onu to get MAC
                onu_for_mac = db.query(ONU).filter(
                    ONU.olt_id == olt.id,
                    ONU.pon_port == pon_port,
                    ONU.onu_id == onu_id
                ).first()
                mac = onu_for_mac.mac_address if onu_for_mac else None
                if not mac:
                    continue  # Skip if we can't find the ONU
            else:
                mac = key  # Key is already a MAC address

            rx_kbps = 0
            tx_kbps = 0

            # Look up the ONU first — the rate helper needs its online state.
            onu = db.query(ONU).filter(
                ONU.olt_id == olt.id,
                ONU.mac_address == mac
            ).first()
            onu_online = bool(onu and onu.is_online)

            if mac in prev_snapshots:
                prev = prev_snapshots[mac]
                res = compute_traffic_rate(
                    RateInput(prev.rx_bytes, prev.tx_bytes, prev.timestamp,
                              getattr(prev, 'last_rx_kbps', 0) or 0,
                              getattr(prev, 'last_tx_kbps', 0) or 0),
                    rx_bytes, tx_bytes, current_time, onu_online,
                )
                rx_kbps, tx_kbps = res.rx_kbps, res.tx_kbps
                res.apply_to(prev)
            else:
                db.add(TrafficSnapshot(
                    olt_id=olt.id,
                    mac_address=mac,
                    rx_bytes=rx_bytes,
                    tx_bytes=tx_bytes,
                    timestamp=current_time,
                    last_rx_kbps=0,
                    last_tx_kbps=0
                ))

            # Override with Mikrotik rates if available (more accurate) — but
            # never for an offline ONU (would resurrect stale/other traffic).
            if onu_online and mac in mk_rates:
                rx_kbps = mk_rates[mac]['rx_kbps']
                tx_kbps = mk_rates[mac]['tx_kbps']

            traffic_data.append({
                'onu': onu,
                'pon_port': pon_port,
                'rx_kbps': rx_kbps,
                'tx_kbps': tx_kbps
            })

            # Save ONU traffic history — write for every online ONU (including a
            # genuine 0) so idle periods render as a continuous zero line;
            # offline ONUs are skipped so their graph stops instead of plateauing.
            if onu and onu.is_online:
                onu_history = TrafficHistory(
                    entity_type='onu',
                    entity_id=str(onu.id),
                    olt_id=olt.id,
                    pon_port=pon_port,
                    onu_db_id=onu.id,
                    rx_kbps=rx_kbps,
                    tx_kbps=tx_kbps,
                    timestamp=current_time
                )
                db.add(onu_history)

        # Aggregate and save PON port history
        pon_traffic = {}
        for t in traffic_data:
            pon = t['pon_port']
            if pon not in pon_traffic:
                pon_traffic[pon] = {'rx_kbps': 0, 'tx_kbps': 0}
            pon_traffic[pon]['rx_kbps'] += t['rx_kbps']
            pon_traffic[pon]['tx_kbps'] += t['tx_kbps']

        for pon, traffic in pon_traffic.items():
            pon_history = TrafficHistory(
                entity_type='pon',
                entity_id=f"{olt.id}:{pon}",
                olt_id=olt.id,
                pon_port=pon,
                onu_db_id=None,
                rx_kbps=traffic['rx_kbps'],
                tx_kbps=traffic['tx_kbps'],
                timestamp=current_time
            )
            db.add(pon_history)

        # Save OLT total history
        total_rx = sum(t['rx_kbps'] for t in traffic_data)
        total_tx = sum(t['tx_kbps'] for t in traffic_data)
        olt_history = TrafficHistory(
            entity_type='olt',
            entity_id=str(olt.id),
            olt_id=olt.id,
            pon_port=None,
            onu_db_id=None,
            rx_kbps=total_rx,
            tx_kbps=total_tx,
            timestamp=current_time
        )
        db.add(olt_history)

        # Collect uplink port traffic via SNMP
        from models import PortTraffic
        port_counters = await loop.run_in_executor(
            thread_executor,
            poll_port_traffic_snmp,
            olt.ip_address,
            "public"
        )

        if port_counters:
            logger.info(f"Got {len(port_counters)} port counters for {olt.name}")
            port_rates = calculate_port_rates(olt.id, olt.ip_address, port_counters)
            logger.info(f"Calculated {len(port_rates)} port rates for {olt.name}")

            # Port mapping comes from the OLT driver's declared port layout.
            # New OLT models add a driver class — no edits needed here.
            try:
                port_mapping = get_driver(olt).get_port_layout().to_port_mapping()
            except ValueError:
                # Unknown model: fall back to a generic 1-8 GE layout so we
                # still record uplink traffic instead of dropping it silently.
                port_mapping = {i: ('ge', i) for i in range(1, 9)}

            uplink_count = 0
            for if_idx, rates in port_rates.items():
                if if_idx in port_mapping:
                    port_type, port_num = port_mapping[if_idx]
                    rx = rates['rx_kbps']
                    tx = rates['tx_kbps']

                    # Sanity guard: never persist an impossible port rate
                    # (negative, or above the 10 Gbps uplink ceiling). A single
                    # garbage sample otherwise blows up the graph scale and drags
                    # the average negative.
                    if rx < 0 or tx < 0 or rx > 10_000_000 or tx > 10_000_000:
                        continue

                    # Save to PortTraffic table for per-port graphs
                    port_traffic = PortTraffic(
                        olt_id=olt.id,
                        port_type=port_type,
                        port_number=port_num,  # Use mapped port number
                        rx_kbps=rx,
                        tx_kbps=tx,
                        timestamp=current_time
                    )
                    db.add(port_traffic)

                    # Also save to TrafficHistory for historical graphs
                    uplink_history = TrafficHistory(
                        entity_type=port_type,  # 'ge' or 'xge'
                        entity_id=f"{olt.id}:{port_type}:{port_num}",
                        olt_id=olt.id,
                        pon_port=None,
                        onu_db_id=None,
                        rx_kbps=rx,
                        tx_kbps=tx,
                        timestamp=current_time
                    )
                    db.add(uplink_history)
                    uplink_count += 1

            if uplink_count > 0:
                logger.info(f"Uplink traffic saved for {olt.name}: {uplink_count} ports")

        # Zero live-rate snapshots for offline ONUs so no read path (incl. the
        # WebSocket cache) can serve their last-known rate. Deregistered offline
        # ONUs don't appear in the SNMP counter table above, so they'd otherwise
        # keep a stale snapshot forever — handle them here.
        # NOTE: the same MAC can have duplicate ONU rows at stale positions
        # (one online + several stale offline). Only zero MACs that are offline
        # in ALL rows, else we'd wipe a currently-online ONU's live rate.
        online_macs = {
            m[0] for m in db.query(ONU.mac_address).filter(
                ONU.olt_id == olt.id, ONU.is_online == True
            ).all()
        }
        offline_macs = [
            m[0] for m in db.query(ONU.mac_address).filter(
                ONU.olt_id == olt.id, ONU.is_online == False
            ).all()
            if m[0] not in online_macs
        ]
        if offline_macs:
            db.query(TrafficSnapshot).filter(
                TrafficSnapshot.olt_id == olt.id,
                TrafficSnapshot.mac_address.in_(offline_macs),
            ).update(
                {TrafficSnapshot.last_rx_kbps: 0, TrafficSnapshot.last_tx_kbps: 0},
                synchronize_session=False,
            )

        logger.info(f"Traffic history saved for {olt.name}: {len(traffic_data)} ONUs, total {total_rx:.0f}/{total_tx:.0f} kbps")

    except Exception as e:
        logger.error(f"Failed to collect traffic history for {olt.name}: {e}")


async def poll_all_olts(db_session_factory, use_snmp: bool = True, tenant_id: Optional[str] = None, skip_optical: bool = False):
    """Poll all OLTs for a single tenant and update the database.

    Phase 1 (multi-tenant): this function now operates within ONE tenant's
    scope. The session is tagged with `tenant_id` so PostgreSQL RLS filters
    every query to that tenant automatically. The outer
    `poll_all_tenants` driver iterates tenants.

    `tenant_id=None` means "no tenant filtering" — only valid for the legacy
    SQLite single-tenant binary, which has no tenants table.

    When ``skip_optical`` is True the driver skips the web scrape for optical
    metrics (RX power, distance, temperature) to save ~4s per OLT.

    Uses SNMP polling for fast ONU data retrieval (~2 seconds).
    Falls back to SSH if SNMP fails.
    """
    if tenant_id:
        logger.info(f"Starting OLT polling cycle (tenant={tenant_id})")
    else:
        logger.info("Starting OLT polling cycle (legacy single-tenant mode)")

    db = db_session_factory()
    if tenant_id:
        set_session_tenant(db, tenant_id)
    try:
        olts = db.query(OLT).all()

        for olt in olts:
            # Track OLT's previous online status for alarm notifications
            olt_was_online = olt.is_online

            # Tag the session with this OLT's workspace_id so the
            # `before_flush` hook in models.py auto-fills workspace_id on
            # newly-created ONU / TrafficHistory / PortTraffic rows. Without
            # this, those inserts hit a NOT NULL violation and the entire
            # poll cycle rolls back. tenant_id was already set above.
            if olt.workspace_id:
                db.info["workspace_id"] = olt.workspace_id

            try:
                logger.info(f"Polling OLT: {olt.name} ({olt.ip_address})")

                # Get existing ONUs for this OLT
                existing_onus = {
                    o.mac_address: o
                    for o in db.query(ONU).filter(ONU.olt_id == olt.id).all()
                }

                # Get web credentials (still needed for ad-hoc lookups like
                # the offline-reason scrape further down).
                web_user = olt.web_username or olt.username or 'admin'
                web_pass = decrypt_sensitive(olt.web_password) if olt.web_password else decrypt_sensitive(olt.password) if olt.password else 'admin'

                # Resolve and run the driver for this OLT model. The driver
                # encapsulates all model-specific polling logic — adding a new
                # OLT model only requires writing a new driver class.
                snmp_onus_data = []
                snmp_status_map = {}
                web_opm_data: Dict[str, Dict] = {}
                web_model_data: Dict[str, str] = {}
                web_status_data: Dict[str, Dict] = {}
                health_data: Dict = {}

                if use_snmp:
                    loop = asyncio.get_event_loop()
                    try:
                        driver = get_driver(olt)
                    except ValueError as drv_err:
                        logger.error(f"No driver for OLT {olt.name} (model={olt.model!r}): {drv_err}")
                        raise

                    logger.info(f"Polling {olt.name} via {driver.__class__.__name__}" + (" (skip optical)" if skip_optical else ""))
                    poll_result: DriverPollResult = await loop.run_in_executor(
                        thread_executor, lambda: driver.poll(skip_optical=skip_optical)
                    )

                    snmp_onus_data = list(poll_result.onus or [])
                    snmp_status_map = dict(poll_result.status_map or {})
                    web_opm_data = dict(poll_result.optical_data or {})
                    web_model_data = dict(poll_result.onu_models or {})
                    web_status_data = dict(poll_result.olt_alive_times or {})
                    health_data = dict(poll_result.health or {})

                    # Surface OPM errors the same way the legacy code path did,
                    # so the existing dashboard "last error" UI keeps working.
                    if web_opm_data:
                        if olt.last_error and 'OPM' in olt.last_error:
                            olt.last_error = None
                    elif not (web_user and web_pass):
                        olt.last_error = "OPM failed: Web credentials not configured"

                # Common processing for all OLT types (after getting ONU list)
                if snmp_onus_data:
                    logger.info(f"ONU poll successful for {olt.name}: {len(snmp_onus_data)} ONUs")

                    # Update OLT status
                    olt.is_online = True
                    olt.last_poll = get_current_time_in_timezone(db)
                    olt.last_error = None

                    # Health metrics from the driver poll above (CPU/temp/uptime + PON transceiver data)
                    try:
                        if health_data:
                            olt.cpu_usage = health_data.get('cpu_usage')
                            olt.memory_usage = health_data.get('memory_usage')
                            olt.temperature = health_data.get('temperature')
                            olt.uptime_seconds = health_data.get('uptime_seconds')

                            # Check high temperature alarm
                            if olt.temperature is not None:
                                alarm_settings = get_alarm_settings(db)
                                temp_threshold = float(alarm_settings.get("high_temperature_threshold", 60))
                                if olt.temperature > temp_threshold:
                                    _now = get_current_time_in_timezone(db).replace(tzinfo=None)
                                    _last = high_temp_alert_cache.get(olt.id)
                                    if _last is None or (_now - _last).total_seconds() >= HIGH_TEMP_COOLDOWN_S:
                                        send_high_temperature_notification(db, olt, olt.temperature)
                                        high_temp_alert_cache[olt.id] = _now
                                elif olt.id in high_temp_alert_cache:
                                    # Recovered below threshold — reset so the next
                                    # spike alerts immediately.
                                    del high_temp_alert_cache[olt.id]

                            # Save PON port transceiver diagnostics to OLTPort table
                            pon_ports_data = health_data.get('pon_ports', [])
                            for port_info in pon_ports_data:
                                port_num = port_info.get('port')
                                if port_num:
                                    # Find or create OLTPort entry for this PON port
                                    olt_port = db.query(OLTPort).filter(
                                        OLTPort.olt_id == olt.id,
                                        OLTPort.port_type == 'pon',
                                        OLTPort.port_number == port_num
                                    ).first()
                                    if not olt_port:
                                        olt_port = OLTPort(
                                            olt_id=olt.id,
                                            port_type='pon',
                                            port_number=port_num
                                        )
                                        db.add(olt_port)
                                    # Update transceiver data
                                    if port_info.get('temperature') is not None:
                                        olt_port.temperature = port_info['temperature']
                                    if port_info.get('tx_power') is not None:
                                        olt_port.tx_power = port_info['tx_power']
                                    olt_port.last_updated = datetime.utcnow()
                    except Exception as health_err:
                        logger.warning(f"Health poll failed for {olt.name}: {health_err}")

                    # Send OLT back online notification if it was offline
                    if not olt_was_online and olt.is_online:
                        send_olt_status_notification(db, olt, is_online=True)

                    # Re-index existing ONUs by (pon_port, onu_id) for proper matching
                    existing_by_key = {
                        (o.pon_port, o.onu_id): o
                        for o in db.query(ONU).filter(ONU.olt_id == olt.id).all()
                    }
                    # Also index by MAC so an ONU that re-registers at a new
                    # (pon,onu) position is MOVED, not duplicated. Prefer the
                    # online / most-recently-seen row when duplicates still exist.
                    existing_by_mac = {}
                    for _o in existing_by_key.values():
                        _cur = existing_by_mac.get(_o.mac_address)
                        if _cur is None or (_o.is_online and not _cur.is_online):
                            existing_by_mac[_o.mac_address] = _o

                    # Track which ONUs we've seen
                    seen_keys = set()

                    # Collect status changes for batched notification
                    onus_went_online = []
                    onus_went_offline = []
                    new_onus = []
                    # First-ever poll of this OLT: every ONU looks "new" but this
                    # is just initial discovery — don't fire a per-ONU storm.
                    is_first_discovery = len(existing_by_key) == 0

                    # Check license ONU limit
                    from license_manager import license_manager
                    license_info = license_manager.get_license_info()
                    max_onus = license_info.get('max_onus', 100)
                    current_onu_count = db.query(ONU).count()
                    onu_limit_reached = current_onu_count >= max_onus

                    # Update or create ONUs from SNMP data
                    for onu_data in snmp_onus_data:
                        key = (onu_data.pon_port, onu_data.onu_id)
                        seen_keys.add(key)

                        # Get online status from status_map using pon:onu key
                        # (handles duplicate MACs across PON ports correctly)
                        status_key = f"{onu_data.pon_port}:{onu_data.onu_id}"
                        is_online = snmp_status_map.get(status_key, False)

                        # Get RX power from SNMP
                        rx_power = onu_data.rx_power

                        # Get ONU self-reported optical data from web scraping
                        onu_rx_power = None
                        onu_tx_power = None
                        onu_temperature = None
                        onu_voltage = None
                        onu_tx_bias = None
                        web_distance = None
                        # Try MAC lookup first, then fallback to pon:onu key (for GPON OLTs)
                        web_data = None
                        if onu_data.mac_address in web_opm_data:
                            web_data = web_opm_data[onu_data.mac_address]
                        else:
                            # GPON OLTs may use pon:onu key format
                            pon_onu_key = f"{onu_data.pon_port}:{onu_data.onu_id}"
                            if pon_onu_key in web_opm_data:
                                web_data = web_opm_data[pon_onu_key]
                        if web_data:
                            # ONU self-reported RX power (what ONU sees from OLT, ~-19dBm)
                            onu_rx_power = web_data.get('onu_rx_power')
                            # OLT RX from ONU via web (fallback if SNMP rx_power not available)
                            web_olt_rx = web_data.get('rx_power')
                            if rx_power is None and web_olt_rx is not None:
                                rx_power = web_olt_rx  # Use web-scraped OLT RX as fallback
                            onu_tx_power = web_data.get('tx_power')
                            onu_temperature = web_data.get('temperature')
                            onu_voltage = web_data.get('voltage')
                            onu_tx_bias = web_data.get('tx_bias')
                            web_distance = web_data.get('distance')  # From OLT web interface

                        # Use ONLY web distance (no SNMP fallback)
                        final_distance = web_distance

                        existing = existing_by_key.get(key)
                        if existing is None:
                            # Not at this (pon,onu) — the same MAC may already exist
                            # at another position. Some OLTs list the SAME MAC at
                            # several (pon,onu) at once (one real + stale "ghost"
                            # entries), so ONLY move the row when this position is
                            # genuinely ONLINE (a real re-registration). If the MAC
                            # is already tracked and this position is offline, it's
                            # a ghost — skip it entirely (don't move, don't dupe).
                            moved = existing_by_mac.get(onu_data.mac_address)
                            if moved is not None:
                                if is_online:
                                    old_key = (moved.pon_port, moved.onu_id)
                                    seen_keys.discard(old_key)
                                    existing_by_key.pop(old_key, None)
                                    moved.pon_port = onu_data.pon_port
                                    moved.onu_id = onu_data.onu_id
                                    existing_by_key[key] = moved
                                    existing = moved
                                    logger.info(
                                        f"ONU {onu_data.mac_address} moved "
                                        f"{old_key} -> {key} (re-registered) on {olt.name}"
                                    )
                                else:
                                    # Ghost/stale offline position for a known MAC.
                                    continue

                        if existing is not None:
                            # Update existing ONU
                            was_online = existing.is_online
                            existing.mac_address = onu_data.mac_address
                            existing.is_online = is_online
                            # Update optical diagnostics only when online
                            if onu_data.description:
                                existing.description = onu_data.description
                            # Update model from SNMP or web scraping (for GPON OLTs)
                            if onu_data.model:
                                existing.model = onu_data.model
                            elif not existing.model and web_model_data:
                                # Fallback to web-scraped model for GPON OLTs
                                pon_onu_key = f"{onu_data.pon_port}:{onu_data.onu_id}"
                                if pon_onu_key in web_model_data:
                                    existing.model = web_model_data[pon_onu_key]
                            if is_online:
                                # Only update distance/rx_power when ONU is online
                                if final_distance is not None:
                                    existing.distance = final_distance
                                if rx_power is not None:
                                    existing.rx_power = rx_power
                                # Update ONU self-reported optical data from web scraping
                                if onu_rx_power is not None:
                                    existing.onu_rx_power = onu_rx_power
                                if onu_tx_power is not None:
                                    existing.onu_tx_power = onu_tx_power
                                if onu_temperature is not None:
                                    existing.onu_temperature = onu_temperature
                                if onu_voltage is not None:
                                    existing.onu_voltage = onu_voltage
                                if onu_tx_bias is not None:
                                    existing.onu_tx_bias = onu_tx_bias
                                # Update real alive time from OLT (scraped from onustatusinfo.html)
                                status_key = f"{onu_data.pon_port}:{onu_data.onu_id}"
                                if status_key in web_status_data:
                                    status_info = web_status_data[status_key]
                                    if status_info.get('alive_time_seconds') is not None:
                                        existing.olt_alive_time = status_info['alive_time_seconds']
                                existing.last_seen = datetime.utcnow()
                                # Initialize online_since for existing online ONUs (first poll after upgrade)
                                if existing.online_since is None:
                                    existing.online_since = datetime.utcnow()
                            else:
                                # Keep last known optical data when ONU goes offline
                                # This preserves the "last signal" for notifications and troubleshooting
                                # Only clear if explicitly requested (e.g., ONU removed)
                                pass
                            existing.updated_at = datetime.utcnow()

                            # Collect status changes for batched notification
                            if was_online and not is_online:
                                logger.info(f"ONU went OFFLINE: {existing.mac_address} (PON {existing.pon_port}/{existing.onu_id})")
                                # Get offline reason from onustatusinfo.html (primary) or alarm log (fallback)
                                status_key = f"{onu_data.pon_port}:{onu_data.onu_id}"
                                offline_reason = None
                                if status_key in web_status_data:
                                    # Use deregister_reason from onustatusinfo.html (more reliable)
                                    dereg_reason = web_status_data[status_key].get('deregister_reason')
                                    if dereg_reason:
                                        # Map OLT values to our standard values
                                        dereg_lower = dereg_reason.lower()
                                        if 'power' in dereg_lower or 'dying' in dereg_lower:
                                            offline_reason = "Power Off"
                                        elif 'los' in dereg_lower or 'fiber' in dereg_lower or 'link' in dereg_lower:
                                            offline_reason = "Fiber Cut"
                                        else:
                                            offline_reason = dereg_reason  # Use raw value
                                # Fallback to alarm log if no status data
                                if not offline_reason and web_user and web_pass:
                                    try:
                                        offline_reason = get_onu_offline_reason_web(
                                            olt.ip_address, existing.pon_port, existing.onu_id,
                                            existing.mac_address, web_user, web_pass
                                        )
                                    except Exception as e:
                                        logger.debug(f"Could not get offline reason from alarm log: {e}")
                                existing.offline_reason = offline_reason or "Unknown"
                                existing.olt_alive_time = None  # Clear alive time when offline
                                logger.info(f"ONU offline reason: {existing.offline_reason}")
                                onus_went_offline.append(existing)
                            elif not was_online and is_online:
                                logger.info(f"ONU came back ONLINE: {existing.mac_address} (PON {existing.pon_port}/{existing.onu_id})")
                                # Set online_since for uptime tracking
                                existing.online_since = datetime.utcnow()
                                existing.offline_reason = None  # Clear offline reason
                                onus_went_online.append(existing)
                        else:
                            # Check ONU limit before creating new ONU
                            if onu_limit_reached:
                                logger.warning(f"ONU limit reached ({max_onus}). Skipping new ONU: {onu_data.mac_address}")
                                # Log event so user can see in Event History
                                log_event(db, 'onu_limit_reached', 'system', 0, olt.id,
                                          f"ONU limit reached ({max_onus}). New ONU {onu_data.mac_address} not added. Upgrade license to add more ONUs.")
                                continue

                            # Create new ONU from SNMP (now includes pon_port and onu_id)
                            # Only save distance/rx_power if ONU is online
                            # Get model from SNMP or web scraping
                            onu_model = onu_data.model
                            if not onu_model and web_model_data:
                                pon_onu_key = f"{onu_data.pon_port}:{onu_data.onu_id}"
                                onu_model = web_model_data.get(pon_onu_key)

                            new_onu = ONU(
                                olt_id=olt.id,
                                pon_port=onu_data.pon_port,
                                onu_id=onu_data.onu_id,
                                mac_address=onu_data.mac_address,
                                description=onu_data.description,
                                model=onu_model,
                                is_online=is_online,
                                distance=final_distance if is_online else None,
                                rx_power=rx_power if is_online else None,
                                onu_rx_power=onu_rx_power if is_online else None,
                                onu_tx_power=onu_tx_power if is_online else None,
                                onu_temperature=onu_temperature if is_online else None,
                                onu_voltage=onu_voltage if is_online else None,
                                onu_tx_bias=onu_tx_bias if is_online else None,
                                online_since=datetime.utcnow() if is_online else None,
                                last_seen=datetime.utcnow() if is_online else None
                            )
                            db.add(new_onu)
                            existing_by_key[key] = new_onu
                            existing_by_mac[onu_data.mac_address] = new_onu
                            current_onu_count += 1  # Track added ONUs
                            onu_limit_reached = current_onu_count >= max_onus
                            # Collect for a bounded, off-loop notification after
                            # the ONU loop (was a blocking POST per ONU = storm).
                            if not is_first_discovery:
                                new_onus.append(new_onu)

                    # Mark ONUs not seen in SNMP as offline (they may be powered off)
                    for key, onu in existing_by_key.items():
                        if key not in seen_keys:
                            if onu.is_online:
                                logger.info(f"Marking ONU offline (not in SNMP): PON {onu.pon_port} ONU {onu.onu_id} ({onu.mac_address})")
                                onu.is_online = False
                                # Get offline reason from OLT alarm log
                                if web_user and web_pass:
                                    try:
                                        offline_reason = get_onu_offline_reason_web(
                                            olt.ip_address, onu.pon_port, onu.onu_id,
                                            onu.mac_address, web_user, web_pass
                                        )
                                        onu.offline_reason = offline_reason
                                        logger.info(f"ONU offline reason: {offline_reason}")
                                    except Exception as e:
                                        logger.debug(f"Could not get offline reason: {e}")
                                        onu.offline_reason = "Unknown"
                                else:
                                    onu.offline_reason = "Unknown"
                                # Keep last known optical data for notifications and troubleshooting
                                # (distance, rx_power, onu_rx_power, etc. are preserved)
                                onu.updated_at = datetime.utcnow()
                                onus_went_offline.append(onu)

                    # Send batched notification for all status changes
                    send_whatsapp_notification_batch(db, onus_went_online, onus_went_offline, olt.name)

                    # New-ONU notifications: send after the loop, and cap to avoid
                    # a storm on a mass (re)registration. Beyond the cap, log only.
                    if new_onus:
                        NEW_ONU_NOTIFY_CAP = 10
                        for _n in new_onus[:NEW_ONU_NOTIFY_CAP]:
                            try:
                                send_new_onu_notification(db, _n, olt.name)
                            except Exception as _e:
                                logger.warning(f"new-ONU notify failed: {_e}")
                        if len(new_onus) > NEW_ONU_NOTIFY_CAP:
                            logger.info(
                                f"{len(new_onus)} new ONUs on {olt.name}; notified first "
                                f"{NEW_ONU_NOTIFY_CAP}, suppressed the rest to avoid a storm"
                            )

                    # Check for weak signal alarms (Danger Zone detection)
                    alarm_settings = get_alarm_settings(db)
                    if is_alarm_enabled(alarm_settings, "weak_signal"):
                        # Upper threshold: signal weaker than this triggers alert (e.g., -25 dBm)
                        upper_threshold = float(alarm_settings.get("weak_signal_threshold", -25))
                        # Lower threshold: signal weaker than this means ONU is likely to disconnect (e.g., -30 dBm)
                        lower_threshold = float(alarm_settings.get("weak_signal_lower_threshold", -30))

                        # Find ONUs in the DANGER ZONE: signal is weak but not yet disconnected
                        # Signal must be: weaker than upper (rx_power < upper) AND stronger than lower (rx_power >= lower)
                        weak_signal_onus = [
                            o for o in db.query(ONU).filter(
                                ONU.olt_id == olt.id,
                                ONU.is_online == True,
                                ONU.rx_power != None,
                                ONU.rx_power < upper_threshold,  # Signal weaker than upper threshold
                                ONU.rx_power >= lower_threshold  # But not yet at disconnection level
                            ).all()
                        ]
                        if weak_signal_onus:
                            send_weak_signal_notification(db, weak_signal_onus, olt.name, upper_threshold, lower_threshold)

                    # Log successful SNMP poll
                    poll_log = PollLog(
                    olt_id=olt.id,
                    status="success",
                    message=f"SNMP: {len(snmp_onus_data)} ONUs",
                    onus_found=len(snmp_onus_data)
                    )
                    db.add(poll_log)
                    db.commit()

                    # Collect traffic history after successful poll
                    await collect_traffic_history(olt, db)
                    db.commit()

            except Exception as e:
                logger.error(f"Failed to poll OLT {olt.name}: {e}")
                olt.is_online = False
                olt.last_poll = get_current_time_in_timezone(db)
                olt.last_error = str(e)

                # Send OLT offline notification if it was online before
                if olt_was_online:
                    send_olt_status_notification(db, olt, is_online=False)

                # Log failed poll
                poll_log = PollLog(
                    olt_id=olt.id,
                    status="error",
                    message=str(e),
                    onus_found=0
                )
                db.add(poll_log)

            db.commit()

    finally:
        db.close()

    logger.info("Completed OLT polling cycle")


async def cleanup_old_data(db_session_factory, retention_days: int = 30):
    """Clean up old data from traffic_history, poll_logs, and port_traffic tables.

    This runs periodically to prevent database bloat.
    Default retention is 7 days.
    """
    from models import PortTraffic

    db = db_session_factory()
    try:
        cutoff_time = datetime.utcnow() - timedelta(days=retention_days)

        # Clean traffic_history (largest table)
        deleted_traffic = db.query(TrafficHistory).filter(
            TrafficHistory.timestamp < cutoff_time
        ).delete(synchronize_session=False)

        # Clean poll_logs (uses polled_at column)
        deleted_polls = db.query(PollLog).filter(
            PollLog.polled_at < cutoff_time
        ).delete(synchronize_session=False)

        # Clean port_traffic
        deleted_ports = db.query(PortTraffic).filter(
            PortTraffic.timestamp < cutoff_time
        ).delete(synchronize_session=False)

        # Self-heal: scrub any impossible traffic samples (negative or above the
        # 10 Gbps port ceiling) so one bad row can't wreck a graph's scale/avg.
        _bad = "(rx_kbps < 0 OR tx_kbps < 0 OR rx_kbps > 10000000 OR tx_kbps > 10000000)"
        try:
            from sqlalchemy import text as _sql_text
            db.execute(_sql_text(f"DELETE FROM port_traffic WHERE {_bad}"))
            db.execute(_sql_text(f"DELETE FROM traffic_history WHERE {_bad}"))
        except Exception as _e:
            logger.warning(f"Impossible-rate scrub skipped: {_e}")

        db.commit()

        total_deleted = deleted_traffic + deleted_polls + deleted_ports
        if total_deleted > 0:
            logger.info(f"Cleanup completed: deleted {deleted_traffic} traffic_history, "
                       f"{deleted_polls} poll_logs, {deleted_ports} port_traffic records "
                       f"older than {retention_days} days")

    except Exception as e:
        logger.error(f"Cleanup error: {e}")
        db.rollback()
    finally:
        db.close()


async def check_and_run_auto_backup(db_session_factory):
    """Check if auto backup is due and run it"""
    logger.info("Checking for scheduled auto backup...")
    db = db_session_factory()
    try:
        settings = db.query(BackupSettings).first()
        if not settings or not settings.auto_backup_enabled:
            logger.info("Auto backup disabled or no settings")
            return

        if not settings.next_backup_at:
            logger.info("No next_backup_at time set")
            return

        now = datetime.now()
        logger.info(f"Auto backup check: now={now}, scheduled={settings.next_backup_at}")
        if now >= settings.next_backup_at:
            logger.info("Auto backup triggered - running scheduled backup")

            # Run the backup
            success, backup_path, file_size = create_system_backup_file(db, include_uploads=False)

            if success:
                # Save backup record
                backup = SystemBackup(
                    filename=backup_path.name,
                    file_size=file_size if isinstance(file_size, int) else 0,
                    backup_type='scheduled',
                    storage_type='local',
                    storage_path=str(backup_path),
                    includes_db=True,
                    includes_config=True,
                    includes_uploads=False,
                    status='completed',
                    notes='Automatic scheduled backup'
                )
                db.add(backup)

                # Update last backup info
                settings.last_backup_at = now
                settings.last_backup_status = 'completed'

                logger.info(f"Auto backup completed: {backup_path.name}")
            else:
                settings.last_backup_status = 'failed'
                logger.error(f"Auto backup failed: {file_size}")

            # Calculate next backup time
            settings.next_backup_at = calculate_next_backup_time(settings)
            db.commit()

    except Exception as e:
        logger.error(f"Auto backup check error: {e}", exc_info=True)
    finally:
        db.close()


async def poll_all_tenants(db_session_factory, use_snmp: bool = True, skip_optical: bool = False):
    """Iterate every active tenant and run the polling cycle for each.

    Phase 1: this is the new outer driver. The single-tenant codepath is
    preserved by treating "no tenants table" as "fall through to the legacy
    one-shot poll".
    """
    # Open an unscoped session ONLY to enumerate tenants. Every nested
    # poll_all_olts call uses its own tenant-scoped session.
    db = db_session_factory()
    try:
        try:
            tenants = (
                db.query(Tenant)
                .filter(Tenant.status.in_(("active", "trial")))
                .filter(Tenant.deleted_at.is_(None))
                .all()
            )
        except Exception:
            # Tenants table missing — legacy single-tenant binary, fall back.
            await poll_all_olts(db_session_factory, use_snmp=use_snmp, tenant_id=None, skip_optical=skip_optical)
            return

        if not tenants:
            logger.info("No active tenants to poll")
            return

        logger.info(f"Polling {len(tenants)} active tenants")
        for tenant in tenants:
            try:
                await poll_all_olts(
                    db_session_factory, use_snmp=use_snmp, tenant_id=tenant.id, skip_optical=skip_optical
                )
            except Exception as e:
                logger.error(
                    f"Polling cycle failed for tenant {tenant.id} ({tenant.name}): {e}",
                    exc_info=True,
                )
    finally:
        db.close()


async def polling_loop(db_session_factory):
    """Background loop to poll OLTs periodically"""
    global cleanup_counter

    # Web scrape (optical data) runs every OPTICAL_EVERY cycles. With a 30s
    # poll interval and OPTICAL_EVERY=5, optical refreshes every ~2.5 min —
    # plenty fast for RX power / distance which change over hours.
    OPTICAL_EVERY = int(os.getenv("OPTICAL_EVERY", "5"))
    optical_counter = 0

    logger.info("Background polling loop started")
    while True:
        try:
            logger.info(f"Waiting {POLL_INTERVAL} seconds before next poll cycle...")
            await asyncio.sleep(POLL_INTERVAL)

            optical_counter += 1
            skip_optical = (optical_counter % OPTICAL_EVERY) != 0

            logger.info(f"Starting scheduled poll cycle (optical={'skip' if skip_optical else 'run'})")
            await poll_all_tenants(db_session_factory, skip_optical=skip_optical)

            # Check for auto backup
            await check_and_run_auto_backup(db_session_factory)

            # Run any due scheduled tasks (backups, ONU reboots, …)
            await run_due_scheduled_tasks(db_session_factory)

            # Run cleanup periodically (every CLEANUP_INTERVAL_CYCLES polls)
            cleanup_counter += 1
            if cleanup_counter >= CLEANUP_INTERVAL_CYCLES:
                cleanup_counter = 0
                logger.info("Running scheduled database cleanup...")
                # Keep 30 days so the 1-month (1M) traffic graph isn't clipped.
                await cleanup_old_data(db_session_factory, retention_days=30)

        except asyncio.CancelledError:
            logger.info("Polling loop cancelled")
            break
        except Exception as e:
            logger.error(f"Polling loop error: {e}", exc_info=True)


async def saas_fallback_polling_loop(db_session_factory):
    """SaaS-mode fallback: directly poll OLTs that the agent can't reach.

    When the agent can't reach an OLT (e.g. it's on a different subnet), it
    sets ``last_error = 'agent:unreachable'`` during ingest.  The SaaS may
    still be able to reach the OLT through the WireGuard tunnel.  This loop
    runs every 60s and polls any OLT with ``last_error = 'agent:unreachable'``.

    Uses a PostgreSQL advisory lock so only one uvicorn worker runs this at a
    time — prevents hammering OLTs with duplicate concurrent SNMP requests.
    """
    FALLBACK_INTERVAL = 60
    ADVISORY_LOCK_ID = 999_999_001  # unique lock ID for fallback polling
    OPTICAL_EVERY = 5  # collect optical data every Nth cycle (~5 min)

    logger.info("SaaS fallback polling loop started (every %ds)", FALLBACK_INTERVAL)
    await asyncio.sleep(15)

    cycle_count = OPTICAL_EVERY - 1  # first cycle collects optical immediately
    while True:
        try:
            await asyncio.sleep(FALLBACK_INTERVAL)

            db = db_session_factory()
            try:
                # Advisory lock — only one worker runs this block at a time.
                # pg_try_advisory_lock returns True immediately if lock is free.
                from sqlalchemy import text
                got_lock = db.execute(
                    text(f"SELECT pg_try_advisory_lock({ADVISORY_LOCK_ID})")
                ).scalar()
                if not got_lock:
                    continue  # Other worker is already running

                try:
                    cycle_count += 1
                    skip_optical = (cycle_count % OPTICAL_EVERY) != 0
                    logger.info(
                        "Fallback poll cycle #%d starting%s",
                        cycle_count,
                        "" if skip_optical else " (with optical)",
                    )

                    # Get active tenants (tenants table has no RLS)
                    tenants = db.execute(text(
                        "SELECT id, name FROM tenants "
                        "WHERE status IN ($$active$$,$$trial$$) AND deleted_at IS NULL"
                    )).fetchall()

                    loop = asyncio.get_event_loop()

                    for tenant_row in tenants:
                        tid = str(tenant_row[0])
                        tname = tenant_row[1]

                        # Open a tenant-scoped session to query OLTs (RLS)
                        tdb = db_session_factory()
                        set_session_tenant(tdb, tid)
                        try:
                            # Find OLTs the agent flagged as unreachable
                            # OR newly-added OLTs that have never been polled
                            fallback_olts = (
                                tdb.query(OLT)
                                .filter(or_(
                                    OLT.last_error == "agent:unreachable",
                                    OLT.last_error == "pending:first_poll",
                                    (OLT.last_poll.is_(None)) & (OLT.is_online == False),
                                ))
                                .all()
                            )
                            if not fallback_olts:
                                continue

                            logger.info(
                                "Fallback poll: %d OLT(s) need polling for tenant %s",
                                len(fallback_olts), tname,
                            )

                            for olt in fallback_olts:
                                wid = olt.workspace_id
                                if wid:
                                    tdb.info["workspace_id"] = str(wid)
                                try:
                                    driver = get_driver(olt)
                                    poll_result = await loop.run_in_executor(
                                        thread_executor,
                                        lambda d=driver, so=skip_optical: d.poll(skip_optical=so),
                                    )

                                    if not poll_result.onus:
                                        logger.info("Fallback: %s (%s) still unreachable", olt.name, olt.ip_address)
                                        continue

                                    logger.info(
                                        "Fallback: %s (%s) found %d ONUs — updating",
                                        olt.name, olt.ip_address, len(poll_result.onus),
                                    )
                                    olt.is_online = True
                                    olt.last_poll = datetime.utcnow()
                                    # Keep last_error as "agent:unreachable" so the
                                    # fallback loop continues polling this OLT on
                                    # subsequent cycles (no agent will claim it).
                                    olt.last_error = "agent:unreachable"

                                    health = poll_result.health or {}
                                    if health:
                                        olt.cpu_usage = health.get("cpu_usage")
                                        olt.memory_usage = health.get("memory_usage")
                                        olt.temperature = health.get("temperature")
                                        olt.uptime_seconds = health.get("uptime_seconds")

                                    existing_by_key = {
                                        (o.pon_port, o.onu_id): o
                                        for o in tdb.query(ONU).filter(ONU.olt_id == olt.id).all()
                                    }
                                    web_opm_data = dict(poll_result.optical_data or {})
                                    web_model_data = dict(poll_result.onu_models or {})
                                    web_status_data = dict(poll_result.olt_alive_times or {})
                                    now = datetime.utcnow()
                                    for onu_data in poll_result.onus:
                                        key = (onu_data.pon_port, onu_data.onu_id)
                                        status_key = f"{onu_data.pon_port}:{onu_data.onu_id}"
                                        is_online = poll_result.status_map.get(status_key, False)

                                        # Resolve optical data (same logic as main poll)
                                        rx_power = onu_data.rx_power
                                        onu_rx_power = onu_tx_power = onu_temperature = onu_voltage = onu_tx_bias = None
                                        web_distance = None
                                        web_data = web_opm_data.get(onu_data.mac_address)
                                        if not web_data:
                                            web_data = web_opm_data.get(status_key)
                                        if web_data:
                                            onu_rx_power = web_data.get('onu_rx_power')
                                            web_olt_rx = web_data.get('rx_power')
                                            if rx_power is None and web_olt_rx is not None:
                                                rx_power = web_olt_rx
                                            onu_tx_power = web_data.get('tx_power')
                                            onu_temperature = web_data.get('temperature')
                                            onu_voltage = web_data.get('voltage')
                                            onu_tx_bias = web_data.get('tx_bias')
                                            web_distance = web_data.get('distance')

                                        if key in existing_by_key:
                                            onu = existing_by_key[key]
                                            onu.is_online = is_online
                                            onu.mac_address = onu_data.mac_address
                                            if onu_data.description:
                                                onu.description = onu_data.description
                                            if onu_data.model:
                                                onu.model = onu_data.model
                                            elif not onu.model and web_model_data:
                                                onu.model = web_model_data.get(status_key)
                                            if is_online:
                                                if web_distance is not None:
                                                    onu.distance = web_distance
                                                if rx_power is not None:
                                                    onu.rx_power = rx_power
                                                if onu_rx_power is not None:
                                                    onu.onu_rx_power = onu_rx_power
                                                if onu_tx_power is not None:
                                                    onu.onu_tx_power = onu_tx_power
                                                if onu_temperature is not None:
                                                    onu.onu_temperature = onu_temperature
                                                if onu_voltage is not None:
                                                    onu.onu_voltage = onu_voltage
                                                if onu_tx_bias is not None:
                                                    onu.onu_tx_bias = onu_tx_bias
                                                if status_key in web_status_data:
                                                    at = web_status_data[status_key].get('alive_time_seconds')
                                                    if at is not None:
                                                        onu.olt_alive_time = at
                                                onu.last_seen = now
                                                if onu.online_since is None:
                                                    onu.online_since = now
                                            onu.updated_at = now
                                        else:
                                            onu = ONU(
                                                tenant_id=tid,
                                                workspace_id=str(wid) if wid else None,
                                                olt_id=olt.id,
                                                pon_port=onu_data.pon_port,
                                                onu_id=onu_data.onu_id,
                                                mac_address=onu_data.mac_address,
                                                description=onu_data.description or "",
                                                model=onu_data.model or web_model_data.get(status_key),
                                                is_online=is_online,
                                                distance=web_distance if is_online else None,
                                                rx_power=rx_power if is_online else None,
                                                onu_rx_power=onu_rx_power if is_online else None,
                                                onu_tx_power=onu_tx_power if is_online else None,
                                                onu_temperature=onu_temperature if is_online else None,
                                                onu_voltage=onu_voltage if is_online else None,
                                                onu_tx_bias=onu_tx_bias if is_online else None,
                                                online_since=now if is_online else None,
                                                last_seen=now if is_online else None,
                                            )
                                            tdb.add(onu)

                                    seen_keys = {
                                        (d.pon_port, d.onu_id) for d in poll_result.onus
                                    }
                                    for key, onu in existing_by_key.items():
                                        if key not in seen_keys and onu.is_online:
                                            onu.is_online = False
                                            onu.updated_at = now

                                    # ---- Traffic processing ----
                                    current_counters = poll_result.port_traffic or {}
                                    if current_counters:
                                        prev_snapshots = {
                                            s.mac_address: s
                                            for s in tdb.query(TrafficSnapshot).filter(
                                                TrafficSnapshot.olt_id == olt.id
                                            ).all()
                                        }
                                        traffic_data = []
                                        for tkey, counters in current_counters.items():
                                            rx_bytes = counters['rx_bytes']
                                            tx_bytes = counters['tx_bytes']
                                            pon_port = counters.get('pon_port', 0)
                                            onu_id_t = counters.get('onu_id', 0)

                                            # Resolve MAC
                                            if ':' in tkey and len(tkey) < 10:
                                                onu_for_mac = tdb.query(ONU).filter(
                                                    ONU.olt_id == olt.id,
                                                    ONU.pon_port == pon_port,
                                                    ONU.onu_id == onu_id_t,
                                                ).first()
                                                mac = onu_for_mac.mac_address if onu_for_mac else None
                                                if not mac:
                                                    continue
                                            else:
                                                mac = tkey

                                            rx_kbps = 0
                                            tx_kbps = 0
                                            if mac in prev_snapshots:
                                                prev = prev_snapshots[mac]
                                                time_diff = (now - prev.timestamp).total_seconds()
                                                if 0 < time_diff <= 300:
                                                    rx_diff = rx_bytes - prev.rx_bytes
                                                    tx_diff = tx_bytes - prev.tx_bytes
                                                    # Counter reset — discard sample
                                                    if rx_diff < 0 or tx_diff < 0:
                                                        rx_diff = tx_diff = 0
                                                    if rx_diff > 0 or tx_diff > 0:
                                                        rx_kbps = round((rx_diff * 8) / time_diff / 1000, 2)
                                                        tx_kbps = round((tx_diff * 8) / time_diff / 1000, 2)
                                                        MAX_VALID = 1_500_000
                                                        if rx_kbps > MAX_VALID or tx_kbps > MAX_VALID:
                                                            rx_kbps = tx_kbps = 0
                                                        prev.last_rx_kbps = rx_kbps
                                                        prev.last_tx_kbps = tx_kbps
                                                prev.rx_bytes = rx_bytes
                                                prev.tx_bytes = tx_bytes
                                                prev.timestamp = now
                                            else:
                                                tdb.add(TrafficSnapshot(
                                                    tenant_id=tid,
                                                    olt_id=olt.id,
                                                    mac_address=mac,
                                                    rx_bytes=rx_bytes,
                                                    tx_bytes=tx_bytes,
                                                    timestamp=now,
                                                    last_rx_kbps=0,
                                                    last_tx_kbps=0,
                                                ))

                                            onu_obj = tdb.query(ONU).filter(
                                                ONU.olt_id == olt.id,
                                                ONU.mac_address == mac,
                                            ).first()
                                            if onu_obj:
                                                traffic_data.append({
                                                    'onu': onu_obj,
                                                    'pon_port': onu_obj.pon_port,
                                                    'rx_kbps': rx_kbps,
                                                    'tx_kbps': tx_kbps,
                                                })

                                        # Save ONU traffic history
                                        for td in traffic_data:
                                            if td['rx_kbps'] > 0 or td['tx_kbps'] > 0:
                                                tdb.add(TrafficHistory(
                                                    tenant_id=tid,
                                                    entity_type='onu',
                                                    entity_id=str(td['onu'].id),
                                                    olt_id=olt.id,
                                                    pon_port=td['pon_port'],
                                                    onu_db_id=td['onu'].id,
                                                    rx_kbps=td['rx_kbps'],
                                                    tx_kbps=td['tx_kbps'],
                                                    timestamp=now,
                                                ))

                                        # PON aggregation
                                        pon_agg = {}
                                        for td in traffic_data:
                                            p = td['pon_port']
                                            if p not in pon_agg:
                                                pon_agg[p] = {'rx': 0, 'tx': 0}
                                            pon_agg[p]['rx'] += td['rx_kbps']
                                            pon_agg[p]['tx'] += td['tx_kbps']
                                        for p, agg in pon_agg.items():
                                            tdb.add(TrafficHistory(
                                                tenant_id=tid,
                                                entity_type='pon',
                                                entity_id=f"{olt.id}:{p}",
                                                olt_id=olt.id,
                                                pon_port=p,
                                                onu_db_id=None,
                                                rx_kbps=agg['rx'],
                                                tx_kbps=agg['tx'],
                                                timestamp=now,
                                            ))

                                        # OLT total
                                        total_rx = sum(t['rx_kbps'] for t in traffic_data)
                                        total_tx = sum(t['tx_kbps'] for t in traffic_data)
                                        if total_rx > 0 or total_tx > 0:
                                            tdb.add(TrafficHistory(
                                                tenant_id=tid,
                                                entity_type='olt',
                                                entity_id=str(olt.id),
                                                olt_id=olt.id,
                                                pon_port=None,
                                                onu_db_id=None,
                                                rx_kbps=total_rx,
                                                tx_kbps=total_tx,
                                                timestamp=now,
                                            ))

                                    # ---- Uplink port traffic ----
                                    try:
                                        snmp_comm = getattr(olt, 'snmp_community', None) or "public"
                                        uplink_counters = await loop.run_in_executor(
                                            thread_executor,
                                            poll_port_traffic_snmp,
                                            olt.ip_address,
                                            snmp_comm,
                                        )
                                        if uplink_counters:
                                            from models import PortTraffic
                                            up_rates = calculate_port_rates(
                                                olt.id, olt.ip_address, uplink_counters
                                            )
                                            try:
                                                pm = driver.get_port_layout().to_port_mapping()
                                            except ValueError:
                                                pm = {i: ('ge', i) for i in range(1, 9)}
                                            for if_idx, rates in up_rates.items():
                                                if if_idx in pm:
                                                    pt, pn = pm[if_idx]
                                                    tdb.add(PortTraffic(
                                                        tenant_id=tid,
                                                        olt_id=olt.id,
                                                        port_type=pt,
                                                        port_number=pn,
                                                        rx_kbps=rates['rx_kbps'],
                                                        tx_kbps=rates['tx_kbps'],
                                                        timestamp=now,
                                                    ))
                                                    tdb.add(TrafficHistory(
                                                        tenant_id=tid,
                                                        entity_type=pt,
                                                        entity_id=f"{olt.id}:{pt}:{pn}",
                                                        olt_id=olt.id,
                                                        pon_port=None,
                                                        onu_db_id=None,
                                                        rx_kbps=rates['rx_kbps'],
                                                        tx_kbps=rates['tx_kbps'],
                                                        timestamp=now,
                                                    ))
                                    except Exception as exc:
                                        logger.warning("Fallback uplink traffic failed for %s: %s", olt.name, exc)

                                    tdb.commit()

                                except Exception as exc:
                                    logger.warning(
                                        "Fallback poll failed for %s (%s): %s",
                                        olt.name, olt.ip_address, exc,
                                    )
                                    tdb.rollback()
                        finally:
                            tdb.close()
                finally:
                    # Always release advisory lock
                    db.execute(text(f"SELECT pg_advisory_unlock({ADVISORY_LOCK_ID})"))
                    db.commit()
            finally:
                db.close()

        except asyncio.CancelledError:
            logger.info("SaaS fallback polling loop cancelled")
            break
        except Exception as e:
            logger.error("Fallback polling loop error: %s", e, exc_info=True)


async def handle_trap_event(event: TrapEvent, db_session_factory):
    """Handle SNMP trap event - update ONU status and send notification"""
    from models import SessionLocal

    logger.info(f"Processing trap: {event.event_type} from {event.source_ip}, "
                f"PON:{event.pon_port}, ONU:{event.onu_id}, MAC:{event.mac_address}")

    db = db_session_factory()
    try:
        # Find the OLT by IP address
        olt = db.query(OLT).filter(OLT.ip_address == event.source_ip).first()
        if not olt:
            logger.warning(f"Trap from unknown OLT: {event.source_ip}")
            return

        # Find the ONU
        onu = None
        if event.mac_address:
            onu = db.query(ONU).filter(
                ONU.olt_id == olt.id,
                ONU.mac_address == event.mac_address
            ).first()
        elif event.pon_port and event.onu_id:
            onu = db.query(ONU).filter(
                ONU.olt_id == olt.id,
                ONU.pon_port == event.pon_port,
                ONU.onu_id == event.onu_id
            ).first()

        if not onu:
            logger.warning(f"Trap for unknown ONU: PON={event.pon_port}, ID={event.onu_id}, MAC={event.mac_address}")
            return

        # Check if status actually changed
        new_status = event.event_type == 'online'
        if onu.is_online == new_status:
            logger.debug(f"ONU {onu.description or onu.mac_address} status unchanged")
            return

        # Update ONU status
        old_status = onu.is_online
        onu.is_online = new_status
        onu.last_seen = datetime.utcnow()
        db.commit()

        logger.info(f"ONU {onu.description or onu.mac_address} status changed: "
                    f"{'online' if old_status else 'offline'} -> {'online' if new_status else 'offline'} (via TRAP)")

        # Send WhatsApp notification for instant alert
        if event.event_type == 'offline':
            send_whatsapp_notification_batch(db, [], [onu], olt.name)
        else:
            send_whatsapp_notification_batch(db, [onu], [], olt.name)

    except Exception as e:
        logger.error(f"Error handling trap event: {e}", exc_info=True)
    finally:
        db.close()


def get_trap_settings(db: Session) -> dict:
    """Get SNMP trap settings from database"""
    settings = {}
    for key in ['trap_enabled', 'trap_port', 'trap_community']:
        setting = db.query(Settings).filter(Settings.key == key).first()
        if setting:
            settings[key] = setting.value
    return settings


async def start_trap_receiver(db_session_factory):
    """Start SNMP trap receiver if enabled"""
    global trap_receiver, trap_task

    db = db_session_factory()
    try:
        settings = get_trap_settings(db)
        trap_enabled = str(settings.get('trap_enabled', 'true')).lower() == 'true'
        trap_port = int(settings.get('trap_port', '162'))

        if not trap_enabled:
            logger.info("SNMP Trap receiver is disabled in settings")
            return

        trap_receiver = SimpleTrapReceiver(port=trap_port)

        async def on_trap(event: TrapEvent):
            await handle_trap_event(event, db_session_factory)

        trap_receiver.set_callback(on_trap)

        try:
            await trap_receiver.start()
            logger.info(f"SNMP Trap receiver started on port {trap_port}")
        except PermissionError:
            logger.warning(f"Cannot bind to port {trap_port} (requires root). "
                          f"Trying port 1620 instead...")
            trap_receiver = SimpleTrapReceiver(port=1620)
            trap_receiver.set_callback(on_trap)
            await trap_receiver.start()
            logger.info("SNMP Trap receiver started on port 1620")

    except Exception as e:
        logger.error(f"Failed to start trap receiver: {e}")
    finally:
        db.close()


def dedupe_onus(db) -> int:
    """Collapse duplicate ONU rows that share (olt_id, mac_address).

    A MAC identifies exactly one physical ONU, but historically a re-registration
    at a new (pon,onu) created a new row and orphaned the old one (which then
    showed the online twin's MAC-keyed traffic). Keep the best row per (olt,mac)
    — online preferred, then most-recently-seen — and delete the rest with their
    now-bogus history. Idempotent: a no-op once the table is clean.
    """
    from sqlalchemy import func
    from datetime import datetime as _dt
    removed = 0
    dup_groups = (
        db.query(ONU.olt_id, ONU.mac_address)
        .group_by(ONU.olt_id, ONU.mac_address)
        .having(func.count() > 1)
        .all()
    )
    for olt_id, mac in dup_groups:
        rows = db.query(ONU).filter(ONU.olt_id == olt_id, ONU.mac_address == mac).all()
        rows.sort(
            key=lambda o: (
                1 if o.is_online else 0,
                o.last_seen or o.online_since or o.updated_at or _dt.min,
            ),
            reverse=True,
        )
        for loser in rows[1:]:
            db.query(TrafficHistory).filter(
                TrafficHistory.onu_db_id == loser.id
            ).delete(synchronize_session=False)
            db.delete(loser)
            removed += 1
    if removed:
        db.commit()
        logger.info(f"Deduped {removed} duplicate ONU rows (kept 1 per olt+MAC)")
    return removed


@asynccontextmanager
async def lifespan(app: FastAPI):
    """Application lifespan manager"""
    global polling_task, trap_task, trap_receiver, fallback_task

    # Validate license first (but don't crash - allow read-only mode)
    from license_manager import validate_license_on_startup, license_manager, LicenseError, license_check_loop
    license_valid = validate_license_on_startup()

    if license_valid:
        license_info = license_manager.get_license_info()
        logger.info(f"License valid: {license_info.get('customer_name')} | Max OLTs: {license_info.get('max_olts')} | Max ONUs: {license_info.get('max_onus')}")
    else:
        logger.warning("=" * 50)
        logger.warning("LICENSE INVALID - Running in READ-ONLY mode")
        logger.warning(f"Reason: {license_manager.error_message}")
        logger.warning("=" * 50)

    # Initialize database
    init_db()
    logger.info("Database initialized")

    # Create default admin user if needed (single-tenant legacy mode only).
    # In SaaS mode (Phase 1+) tenants are created via the signup flow and a
    # default tenant_id-less admin would violate row-level security.
    from models import SessionLocal
    if not os.getenv("SAAS_MODE"):
        db = SessionLocal()
        try:
            create_default_admin(db)
            # One-time (idempotent) cleanup of legacy duplicate ONU rows.
            try:
                dedupe_onus(db)
            except Exception as _dedup_err:
                logger.warning(f"ONU dedupe skipped: {_dedup_err}")
        finally:
            db.close()
    else:
        logger.info("SAAS_MODE=1 set; skipping legacy default-admin bootstrap")

    # Start background polling (disabled in SaaS mode — agents push data)
    if os.getenv("SAAS_MODE"):
        logger.info("SAAS_MODE — primary polling disabled, agents push data via /api/agent/ingest")
        # Start fallback polling loop for OLTs that the agent can't reach
        # but the SaaS can reach through WireGuard tunnel.
        fallback_task = asyncio.create_task(saas_fallback_polling_loop(SessionLocal))
        logger.info("Started SaaS fallback polling loop for agent-unreachable OLTs")
    else:
        polling_task = asyncio.create_task(polling_loop(SessionLocal))
        logger.info(f"Started background polling (interval: {POLL_INTERVAL}s)")

    # Start SNMP trap receiver (not needed in SaaS mode)
    if not os.getenv("SAAS_MODE"):
        trap_task = asyncio.create_task(start_trap_receiver(SessionLocal))

    # Start periodic license check
    license_task = asyncio.create_task(license_check_loop())
    logger.info("Started periodic license check (every 5 minutes)")

    yield

    # Cleanup
    if polling_task:
        polling_task.cancel()
        try:
            await polling_task
        except asyncio.CancelledError:
            pass

    if fallback_task:
        fallback_task.cancel()
        try:
            await fallback_task
        except asyncio.CancelledError:
            pass

    if trap_receiver:
        await trap_receiver.stop()

    logger.info("Application shutdown complete")


# Phase 5 — Sentry must be initialised before FastAPI is constructed so the
# integration can hook into request handling. No-op if SENTRY_DSN isn't set.
try:
    from observability import init_sentry, setup_observability
    init_sentry()
except Exception as _e:
    logger.warning(f"observability init failed: {_e}")
    setup_observability = None  # type: ignore[assignment]

# Create FastAPI app
app = FastAPI(
    title="EPON OLT Manager",
    description="Dashboard for managing VSOL EPON OLTs and ONUs",
    version="1.0.0",
    lifespan=lifespan
)

# Phase 5 — wire /health, /metrics and request-timing middleware.
if setup_observability is not None:
    try:
        setup_observability(app)
    except Exception as _e:
        logger.warning(f"observability setup failed: {_e}")

# CORS middleware - configurable origins from environment
# For production: set CORS_ORIGINS environment variable (comma-separated)
# Example: CORS_ORIGINS=https://olt.example.com,https://admin.example.com
CORS_ORIGINS = os.getenv("CORS_ORIGINS", "").strip()
if CORS_ORIGINS:
    allowed_origins = [origin.strip() for origin in CORS_ORIGINS.split(",")]
else:
    # Default: allow local development and same-origin requests
    allowed_origins = [
        "http://localhost:3000",
        "http://localhost:8000",
        "http://127.0.0.1:3000",
        "http://127.0.0.1:8000",
    ]

app.add_middleware(
    CORSMiddleware,
    allow_origins=allowed_origins,
    allow_credentials=True,
    allow_methods=["GET", "POST", "PUT", "DELETE", "OPTIONS"],
    allow_headers=["Authorization", "Content-Type", "X-Requested-With"],
)

# Phase 2 routers — auth (signup/login/reset) and Stripe billing.
# Imported lazily so a missing optional dep (like `stripe`) doesn't crash
# boot in dev environments where billing isn't configured yet.
try:
    from auth_routes import router as auth_router
    app.include_router(auth_router)
except Exception as _e:
    logger.warning(f"auth_routes not loaded: {_e}")

try:
    from billing import router as billing_router
    app.include_router(billing_router)
except Exception as _e:
    logger.warning(f"billing routes not loaded: {_e}")

# Phase 3 — WireGuard provisioning routes.
try:
    from wireguard.routes import router as wg_router
    app.include_router(wg_router)
except Exception as _e:
    logger.warning(f"wireguard routes not loaded: {_e}")

# Phase 6 — in-app feedback router (used by the dashboard widget).
try:
    from feedback_routes import router as feedback_router
    app.include_router(feedback_router)
except Exception as _e:
    logger.warning(f"feedback routes not loaded: {_e}")

# Local agent routes (key management + data ingest).
try:
    from agent_routes import router as agent_router
    app.include_router(agent_router)
except Exception as _e:
    logger.warning(f"agent routes not loaded: {_e}")

# Create uploads directory
UPLOAD_DIR = os.path.join(os.path.dirname(__file__), "uploads")
os.makedirs(UPLOAD_DIR, exist_ok=True)

# Mount static files for uploads
app.mount("/uploads", StaticFiles(directory=UPLOAD_DIR), name="uploads")

# Mount frontend static files (for compiled installations without nginx)
# Check multiple possible locations for static directory
def find_static_dir():
    possible_paths = [
        os.path.join(os.path.dirname(__file__), "static"),  # Development
        os.path.join(os.getcwd(), "static"),                 # Current working directory
        os.path.join(os.path.dirname(sys.executable), "static"),  # Next to executable
        "/opt/olt-manager/static",                           # Standard install location
    ]
    for path in possible_paths:
        if os.path.exists(path) and os.path.isdir(path):
            return path
    return None

STATIC_DIR = find_static_dir()
if STATIC_DIR:
    logger.info(f"Found static directory at: {STATIC_DIR}")
    # Mount the JS/CSS static folder
    static_assets = os.path.join(STATIC_DIR, "static")
    if os.path.exists(static_assets):
        app.mount("/static", StaticFiles(directory=static_assets), name="static_assets")


# ============ Helper Functions ============

def get_user_allowed_olt_ids(user: User, db: Session) -> Optional[List[int]]:
    """Get list of OLT IDs a user can access. Returns None if user has access to all.

    `owner` is the SaaS top-level role (created at signup) and acts as a
    superset of `admin` — both can see every OLT in their tenant. RLS
    handles the cross-tenant boundary automatically, so returning None
    here is safe: PostgreSQL only ever returns rows tagged with the
    current session's tenant_id.
    """
    if user.role in ("admin", "owner"):
        return None  # Admin/owner has access to all OLTs in the tenant

    # Get assigned OLT IDs for this operator
    assigned_ids = db.query(user_olts.c.olt_id).filter(
        user_olts.c.user_id == user.id
    ).all()

    return [row[0] for row in assigned_ids]


def get_user_olt_ids_list(user: User, db: Session) -> List[int]:
    """Get assigned OLT IDs for a user (for response serialization)."""
    assigned_ids = db.query(user_olts.c.olt_id).filter(
        user_olts.c.user_id == user.id
    ).all()
    return [row[0] for row in assigned_ids]


# ============ Dashboard Endpoints ============

@app.get("/api/dashboard", response_model=DashboardStats)
def get_dashboard_stats(user: User = Depends(require_auth), db: Session = Depends(get_db)):
    """Get dashboard statistics (filtered by user access)"""
    allowed_olt_ids = get_user_allowed_olt_ids(user, db)

    # Build OLT query
    olt_query = db.query(OLT)
    if allowed_olt_ids is not None:
        olt_query = olt_query.filter(OLT.id.in_(allowed_olt_ids))

    total_olts = olt_query.count()
    online_olts = olt_query.filter(OLT.is_online == True).count()

    # Build ONU query
    onu_query = db.query(ONU)
    if allowed_olt_ids is not None:
        onu_query = onu_query.filter(ONU.olt_id.in_(allowed_olt_ids))

    total_onus = onu_query.count()
    online_onus = onu_query.filter(ONU.is_online == True).count()

    return DashboardStats(
        total_olts=total_olts,
        online_olts=online_olts,
        offline_olts=total_olts - online_olts,
        total_onus=total_onus,
        online_onus=online_onus,
        offline_onus=total_onus - online_onus
    )


# ============ OLT Endpoints ============

@app.get("/api/olts", response_model=OLTListResponse)
def list_olts(user: User = Depends(require_auth), db: Session = Depends(get_db)):
    """List all OLTs (filtered by user access)"""
    allowed_olt_ids = get_user_allowed_olt_ids(user, db)

    query = db.query(OLT)
    if allowed_olt_ids is not None:
        query = query.filter(OLT.id.in_(allowed_olt_ids))

    olts = query.all()

    # Get ONU counts in a single query (fixes N+1 problem)
    olt_ids = [olt.id for olt in olts]
    counts_map = {}
    if olt_ids:
        # Get total counts per OLT
        total_counts = db.query(
            ONU.olt_id,
            func.count(ONU.id).label('total')
        ).filter(ONU.olt_id.in_(olt_ids)).group_by(ONU.olt_id).all()

        # Get online counts per OLT
        online_counts = db.query(
            ONU.olt_id,
            func.count(ONU.id).label('online')
        ).filter(ONU.olt_id.in_(olt_ids), ONU.is_online == True).group_by(ONU.olt_id).all()

        # Build lookup dictionary
        for row in total_counts:
            counts_map[row.olt_id] = {'total': row.total, 'online': 0}
        for row in online_counts:
            if row.olt_id in counts_map:
                counts_map[row.olt_id]['online'] = row.online

    response_olts = []
    for olt in olts:
        counts = counts_map.get(olt.id, {'total': 0, 'online': 0})

        response_olts.append(OLTResponse(
            id=olt.id,
            name=olt.name,
            ip_address=olt.ip_address,
            model=olt.model,
            pon_ports=olt.pon_ports,
            snmp_community=olt.snmp_community,
            is_online=olt.is_online,
            last_poll=olt.last_poll,
            last_error=_display_olt_error(olt.last_error),
            onu_count=counts['total'],
            online_onu_count=counts['online'],
            cpu_usage=olt.cpu_usage,
            memory_usage=olt.memory_usage,
            temperature=olt.temperature,
            uptime_seconds=olt.uptime_seconds,
            created_at=olt.created_at,
            updated_at=olt.updated_at
        ))

    return OLTListResponse(olts=response_olts, total=len(response_olts))


@app.get("/api/olts/{olt_id}", response_model=OLTResponse)
def get_olt(olt_id: int, user: User = Depends(require_auth), db: Session = Depends(get_db)):
    """Get specific OLT by ID"""
    olt = db.query(OLT).filter(OLT.id == olt_id).first()
    if not olt:
        raise HTTPException(status_code=404, detail="OLT not found")

    onu_count = db.query(ONU).filter(ONU.olt_id == olt.id).count()
    online_onu_count = db.query(ONU).filter(
        ONU.olt_id == olt.id,
        ONU.is_online == True
    ).count()

    return OLTResponse(
        id=olt.id,
        name=olt.name,
        ip_address=olt.ip_address,
        model=olt.model,
        pon_ports=olt.pon_ports,
        snmp_community=olt.snmp_community,
        is_online=olt.is_online,
        last_poll=olt.last_poll,
        last_error=_display_olt_error(olt.last_error),
        onu_count=onu_count,
        online_onu_count=online_onu_count,
        cpu_usage=olt.cpu_usage,
        memory_usage=olt.memory_usage,
        temperature=olt.temperature,
        uptime_seconds=olt.uptime_seconds,
        created_at=olt.created_at,
        updated_at=olt.updated_at
    )


@app.get("/api/supported-olt-models")
def get_supported_olt_models(user: User = Depends(require_auth)):
    """Return the OLT models the backend actually has drivers for.

    Single source of truth for the 'Add OLT' dropdown — replaces the hardcoded
    frontend list so the UI can't offer models with no driver. Each entry has an
    ``implemented`` flag (False = registered stub, e.g. Huawei/ZTE).
    """
    return {"models": list_supported_models()}


def _display_olt_error(last_error):
    """Translate internal status sentinels into human-readable text for the UI.

    Previously these sentinels were blanked to None, hiding the reason an OLT
    showed no data from first-time users. Now we surface a friendly message.
    """
    if last_error == "agent:unreachable":
        return "Local agent can't reach this OLT (check agent/network)"
    if last_error == "pending:first_poll":
        return "Awaiting first poll…"
    return last_error


def _probe_olt_snmp(ip: str, community: str, timeout: int = 4) -> bool:
    """Quick reachability check: SNMP GET of sysUpTime.

    Returns True if the OLT answers SNMP with the given community. Used to give
    an immediate, non-blocking warning at add-time instead of a silent failure
    on the first poll cycle. Never raises.
    """
    import subprocess
    try:
        result = subprocess.run(
            ["snmpget", "-v2c", "-c", community, "-t", "2", "-r", "1",
             f"{ip}:161", "1.3.6.1.2.1.1.3.0"],
            capture_output=True, text=True, timeout=timeout
        )
        return result.returncode == 0 and bool(result.stdout.strip())
    except FileNotFoundError:
        # snmpget not installed — can't probe; don't warn misleadingly
        logger.warning("snmpget not found; skipping add-OLT SNMP probe")
        return True
    except Exception:
        return False


@app.post("/api/olts", response_model=OLTResponse, status_code=201)
def create_olt(olt_data: OLTCreate, user: User = Depends(require_admin), db: Session = Depends(get_db)):
    """Create new OLT (admin only)"""
    if os.getenv("SAAS_MODE"):
        # SaaS: enforce the tenant's billing-plan OLT limit (402 if exceeded).
        from plans import enforce_plan_limit
        tenant = db.query(Tenant).filter(Tenant.id == user.tenant_id).first()
        if tenant:
            enforce_plan_limit(db, tenant, "olts")
    else:
        # Local (single-instance license) OLT limit.
        from license_manager import license_manager
        license_info = license_manager.get_license_info()
        max_olts = license_info.get('max_olts', 1)
        current_olt_count = db.query(OLT).count()
        if current_olt_count >= max_olts:
            package = license_info.get('package_type', 'trial')
            raise HTTPException(
                status_code=403,
                detail=f"OLT limit reached ({max_olts}). Upgrade your package to add more OLTs. Current package: {package}"
            )

    # Check for duplicate IP
    existing = db.query(OLT).filter(OLT.ip_address == olt_data.ip_address).first()
    if existing:
        raise HTTPException(status_code=400, detail="OLT with this IP already exists")

    olt = OLT(
        name=olt_data.name,
        ip_address=olt_data.ip_address,
        username=olt_data.username,
        password=encrypt_sensitive(olt_data.password),  # Encrypt password
        model=olt_data.model,
        pon_ports=olt_data.pon_ports,
        snmp_community=(olt_data.snmp_community or "public")
    )
    db.add(olt)
    db.commit()
    db.refresh(olt)

    # Log event
    log_event(db, 'olt_created', 'olt', olt.id, olt.id,
              f"OLT '{olt.name}' ({olt.ip_address}) added by {user.username}")

    # SaaS: auto-route OLT's /24 through WireGuard so fallback polling works
    if os.getenv("SAAS_MODE") and olt.workspace_id:
        try:
            from wireguard.routes import recalculate_workspace_routes
            ws = db.query(Workspace).filter(Workspace.id == olt.workspace_id).first()
            if ws and ws.wg_pubkey:
                recalculate_workspace_routes(ws, db)
                # Mark OLT for fallback pickup
                olt.last_error = "pending:first_poll"
                db.commit()
                db.refresh(olt)
        except Exception as e:
            logger.warning("Auto-route after OLT create failed: %s", e)

    # Non-blocking diagnostics so users learn about problems at add-time instead
    # of via a silent never-populating poll.
    warnings = []

    # (a) Model support: an unknown or stub (unimplemented) model will never poll.
    support = check_model_support(olt.model)
    if olt.model and support["status"] == "unknown":
        warnings.append(
            f"Model '{olt.model}' has no backend driver — this OLT will not poll. "
            f"Pick a supported model (see /api/supported-olt-models)."
        )
    elif support["status"] == "unimplemented":
        warnings.append(
            f"Model '{olt.model}' is registered but not yet implemented — polling "
            f"and actions are not available for it yet."
        )

    # (b) Connectivity probe (direct-poll deployments only). In SaaS mode the
    # backend can't reach the OLT directly — the agent does — so skip.
    if not os.getenv("SAAS_MODE"):
        if not _probe_olt_snmp(olt.ip_address, olt.snmp_community or "public"):
            warnings.append(
                f"OLT did not answer SNMP at {olt.ip_address} with community "
                f"'{olt.snmp_community or 'public'}'. Check reachability, that SNMP "
                f"v2c is enabled on the OLT, and the community string."
            )

    warning = " ".join(warnings) if warnings else None
    if warning:
        logger.warning("Add-OLT diagnostics for %s: %s", olt.ip_address, warning)

    return OLTResponse(
        id=olt.id,
        name=olt.name,
        ip_address=olt.ip_address,
        model=olt.model,
        pon_ports=olt.pon_ports,
        snmp_community=olt.snmp_community,
        is_online=olt.is_online,
        last_poll=olt.last_poll,
        last_error=_display_olt_error(olt.last_error),
        warning=warning,
        onu_count=0,
        online_onu_count=0,
        created_at=olt.created_at,
        updated_at=olt.updated_at
    )


@app.put("/api/olts/{olt_id}", response_model=OLTResponse)
def update_olt(olt_id: int, olt_data: OLTUpdate, user: User = Depends(require_admin), db: Session = Depends(get_db)):
    """Update OLT (admin only)"""
    olt = db.query(OLT).filter(OLT.id == olt_id).first()
    if not olt:
        raise HTTPException(status_code=404, detail="OLT not found")

    if olt_data.name is not None:
        olt.name = olt_data.name
    if olt_data.username is not None:
        olt.username = olt_data.username
    if olt_data.password is not None:
        olt.password = encrypt_sensitive(olt_data.password)  # Encrypt password
    if olt_data.model is not None:
        olt.model = olt_data.model
    if olt_data.pon_ports is not None:
        olt.pon_ports = olt_data.pon_ports
    if olt_data.snmp_community is not None:
        olt.snmp_community = olt_data.snmp_community

    db.commit()
    db.refresh(olt)

    onu_count = db.query(ONU).filter(ONU.olt_id == olt.id).count()
    online_onu_count = db.query(ONU).filter(
        ONU.olt_id == olt.id,
        ONU.is_online == True
    ).count()

    return OLTResponse(
        id=olt.id,
        name=olt.name,
        ip_address=olt.ip_address,
        model=olt.model,
        pon_ports=olt.pon_ports,
        snmp_community=olt.snmp_community,
        is_online=olt.is_online,
        last_poll=olt.last_poll,
        last_error=_display_olt_error(olt.last_error),
        onu_count=onu_count,
        online_onu_count=online_onu_count,
        created_at=olt.created_at,
        updated_at=olt.updated_at
    )


@app.delete("/api/olts/{olt_id}", status_code=204)
def delete_olt(olt_id: int, user: User = Depends(require_admin), db: Session = Depends(get_db)):
    """Delete OLT and all its ONUs (admin only)"""
    olt = db.query(OLT).filter(OLT.id == olt_id).first()
    if not olt:
        raise HTTPException(status_code=404, detail="OLT not found")

    # Store info before delete for logging
    olt_name = olt.name
    olt_ip = olt.ip_address
    olt_workspace_id = olt.workspace_id
    onu_count = db.query(ONU).filter(ONU.olt_id == olt_id).count()

    db.delete(olt)
    db.commit()

    # Log event
    log_event(db, 'olt_deleted', 'olt', olt_id, None,
              f"OLT '{olt_name}' ({olt_ip}) deleted by {user.username}" +
              (f" - {onu_count} ONUs also removed" if onu_count > 0 else ""))

    # SaaS: recalculate routes to remove stale subnets
    if os.getenv("SAAS_MODE") and olt_workspace_id:
        try:
            from wireguard.routes import recalculate_workspace_routes
            ws = db.query(Workspace).filter(Workspace.id == olt_workspace_id).first()
            if ws and ws.wg_pubkey:
                recalculate_workspace_routes(ws, db)
        except Exception as e:
            logger.warning("Auto-route after OLT delete failed: %s", e)

    return None


@app.post("/api/olts/{olt_id}/poll", response_model=PollResult)
async def poll_single_olt(olt_id: int, user: User = Depends(require_auth), db: Session = Depends(get_db)):
    """Manually trigger poll for specific OLT using SNMP (fast ~2 seconds)"""
    from license_manager import license_manager
    olt = db.query(OLT).filter(OLT.id == olt_id).first()
    if not olt:
        raise HTTPException(status_code=404, detail="OLT not found")

    try:
        # Use SNMP for fast polling
        loop = asyncio.get_event_loop()
        olt_community = olt.snmp_community or "public"
        onus_data, status_map = await loop.run_in_executor(
            thread_executor,
            poll_olt_snmp,
            olt.ip_address,
            olt_community
        )

        # Silent-failure guard: an empty SNMP result may mean the OLT is
        # unreachable / wrong community, NOT "0 ONUs". Probe before declaring it
        # online, so first-time users get an actionable error instead of a
        # green "online, 0 ONUs". (Direct-poll deployments only.)
        if not onus_data and not status_map and not os.getenv("SAAS_MODE"):
            if not _probe_olt_snmp(olt.ip_address, olt_community):
                olt.is_online = False
                olt.last_poll = get_current_time_in_timezone(db)
                olt.last_error = (
                    f"No SNMP response from {olt.ip_address} (community "
                    f"'{olt_community}'). Check reachability, SNMP v2c enabled, "
                    f"and the community string."
                )
                db.commit()
                return {"success": False, "message": olt.last_error, "onu_count": 0}

        # Update database (same logic as polling loop)
        olt.is_online = True
        olt.last_poll = get_current_time_in_timezone(db)
        olt.last_error = None

        existing_onus = {
            (o.pon_port, o.onu_id): o
            for o in db.query(ONU).filter(ONU.olt_id == olt.id).all()
        }

        seen_keys = set()

        # Collect status changes for batched notification
        onus_went_online = []
        onus_went_offline = []

        # Check license ONU limit for manual poll
        license_info = license_manager.get_license_info()
        max_onus = license_info.get('max_onus', 100)
        current_onu_count = db.query(ONU).count()
        onu_limit_reached = current_onu_count >= max_onus

        for onu_data in onus_data:
            key = (onu_data.pon_port, onu_data.onu_id)
            seen_keys.add(key)

            # Get online status from status_map using pon:onu key
            # (handles duplicate MACs across PON ports correctly)
            status_key = f"{onu_data.pon_port}:{onu_data.onu_id}"
            is_online = status_map.get(status_key, False)

            # Get RX power from SNMP
            rx_power = onu_data.rx_power

            if key in existing_onus:
                existing = existing_onus[key]
                was_online = existing.is_online
                existing.mac_address = onu_data.mac_address
                existing.is_online = is_online
                existing.missing_polls = 0  # Reset counter - ONU found in SNMP
                # Update description
                if onu_data.description:
                    existing.description = onu_data.description
                # Always update model if available from SNMP
                if onu_data.model:
                    existing.model = onu_data.model
                # Update optical diagnostics only when online
                if is_online:
                    if onu_data.distance is not None:
                        existing.distance = onu_data.distance
                    if rx_power is not None:
                        existing.rx_power = rx_power
                    existing.last_seen = datetime.utcnow()
                else:
                    # Clear optical data when ONU is offline (no live traffic)
                    existing.distance = None
                    existing.rx_power = None
                existing.updated_at = datetime.utcnow()

                # Collect status changes for batched notification
                if was_online and not is_online:
                    onus_went_offline.append(existing)
                elif not was_online and is_online:
                    onus_went_online.append(existing)
            else:
                # Check ONU limit before creating new ONU
                if onu_limit_reached:
                    logger.warning(f"[Manual Poll] ONU limit reached ({max_onus}). Skipping new ONU: {onu_data.mac_address}")
                    # Log event so user can see in Event History
                    log_event(db, 'onu_limit_reached', 'system', 0, olt.id,
                              f"ONU limit reached ({max_onus}). New ONU {onu_data.mac_address} not added. Upgrade license to add more ONUs.")
                    continue

                new_onu = ONU(
                    olt_id=olt.id,
                    pon_port=onu_data.pon_port,
                    onu_id=onu_data.onu_id,
                    mac_address=onu_data.mac_address,
                    description=onu_data.description,
                    model=onu_data.model,
                    is_online=is_online,
                    distance=onu_data.distance,
                    rx_power=rx_power,
                    last_seen=datetime.utcnow()
                )
                db.add(new_onu)
                current_onu_count += 1
                onu_limit_reached = current_onu_count >= max_onus

        # Track ONUs not found in SNMP - delete after 3 consecutive missed polls
        # This syncs with OLT when ONUs are deleted from OLT config
        # IMPORTANT: Only process missing polls if SNMP returned actual data.
        # If SNMP returned 0 ONUs, it's likely a timeout/error, not all ONUs deleted.
        onus_to_delete = []
        if len(onus_data) > 0:
            # SNMP over UDP is lossy: a PARTIAL response (some ONUs timed out)
            # must not trigger deletion of the absent-but-live ONUs. Only trust
            # "absence == removed" when the poll clearly saw most known ONUs.
            poll_looks_complete = len(seen_keys) >= max(3, int(0.5 * len(existing_onus)))
            for key, onu in existing_onus.items():
                if key not in seen_keys:
                    onu.missing_polls += 1
                    onu.updated_at = datetime.utcnow()

                    if onu.missing_polls >= 3 and poll_looks_complete:
                        # ONU not seen for 3 polls of a complete-looking poll - delete it
                        onus_to_delete.append(onu)
                        print(f"Auto-deleting ONU {onu.mac_address} (PON {onu.pon_port}/{onu.onu_id}) - not found in {onu.missing_polls} consecutive polls")
                    elif onu.is_online:
                        # Mark offline but keep tracking
                        onu.is_online = False
                        onu.distance = None
                        onu.rx_power = None
                        onus_went_offline.append(onu)

            # Delete ONUs that have been missing for 3+ polls
            for onu in onus_to_delete:
                db.delete(onu)
        else:
            logger.warning(f"Manual poll for {olt.name}: SNMP returned 0 ONUs, skipping offline detection")

        # Send batched notification for all status changes
        send_whatsapp_notification_batch(db, onus_went_online, onus_went_offline, olt.name)

        db.commit()

        # Collect traffic history (including port traffic)
        await collect_traffic_history(olt, db)
        db.commit()

        return PollResult(
            olt_id=olt.id,
            olt_name=olt.name,
            success=True,
            message=f"Successfully polled. Found {len(onus_data)} ONUs.",
            onus_found=len(onus_data)
        )

    except Exception as e:
        olt.is_online = False
        olt.last_poll = get_current_time_in_timezone(db)
        olt.last_error = str(e)
        db.commit()

        return PollResult(
            olt_id=olt.id,
            olt_name=olt.name,
            success=False,
            message=str(e),
            onus_found=0
        )


# ============ OLT Control Panel Endpoints ============

@app.get("/api/olts/{olt_id}/vlans")
async def get_olt_vlans(olt_id: int, user: User = Depends(require_admin), db: Session = Depends(get_db)):
    """Get VLAN configuration from OLT (admin only)"""
    olt = db.query(OLT).filter(OLT.id == olt_id).first()
    if not olt:
        raise HTTPException(status_code=404, detail="OLT not found")

    try:
        loop = asyncio.get_event_loop()
        connector = OLTConnector(olt.ip_address, olt.username, decrypt_sensitive(olt.password))
        get_vlan = getattr(connector, "get_vlan_config", None)
        if not callable(get_vlan):
            # VLAN read isn't implemented for this OLT — degrade gracefully
            # instead of 500 so the dashboard doesn't error out.
            return {"success": False, "vlans": [], "raw_config": "",
                    "message": "VLAN configuration read is not supported for this OLT."}
        vlan_config = await loop.run_in_executor(thread_executor, get_vlan)
        return {"success": True, "vlans": vlan_config.get('vlans', []), "raw_config": vlan_config.get('raw_config', '')}
    except Exception as e:
        logger.warning(f"get_olt_vlans failed for {olt.ip_address}: {e}")
        return {"success": False, "vlans": [], "raw_config": "", "message": f"Could not read VLANs: {e}"}


class SetONUVlanRequest(BaseModel):
    pon_port: int
    onu_id: int
    vlan_id: int
    mode: str = "tag"  # transparent, tag, translate


@app.post("/api/olts/{olt_id}/set-onu-vlan")
async def set_onu_vlan(olt_id: int, request: SetONUVlanRequest, user: User = Depends(require_admin), db: Session = Depends(get_db)):
    """Set VLAN for a specific ONU (admin only)"""
    olt = db.query(OLT).filter(OLT.id == olt_id).first()
    if not olt:
        raise HTTPException(status_code=404, detail="OLT not found")

    try:
        loop = asyncio.get_event_loop()
        connector = OLTConnector(olt.ip_address, olt.username, decrypt_sensitive(olt.password))
        result = await loop.run_in_executor(
            thread_executor,
            lambda: connector.set_onu_vlan(request.pon_port, request.onu_id, request.vlan_id, request.mode)
        )
        return {"success": result, "message": f"VLAN {request.vlan_id} set for ONU {request.pon_port}:{request.onu_id} in {request.mode} mode"}
    except AttributeError:
        raise HTTPException(status_code=501, detail="Setting ONU VLAN is not implemented for this OLT yet.")
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


class SetPortStatusRequest(BaseModel):
    port_type: str  # pon, ge, xge
    port_number: int
    enabled: bool


@app.post("/api/olts/{olt_id}/set-port-status")
async def set_port_status(olt_id: int, request: SetPortStatusRequest, user: User = Depends(require_admin), db: Session = Depends(get_db)):
    """Enable or disable a port on OLT (admin only)"""
    olt = db.query(OLT).filter(OLT.id == olt_id).first()
    if not olt:
        raise HTTPException(status_code=404, detail="OLT not found")

    try:
        loop = asyncio.get_event_loop()
        connector = OLTConnector(olt.ip_address, olt.username, decrypt_sensitive(olt.password))
        result = await loop.run_in_executor(
            thread_executor,
            lambda: connector.set_port_status(request.port_type, request.port_number, request.enabled)
        )
        status = "enabled" if request.enabled else "disabled"
        return {"success": result, "message": f"Port {request.port_type} {request.port_number} {status}"}
    except AttributeError:
        raise HTTPException(status_code=501, detail="Enabling/disabling a port is not implemented for this OLT yet.")
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.post("/api/olts/{olt_id}/reboot")
async def reboot_olt(olt_id: int, user: User = Depends(require_admin), db: Session = Depends(get_db)):
    """Reboot the entire OLT device (admin only)"""
    olt = db.query(OLT).filter(OLT.id == olt_id).first()
    if not olt:
        raise HTTPException(status_code=404, detail="OLT not found")

    try:
        loop = asyncio.get_event_loop()
        connector = OLTConnector(olt.ip_address, olt.username, decrypt_sensitive(olt.password))
        result = await loop.run_in_executor(thread_executor, connector.reboot_olt)

        # Mark OLT as offline since it's rebooting
        olt.is_online = False
        olt.last_error = "Rebooting..."
        db.commit()

        return {"success": result, "message": f"OLT {olt.name} is rebooting. It will take 2-3 minutes to come back online."}
    except AttributeError:
        raise HTTPException(status_code=501, detail="OLT reboot is not implemented for this OLT yet.")
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.post("/api/olts/{olt_id}/save-config")
async def save_olt_config(olt_id: int, user: User = Depends(require_admin), db: Session = Depends(get_db)):
    """Save running config to startup config (admin only)"""
    olt = db.query(OLT).filter(OLT.id == olt_id).first()
    if not olt:
        raise HTTPException(status_code=404, detail="OLT not found")

    try:
        loop = asyncio.get_event_loop()
        connector = OLTConnector(olt.ip_address, olt.username, decrypt_sensitive(olt.password))
        result = await loop.run_in_executor(thread_executor, connector.save_config)
        return {"success": result, "message": "Configuration saved successfully"}
    except AttributeError:
        raise HTTPException(status_code=501, detail="Save-config is not implemented for this OLT yet.")
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


class ExecuteCommandRequest(BaseModel):
    command: str


# Security: Whitelist of allowed commands for OLT execution
ALLOWED_OLT_COMMANDS = [
    'show',           # Read-only show commands
    'display',        # Display commands (some OLTs use this)
    'ping',           # Network diagnostics
    'traceroute',     # Network diagnostics
]

# Security: Blacklist of dangerous commands that are never allowed
BLOCKED_OLT_COMMANDS = [
    'delete', 'erase', 'format', 'reset',      # Destructive
    'no interface', 'no confirm', 'no onu',    # Config deletion
    'system', 'reload', 'reboot',              # System restart
    'copy', 'tftp', 'ftp',                     # File transfer
    'password', 'secret', 'enable',            # Credential access
    'radius', 'tacacs',                        # AAA config
    'crypto', 'key', 'certificate',            # Security config
    'terminal', 'shell', 'bash',               # Shell access
    'rm ', 'wget', 'curl',                     # Linux commands
    ';', '&&', '||', '|', '`', '$(',           # Command injection
]


def validate_olt_command(command: str) -> tuple[bool, str]:
    """Validate OLT command against security rules"""
    cmd_lower = command.lower().strip()

    # Check for command injection patterns
    for blocked in BLOCKED_OLT_COMMANDS:
        if blocked in cmd_lower:
            return False, f"Blocked command or pattern: {blocked}"

    # Check if command starts with allowed prefix
    cmd_starts_valid = False
    for allowed in ALLOWED_OLT_COMMANDS:
        if cmd_lower.startswith(allowed):
            cmd_starts_valid = True
            break

    if not cmd_starts_valid:
        return False, f"Command must start with one of: {', '.join(ALLOWED_OLT_COMMANDS)}"

    # Command length limit
    if len(command) > 500:
        return False, "Command too long (max 500 characters)"

    return True, "OK"


@app.post("/api/olts/{olt_id}/execute-command")
async def execute_olt_command(olt_id: int, request: ExecuteCommandRequest, user: User = Depends(require_admin), db: Session = Depends(get_db)):
    """Execute a custom CLI command on OLT (admin only) - restricted to safe read-only commands"""
    olt = db.query(OLT).filter(OLT.id == olt_id).first()
    if not olt:
        raise HTTPException(status_code=404, detail="OLT not found")

    # Validate command against security rules
    is_valid, message = validate_olt_command(request.command)
    if not is_valid:
        logger.warning(f"[SECURITY] Blocked command attempt by {user.username}: {request.command}")
        raise HTTPException(status_code=400, detail=f"Command not allowed: {message}")

    try:
        loop = asyncio.get_event_loop()
        connector = OLTConnector(olt.ip_address, olt.username, decrypt_sensitive(olt.password))
        output = await loop.run_in_executor(
            thread_executor,
            lambda: connector.execute_custom_command(request.command)
        )
        logger.info(f"[AUDIT] Command executed by {user.username} on OLT {olt.name}: {request.command}")
        return {"success": True, "output": output}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


# ============ ONU Endpoints ============

@app.get("/api/olts/{olt_id}/onus", response_model=ONUListResponse)
def list_onus_by_olt(olt_id: int, user: User = Depends(require_auth), db: Session = Depends(get_db)):
    """List ONUs for specific OLT"""
    olt = db.query(OLT).filter(OLT.id == olt_id).first()
    if not olt:
        raise HTTPException(status_code=404, detail="OLT not found")

    # Sort: online first, then by pon_port/onu_id
    onus = db.query(ONU).filter(ONU.olt_id == olt_id).order_by(
        ONU.is_online.desc(), ONU.pon_port, ONU.onu_id
    ).all()

    response_onus = []
    for onu in onus:
        region_name = None
        region_color = None
        if onu.region_id:
            region = db.query(Region).filter(Region.id == onu.region_id).first()
            if region:
                region_name = region.name
                region_color = region.color
        response_onus.append(ONUResponse(
            id=onu.id,
            olt_id=onu.olt_id,
            olt_name=olt.name,
            region_id=onu.region_id,
            region_name=region_name,
            region_color=region_color,
            pon_port=onu.pon_port,
            onu_id=onu.onu_id,
            mac_address=onu.mac_address,
            description=onu.description,
            is_online=onu.is_online,
            latitude=onu.latitude,
            longitude=onu.longitude,
            address=onu.address,
            google_maps_url=get_google_maps_url(onu.latitude, onu.longitude),
            distance=onu.distance,
            rx_power=onu.rx_power,
            onu_rx_power=onu.onu_rx_power,
            onu_tx_power=onu.onu_tx_power,
            onu_temperature=onu.onu_temperature,
            onu_voltage=onu.onu_voltage,
            onu_tx_bias=onu.onu_tx_bias,
            model=onu.model,
            image_url=onu.image_url,
            image_urls=parse_image_urls(onu.image_urls),
            uptime=get_onu_uptime(onu),
            offline_reason=onu.offline_reason if not onu.is_online else None,
            last_seen=onu.last_seen,
            created_at=onu.created_at
        ))

    return ONUListResponse(onus=response_onus, total=len(response_onus))


@app.get("/api/onus", response_model=ONUListResponse)
def list_all_onus(
    q: Optional[str] = Query(None, description="Search by customer name or MAC"),
    olt_id: Optional[int] = Query(None, description="Filter by OLT ID"),
    pon_port: Optional[int] = Query(None, description="Filter by PON port"),
    region_id: Optional[int] = Query(None, description="Filter by Region ID"),
    online_only: bool = Query(False, description="Show only online ONUs"),
    user: User = Depends(require_auth),
    db: Session = Depends(get_db)
):
    """List all ONUs with optional filters (filtered by user access)"""
    allowed_olt_ids = get_user_allowed_olt_ids(user, db)

    query = db.query(ONU, OLT.name.label("olt_name")).join(OLT)

    # Filter by user's allowed OLTs
    if allowed_olt_ids is not None:
        query = query.filter(ONU.olt_id.in_(allowed_olt_ids))

    if q:
        search = f"%{q}%"
        query = query.filter(
            or_(
                ONU.description.ilike(search),
                ONU.mac_address.ilike(search)
            )
        )

    if olt_id:
        query = query.filter(ONU.olt_id == olt_id)

    if pon_port:
        query = query.filter(ONU.pon_port == pon_port)

    if region_id:
        query = query.filter(ONU.region_id == region_id)

    if online_only:
        query = query.filter(ONU.is_online == True)

    # Sort: online first, then by OLT name, pon_port, onu_id
    results = query.order_by(ONU.is_online.desc(), OLT.name, ONU.pon_port, ONU.onu_id).all()

    # Build a cache for region names and colors
    region_ids = set(onu.region_id for onu, _ in results if onu.region_id)
    region_names = {}
    region_colors = {}
    if region_ids:
        regions = db.query(Region).filter(Region.id.in_(region_ids)).all()
        region_names = {r.id: r.name for r in regions}
        region_colors = {r.id: r.color for r in regions}

    response_onus = [
        ONUResponse(
            id=onu.id,
            olt_id=onu.olt_id,
            olt_name=olt_name,
            region_id=onu.region_id,
            region_name=region_names.get(onu.region_id),
            region_color=region_colors.get(onu.region_id),
            pon_port=onu.pon_port,
            onu_id=onu.onu_id,
            mac_address=onu.mac_address,
            description=onu.description,
            is_online=onu.is_online,
            latitude=onu.latitude,
            longitude=onu.longitude,
            address=onu.address,
            google_maps_url=get_google_maps_url(onu.latitude, onu.longitude),
            distance=onu.distance,
            rx_power=onu.rx_power,
            onu_rx_power=onu.onu_rx_power,
            onu_tx_power=onu.onu_tx_power,
            onu_temperature=onu.onu_temperature,
            onu_voltage=onu.onu_voltage,
            onu_tx_bias=onu.onu_tx_bias,
            model=onu.model,
            image_url=onu.image_url,
            image_urls=parse_image_urls(onu.image_urls),
            uptime=get_onu_uptime(onu),
            offline_reason=onu.offline_reason if not onu.is_online else None,
            last_seen=onu.last_seen,
            created_at=onu.created_at
        )
        for onu, olt_name in results
    ]

    return ONUListResponse(onus=response_onus, total=len(response_onus))


@app.get("/api/onus/search", response_model=ONUListResponse)
def search_onus(
    q: str = Query(..., min_length=1, description="Search query"),
    user: User = Depends(require_auth),
    db: Session = Depends(get_db)
):
    """Search ONUs by customer name or MAC address"""
    search = f"%{q}%"
    # Sort: online first, then by OLT name, pon_port, onu_id
    results = db.query(ONU, OLT.name.label("olt_name")).join(OLT).filter(
        or_(
            ONU.description.ilike(search),
            ONU.mac_address.ilike(search)
        )
    ).order_by(ONU.is_online.desc(), OLT.name, ONU.pon_port, ONU.onu_id).all()

    # Build a cache for region names and colors
    region_ids = set(onu.region_id for onu, _ in results if onu.region_id)
    region_names = {}
    region_colors = {}
    if region_ids:
        regions = db.query(Region).filter(Region.id.in_(region_ids)).all()
        region_names = {r.id: r.name for r in regions}
        region_colors = {r.id: r.color for r in regions}

    response_onus = [
        ONUResponse(
            id=onu.id,
            olt_id=onu.olt_id,
            olt_name=olt_name,
            region_id=onu.region_id,
            region_name=region_names.get(onu.region_id),
            region_color=region_colors.get(onu.region_id),
            pon_port=onu.pon_port,
            onu_id=onu.onu_id,
            mac_address=onu.mac_address,
            description=onu.description,
            is_online=onu.is_online,
            latitude=onu.latitude,
            longitude=onu.longitude,
            address=onu.address,
            google_maps_url=get_google_maps_url(onu.latitude, onu.longitude),
            distance=onu.distance,
            rx_power=onu.rx_power,
            onu_rx_power=onu.onu_rx_power,
            onu_tx_power=onu.onu_tx_power,
            onu_temperature=onu.onu_temperature,
            onu_voltage=onu.onu_voltage,
            onu_tx_bias=onu.onu_tx_bias,
            model=onu.model,
            image_url=onu.image_url,
            image_urls=parse_image_urls(onu.image_urls),
            uptime=get_onu_uptime(onu),
            offline_reason=onu.offline_reason if not onu.is_online else None,
            last_seen=onu.last_seen,
            created_at=onu.created_at
        )
        for onu, olt_name in results
    ]

    return ONUListResponse(onus=response_onus, total=len(response_onus))


@app.get("/api/onus/{onu_id}", response_model=ONUResponse)
def get_onu(onu_id: int, user: User = Depends(require_auth), db: Session = Depends(get_db)):
    """Get specific ONU by ID"""
    result = db.query(ONU, OLT.name.label("olt_name")).join(OLT).filter(
        ONU.id == onu_id
    ).first()

    if not result:
        raise HTTPException(status_code=404, detail="ONU not found")

    onu, olt_name = result

    region_name = None
    region_color = None
    if onu.region_id:
        region = db.query(Region).filter(Region.id == onu.region_id).first()
        if region:
            region_name = region.name
            region_color = region.color

    return ONUResponse(
        id=onu.id,
        olt_id=onu.olt_id,
        olt_name=olt_name,
        region_id=onu.region_id,
        region_name=region_name,
        region_color=region_color,
        pon_port=onu.pon_port,
        onu_id=onu.onu_id,
        mac_address=onu.mac_address,
        description=onu.description,
        is_online=onu.is_online,
        latitude=onu.latitude,
        longitude=onu.longitude,
        address=onu.address,
        google_maps_url=get_google_maps_url(onu.latitude, onu.longitude),
        distance=onu.distance,
        rx_power=onu.rx_power,
        onu_rx_power=onu.onu_rx_power,
        onu_tx_power=onu.onu_tx_power,
        onu_temperature=onu.onu_temperature,
        onu_voltage=onu.onu_voltage,
        onu_tx_bias=onu.onu_tx_bias,
        model=onu.model,
        image_url=onu.image_url,
        image_urls=parse_image_urls(onu.image_urls),
        uptime=get_onu_uptime(onu),
        offline_reason=onu.offline_reason if not onu.is_online else None,
        last_seen=onu.last_seen,
        created_at=onu.created_at
    )


def sync_onu_description_to_olt(olt_ip: str, olt_username: str, olt_password: str,
                                  pon_port: int, onu_id: int, description: str,
                                  olt_model: str = None):
    """Background task to sync ONU description to OLT via the model driver."""
    try:
        driver_cls = get_driver_class(olt_model)
        driver = driver_cls(
            ip=olt_ip,
            snmp_community="public",
            web_username=olt_username,
            web_password=olt_password,
        )
        success = driver.set_onu_description(pon_port, onu_id, description or "")
        if success:
            logger.info(f"Background sync: ONU 0/{pon_port}:{onu_id} description synced to OLT {olt_ip}")
        else:
            logger.warning(f"Background sync: Failed to sync description for ONU 0/{pon_port}:{onu_id} on {olt_ip}")
    except Exception as e:
        logger.error(f"Background sync failed for ONU 0/{pon_port}:{onu_id} on {olt_ip}: {e}")


@app.put("/api/onus/{onu_id}", response_model=ONUResponse)
async def update_onu(onu_id: int, data: dict, background_tasks: BackgroundTasks,
                     user: User = Depends(require_auth), db: Session = Depends(get_db)):
    """Update ONU description and/or region (any logged in user)"""
    result = db.query(ONU, OLT).join(OLT).filter(
        ONU.id == onu_id
    ).first()

    if not result:
        raise HTTPException(status_code=404, detail="ONU not found")

    onu, olt = result

    if "description" in data:
        new_desc = data["description"] or None

        # Update local database immediately
        onu.description = new_desc
        onu.updated_at = datetime.utcnow()

        # Sync to OLT in background via web interface (non-blocking)
        web_user = olt.web_username or olt.username or 'admin'
        web_pass = decrypt_sensitive(olt.web_password) if olt.web_password else decrypt_sensitive(olt.password) or 'admin'
        background_tasks.add_task(
            sync_onu_description_to_olt,
            olt.ip_address, web_user, web_pass,
            onu.pon_port, onu.onu_id, new_desc or "",
            olt.model  # Pass OLT model for correct URL format
        )
        logger.info(f"Queued background sync for ONU {onu_id} to OLT {olt.name} ({olt.model}) via web")

    # Handle region_id update (can be set to null to remove from region)
    if "region_id" in data:
        new_region_id = data["region_id"]
        if new_region_id is not None:
            # Verify region exists
            region = db.query(Region).filter(Region.id == new_region_id).first()
            if not region:
                raise HTTPException(status_code=400, detail="Region not found")
        onu.region_id = new_region_id
        onu.updated_at = datetime.utcnow()

    # Handle location updates
    if "latitude" in data:
        onu.latitude = data["latitude"]
        onu.updated_at = datetime.utcnow()
    if "longitude" in data:
        onu.longitude = data["longitude"]
        onu.updated_at = datetime.utcnow()
    if "address" in data:
        onu.address = data["address"]
        onu.updated_at = datetime.utcnow()
    if "image_url" in data:
        onu.image_url = data["image_url"]
        onu.updated_at = datetime.utcnow()

    db.commit()
    db.refresh(onu)

    region_name = None
    region_color = None
    if onu.region_id:
        region = db.query(Region).filter(Region.id == onu.region_id).first()
        if region:
            region_name = region.name
            region_color = region.color

    # Parse image_urls from JSON if exists
    return ONUResponse(
        id=onu.id,
        olt_id=onu.olt_id,
        olt_name=olt.name,
        region_id=onu.region_id,
        region_name=region_name,
        region_color=region_color,
        pon_port=onu.pon_port,
        onu_id=onu.onu_id,
        mac_address=onu.mac_address,
        description=onu.description,
        is_online=onu.is_online,
        latitude=onu.latitude,
        longitude=onu.longitude,
        address=onu.address,
        google_maps_url=get_google_maps_url(onu.latitude, onu.longitude),
        distance=onu.distance,
        rx_power=onu.rx_power,
        onu_rx_power=onu.onu_rx_power,
        onu_tx_power=onu.onu_tx_power,
        onu_temperature=onu.onu_temperature,
        onu_voltage=onu.onu_voltage,
        onu_tx_bias=onu.onu_tx_bias,
        model=onu.model,
        image_url=onu.image_url,
        image_urls=parse_image_urls(onu.image_urls),
        uptime=get_onu_uptime(onu),
        offline_reason=onu.offline_reason if not onu.is_online else None,
        last_seen=onu.last_seen,
        created_at=onu.created_at
    )


@app.post("/api/onus/{onu_id}/image")
async def upload_onu_image(onu_id: int, file: UploadFile = File(...),
                           user: User = Depends(require_auth), db: Session = Depends(get_db)):
    """Upload building/location image for ONU (supports up to 3 images)"""
    onu = db.query(ONU).filter(ONU.id == onu_id).first()
    if not onu:
        raise HTTPException(status_code=404, detail="ONU not found")

    # Parse existing images
    existing_images = []
    if onu.image_urls:
        try:
            existing_images = json.loads(onu.image_urls)
        except:
            existing_images = []

    # Check if already at max (3 images)
    if len(existing_images) >= 3:
        raise HTTPException(status_code=400, detail="Maximum 3 images allowed. Delete an existing image first.")

    # Validate file type
    allowed_types = ["image/jpeg", "image/png", "image/gif", "image/webp"]
    if file.content_type not in allowed_types:
        raise HTTPException(status_code=400, detail="Invalid file type. Allowed: JPEG, PNG, GIF, WEBP")

    # Limit file size (5MB)
    contents = await file.read()
    if len(contents) > 5 * 1024 * 1024:
        raise HTTPException(status_code=400, detail="File too large. Maximum size is 5MB")

    # Generate unique filename. Whitelist the extension from the (untrusted)
    # client filename so it can't smuggle path separators into filepath.
    ext = (file.filename.rsplit('.', 1)[-1].lower() if '.' in file.filename else 'jpg')
    if ext not in ('jpg', 'jpeg', 'png', 'gif', 'webp'):
        ext = 'jpg'
    filename = f"onu_{onu_id}_{uuid.uuid4().hex[:8]}.{ext}"
    filepath = os.path.join(UPLOAD_DIR, filename)

    # Save new image
    with open(filepath, 'wb') as f:
        f.write(contents)

    # Add to list and update database
    new_image_url = f"/uploads/{filename}"
    existing_images.append(new_image_url)
    onu.image_urls = json.dumps(existing_images)
    onu.image_url = existing_images[0]  # Keep first image as legacy field
    onu.updated_at = datetime.utcnow()
    db.commit()

    return {"image_url": new_image_url, "image_urls": existing_images, "message": "Image uploaded successfully"}


@app.delete("/api/onus/{onu_id}/image")
def delete_onu_image(onu_id: int, image_index: int = 0, user: User = Depends(require_auth), db: Session = Depends(get_db)):
    """Delete building/location image for ONU by index (0, 1, or 2)"""
    onu = db.query(ONU).filter(ONU.id == onu_id).first()
    if not onu:
        raise HTTPException(status_code=404, detail="ONU not found")

    # Parse existing images
    existing_images = []
    if onu.image_urls:
        try:
            existing_images = json.loads(onu.image_urls)
        except:
            existing_images = []

    if image_index < 0 or image_index >= len(existing_images):
        raise HTTPException(status_code=400, detail=f"Invalid image index. Available: 0-{len(existing_images)-1}")

    # Delete file from disk
    image_url = existing_images[image_index]
    filename = image_url.split('/')[-1]
    filepath = os.path.join(UPLOAD_DIR, filename)
    if os.path.exists(filepath):
        os.remove(filepath)

    # Remove from list and update database
    existing_images.pop(image_index)
    onu.image_urls = json.dumps(existing_images) if existing_images else None
    onu.image_url = existing_images[0] if existing_images else None  # Update legacy field
    onu.updated_at = datetime.utcnow()
    db.commit()

    return {"image_urls": existing_images, "message": "Image deleted successfully"}


@app.delete("/api/onus/{onu_id}", status_code=204)
def delete_onu(onu_id: int, delete_from_olt: bool = True, user: User = Depends(require_admin), db: Session = Depends(get_db)):
    """Delete ONU record (admin only). Also deletes from OLT if delete_from_olt=true."""
    logger.info(f"Delete ONU request: onu_id={onu_id}, delete_from_olt={delete_from_olt}, user={user.username}")
    try:
        from olt_web_scraper import delete_onu_web

        onu = db.query(ONU).filter(ONU.id == onu_id).first()
        if not onu:
            raise HTTPException(status_code=404, detail="ONU not found")

        # Try to delete from OLT first if requested
        olt_delete_success = False

        # Delete from OLT if requested and OLT is online
        if delete_from_olt:
            # Get the OLT
            olt = db.query(OLT).filter(OLT.id == onu.olt_id).first()
            if olt and olt.is_online:
                try:
                    # Use web credentials if set, otherwise fall back to standard credentials
                    web_user = olt.web_username or olt.username or 'admin'
                    web_pass = decrypt_sensitive(olt.web_password) if olt.web_password else decrypt_sensitive(olt.password) or 'admin'

                    olt_delete_success = delete_onu_web(
                        ip=olt.ip_address,
                        pon_port=onu.pon_port,
                        onu_id=onu.onu_id,
                        username=web_user,
                        password=web_pass
                    )
                    if olt_delete_success:
                        logger.info(f"Successfully deleted ONU {onu.mac_address} from OLT {olt.name}")
                except Exception as e:
                    logger.warning(f"Failed to delete ONU from OLT: {e}")
            elif olt and not olt.is_online:
                logger.warning(f"OLT {olt.name} is offline - cannot delete ONU from OLT")

        # Store info before delete for logging
        onu_mac = onu.mac_address
        onu_name = onu.description or onu_mac
        olt_id = onu.olt_id

        # Delete related traffic history records first (foreign key constraint)
        logger.info(f"Deleting traffic history for ONU {onu_id}")
        try:
            deleted_count = db.query(TrafficHistory).filter(TrafficHistory.onu_db_id == onu_id).delete(synchronize_session='fetch')
            logger.info(f"Deleted {deleted_count} traffic history records")
        except Exception as e:
            logger.warning(f"Error deleting traffic history: {e}")

        # Delete from database
        logger.info(f"Deleting ONU {onu_id} from database")
        db.delete(onu)

        # Log event in same transaction (optional - don't fail if this fails)
        try:
            current_time = get_current_time_in_timezone(db)
            event = EventLog(
                event_type='onu_deleted',
                entity_type='onu',
                entity_id=onu_id,
                olt_id=olt_id,
                description=f"ONU '{onu_name}' ({onu_mac}) deleted by {user.username}" +
                           (f" - also removed from OLT" if olt_delete_success else ""),
                created_at=current_time
            )
            db.add(event)
            logger.info(f"Created event log for ONU deletion")
        except Exception as e:
            logger.warning(f"Failed to create event log: {e}")

        logger.info(f"Committing ONU deletion")
        db.commit()
        logger.info(f"ONU {onu_id} deleted successfully")
        return None
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error deleting ONU {onu_id}: {e}", exc_info=True)
        db.rollback()
        raise HTTPException(status_code=500, detail=f"Failed to delete ONU: {str(e)}")


@app.post("/api/onus/{onu_id}/reboot")
def reboot_onu(onu_id: int, user: User = Depends(require_auth), db: Session = Depends(get_db)):
    """Reboot an ONU via the OLT model driver."""
    # Get the ONU
    onu = db.query(ONU).filter(ONU.id == onu_id).first()
    if not onu:
        raise HTTPException(status_code=404, detail="ONU not found")

    # Get the OLT
    olt = db.query(OLT).filter(OLT.id == onu.olt_id).first()
    if not olt:
        raise HTTPException(status_code=404, detail="OLT not found")

    # Check if OLT is online
    if not olt.is_online:
        raise HTTPException(status_code=503, detail="OLT is offline")

    try:
        # Resolve the model driver and let it handle the vendor-specific reboot.
        try:
            driver = get_driver(olt)
        except ValueError as drv_err:
            raise HTTPException(status_code=400, detail=str(drv_err))

        success = driver.reboot_onu(onu.pon_port, onu.onu_id)

        if success:
            # Reset online_since and mark offline so polling detects online transition
            onu.online_since = None
            onu.is_online = False  # Mark offline so next poll sets online_since when ONU comes back
            db.commit()
            logger.info(f"User {user.username} rebooted ONU {onu.id} (0/{onu.pon_port}:{onu.onu_id}) on OLT {olt.name}")
            return {"success": True, "message": f"ONU {onu.description or onu.mac_address} reboot command sent"}
        else:
            raise HTTPException(status_code=500, detail="Failed to send reboot command")

    except Exception as e:
        logger.error(f"Failed to reboot ONU {onu_id}: {e}")
        raise HTTPException(status_code=500, detail=f"Failed to reboot ONU: {e}")


def get_google_maps_url(lat: float, lng: float) -> str:
    """Generate Google Maps URL from coordinates"""
    if lat is not None and lng is not None:
        return f"https://www.google.com/maps?q={lat},{lng}"
    return None


def format_uptime_seconds(seconds: int) -> str:
    """Format seconds into uptime string (e.g., '2d 5h 30m')"""
    if seconds is None or seconds < 0:
        return None
    days = seconds // 86400
    hours = (seconds % 86400) // 3600
    minutes = (seconds % 3600) // 60

    if days > 0:
        return f"{days}d {hours}h {minutes}m"
    elif hours > 0:
        return f"{hours}h {minutes}m"
    else:
        return f"{minutes}m"


def calculate_uptime(online_since: datetime) -> str:
    """Calculate uptime string from online_since timestamp"""
    if online_since is None:
        return None
    now = datetime.utcnow()
    delta = now - online_since
    total_seconds = int(delta.total_seconds())
    return format_uptime_seconds(total_seconds)


def get_onu_uptime(onu) -> str:
    """Get ONU uptime - prefer OLT-reported alive time, fallback to calculated uptime"""
    if not onu.is_online:
        return None
    # Prefer real alive time from OLT (from onustatusinfo.html)
    if onu.olt_alive_time is not None and onu.olt_alive_time > 0:
        return format_uptime_seconds(onu.olt_alive_time)
    # Fallback to calculated uptime from online_since
    return calculate_uptime(onu.online_since)

# ============ Region Endpoints ============

@app.get("/api/regions", response_model=RegionListResponse)
def list_regions(user: User = Depends(require_auth), db: Session = Depends(get_db)):
    """List regions visible to user (admin sees all, operator sees only their own)"""
    if user.role == "admin":
        # Admin sees all regions
        regions = db.query(Region).order_by(Region.name).all()
    else:
        # Operator sees only their own regions
        regions = db.query(Region).filter(Region.owner_id == user.id).order_by(Region.name).all()

    response_regions = []
    for region in regions:
        onu_count = db.query(ONU).filter(ONU.region_id == region.id).count()
        owner_name = None
        if region.owner_id:
            owner = db.query(User).filter(User.id == region.owner_id).first()
            if owner:
                owner_name = owner.full_name or owner.username
        response_regions.append(RegionResponse(
            id=region.id,
            name=region.name,
            description=region.description,
            color=region.color,
            owner_id=region.owner_id,
            owner_name=owner_name,
            latitude=region.latitude,
            longitude=region.longitude,
            address=region.address,
            google_maps_url=get_google_maps_url(region.latitude, region.longitude),
            onu_count=onu_count,
            created_at=region.created_at
        ))

    return RegionListResponse(regions=response_regions, total=len(response_regions))


@app.get("/api/regions/{region_id}", response_model=RegionResponse)
def get_region(region_id: int, user: User = Depends(require_auth), db: Session = Depends(get_db)):
    """Get specific region by ID (access controlled)"""
    region = db.query(Region).filter(Region.id == region_id).first()
    if not region:
        raise HTTPException(status_code=404, detail="Region not found")

    # Access control: admin sees all, operator sees only their own
    if user.role != "admin" and region.owner_id != user.id:
        raise HTTPException(status_code=403, detail="Access denied to this region")

    onu_count = db.query(ONU).filter(ONU.region_id == region.id).count()
    owner_name = None
    if region.owner_id:
        owner = db.query(User).filter(User.id == region.owner_id).first()
        if owner:
            owner_name = owner.full_name or owner.username

    return RegionResponse(
        id=region.id,
        name=region.name,
        description=region.description,
        color=region.color,
        owner_id=region.owner_id,
        owner_name=owner_name,
        latitude=region.latitude,
        longitude=region.longitude,
        address=region.address,
        google_maps_url=get_google_maps_url(region.latitude, region.longitude),
        onu_count=onu_count,
        created_at=region.created_at
    )


@app.post("/api/regions", response_model=RegionResponse, status_code=201)
def create_region(region_data: RegionCreate, user: User = Depends(require_auth), db: Session = Depends(get_db)):
    """Create new region (operators own their regions, admin creates global regions)"""
    # Check for duplicate name within user's scope
    if user.role == "admin":
        # Admin: check against all regions
        existing = db.query(Region).filter(Region.name == region_data.name).first()
    else:
        # Operator: check only against their own regions
        existing = db.query(Region).filter(
            Region.name == region_data.name,
            Region.owner_id == user.id
        ).first()

    if existing:
        raise HTTPException(status_code=400, detail="Region with this name already exists")

    # Set owner_id: NULL for admin (visible to all), user.id for operators (private)
    owner_id = None if user.role == "admin" else user.id

    region = Region(
        name=region_data.name,
        description=region_data.description,
        color=region_data.color,
        owner_id=owner_id,
        latitude=region_data.latitude,
        longitude=region_data.longitude,
        address=region_data.address
    )
    db.add(region)
    db.commit()
    db.refresh(region)

    owner_name = None
    if owner_id:
        owner_name = user.full_name or user.username

    return RegionResponse(
        id=region.id,
        name=region.name,
        description=region.description,
        color=region.color,
        owner_id=region.owner_id,
        owner_name=owner_name,
        latitude=region.latitude,
        longitude=region.longitude,
        address=region.address,
        google_maps_url=get_google_maps_url(region.latitude, region.longitude),
        onu_count=0,
        created_at=region.created_at
    )


@app.put("/api/regions/{region_id}", response_model=RegionResponse)
def update_region(region_id: int, region_data: RegionUpdate, user: User = Depends(require_auth), db: Session = Depends(get_db)):
    """Update region (access controlled - owner or admin)"""
    region = db.query(Region).filter(Region.id == region_id).first()
    if not region:
        raise HTTPException(status_code=404, detail="Region not found")

    # Access control: admin can edit all, operator can only edit their own
    if user.role != "admin" and region.owner_id != user.id:
        raise HTTPException(status_code=403, detail="Access denied to this region")

    if region_data.name is not None:
        # Check for duplicate name within scope
        if user.role == "admin":
            existing = db.query(Region).filter(
                Region.name == region_data.name,
                Region.id != region_id
            ).first()
        else:
            existing = db.query(Region).filter(
                Region.name == region_data.name,
                Region.owner_id == user.id,
                Region.id != region_id
            ).first()
        if existing:
            raise HTTPException(status_code=400, detail="Region with this name already exists")
        region.name = region_data.name

    if region_data.description is not None:
        region.description = region_data.description
    if region_data.color is not None:
        region.color = region_data.color
    if region_data.latitude is not None:
        region.latitude = region_data.latitude
    if region_data.longitude is not None:
        region.longitude = region_data.longitude
    if region_data.address is not None:
        region.address = region_data.address

    db.commit()
    db.refresh(region)

    onu_count = db.query(ONU).filter(ONU.region_id == region.id).count()
    owner_name = None
    if region.owner_id:
        owner = db.query(User).filter(User.id == region.owner_id).first()
        if owner:
            owner_name = owner.full_name or owner.username

    return RegionResponse(
        id=region.id,
        name=region.name,
        description=region.description,
        color=region.color,
        owner_id=region.owner_id,
        owner_name=owner_name,
        latitude=region.latitude,
        longitude=region.longitude,
        address=region.address,
        google_maps_url=get_google_maps_url(region.latitude, region.longitude),
        onu_count=onu_count,
        created_at=region.created_at
    )


@app.delete("/api/regions/{region_id}", status_code=204)
def delete_region(region_id: int, user: User = Depends(require_auth), db: Session = Depends(get_db)):
    """Delete region (admin can delete any, operator can delete their own)"""
    region = db.query(Region).filter(Region.id == region_id).first()
    if not region:
        raise HTTPException(status_code=404, detail="Region not found")

    # Access control: admin can delete all, operator can only delete their own
    if user.role != "admin" and region.owner_id != user.id:
        raise HTTPException(status_code=403, detail="Access denied to this region")

    # Clear region_id from ONUs
    db.query(ONU).filter(ONU.region_id == region_id).update({"region_id": None})

    db.delete(region)
    db.commit()
    return None


@app.get("/api/regions/{region_id}/onus", response_model=ONUListResponse)
def list_onus_by_region(region_id: int, user: User = Depends(require_auth), db: Session = Depends(get_db)):
    """List ONUs in a specific region (access controlled)"""
    region = db.query(Region).filter(Region.id == region_id).first()
    if not region:
        raise HTTPException(status_code=404, detail="Region not found")

    # Access control: admin sees all, operator sees only their own regions
    if user.role != "admin" and region.owner_id != user.id:
        raise HTTPException(status_code=403, detail="Access denied to this region")

    results = db.query(ONU, OLT.name.label("olt_name")).join(OLT).filter(
        ONU.region_id == region_id
    ).order_by(OLT.name, ONU.pon_port, ONU.onu_id).all()

    response_onus = [
        ONUResponse(
            id=onu.id,
            olt_id=onu.olt_id,
            olt_name=olt_name,
            region_id=onu.region_id,
            region_name=region.name,
            region_color=region.color,
            pon_port=onu.pon_port,
            onu_id=onu.onu_id,
            mac_address=onu.mac_address,
            description=onu.description,
            is_online=onu.is_online,
            latitude=onu.latitude,
            longitude=onu.longitude,
            address=onu.address,
            google_maps_url=get_google_maps_url(onu.latitude, onu.longitude),
            distance=onu.distance,
            rx_power=onu.rx_power,
            onu_rx_power=onu.onu_rx_power,
            onu_tx_power=onu.onu_tx_power,
            onu_temperature=onu.onu_temperature,
            onu_voltage=onu.onu_voltage,
            onu_tx_bias=onu.onu_tx_bias,
            model=onu.model,
            image_url=onu.image_url,
            image_urls=parse_image_urls(onu.image_urls),
            uptime=get_onu_uptime(onu),
            offline_reason=onu.offline_reason if not onu.is_online else None,
            last_seen=onu.last_seen,
            created_at=onu.created_at
        )
        for onu, olt_name in results
    ]

    return ONUListResponse(onus=response_onus, total=len(response_onus))


# ============ Authentication Endpoints ============

@app.post("/api/auth/login", response_model=LoginResponse)
def login(credentials: UserLogin, db: Session = Depends(get_db)):
    """Login and get access token"""
    # Check if account is locked
    user_check = db.query(User).filter(User.email == credentials.username).first()
    if user_check and user_check.locked_until:
        if datetime.utcnow() < user_check.locked_until:
            remaining = (user_check.locked_until - datetime.utcnow()).seconds // 60
            raise HTTPException(
                status_code=423,
                detail=f"Account locked. Try again in {remaining + 1} minutes."
            )
        else:
            # Lockout expired, reset
            user_check.locked_until = None
            user_check.failed_login_attempts = 0
            db.commit()

    user = authenticate_user(db, credentials.username, credentials.password)
    if not user:
        # Track failed login attempts
        if user_check:
            user_check.failed_login_attempts = (user_check.failed_login_attempts or 0) + 1
            # Lock account after 5 failed attempts for 15 minutes
            if user_check.failed_login_attempts >= 5:
                user_check.locked_until = datetime.utcnow() + timedelta(minutes=15)
            db.commit()
        raise HTTPException(
            status_code=401,
            detail="Invalid username or password"
        )

    # Reset failed login attempts on successful login
    user.failed_login_attempts = 0
    user.locked_until = None
    user.last_login = datetime.utcnow()
    db.commit()

    # Log login event
    log_event(db, 'user_login', 'user', user.id, None,
              f"User '{user.username}' logged in")

    # Create access token
    token = create_access_token({"user_id": user.id, "role": user.role})

    assigned_olt_ids = get_user_olt_ids_list(user, db)

    return LoginResponse(
        token=token,
        user=UserResponse(
            id=user.id,
            username=user.username,
            role=user.role,
            full_name=user.full_name,
            is_active=user.is_active,
            created_at=user.created_at,
            last_login=user.last_login,
            assigned_olt_ids=assigned_olt_ids
        ),
        must_change_password=user.must_change_password or False
    )


@app.get("/api/auth/me", response_model=UserResponse)
def get_current_user_info(user: User = Depends(require_auth), db: Session = Depends(get_db)):
    """Get current logged in user info"""
    assigned_olt_ids = get_user_olt_ids_list(user, db)

    return UserResponse(
        id=user.id,
        username=user.username,
        role=user.role,
        full_name=user.full_name,
        is_active=user.is_active,
        created_at=user.created_at,
        last_login=user.last_login,
        assigned_olt_ids=assigned_olt_ids
    )


# ============ User Management Endpoints (Admin Only) ============

@app.get("/api/users", response_model=UserListResponse)
def list_users(user: User = Depends(require_admin), db: Session = Depends(get_db)):
    """List all users (admin only)"""
    # User.username is a Python @property aliasing email after Phase 1.
    # SQLAlchemy needs the actual column name for ORDER BY.
    users = db.query(User).order_by(User.email).all()

    response_users = []
    for u in users:
        assigned_olt_ids = get_user_olt_ids_list(u, db)
        response_users.append(UserResponse(
            id=u.id,
            username=u.username,
            role=u.role,
            full_name=u.full_name,
            is_active=u.is_active,
            created_at=u.created_at,
            last_login=u.last_login,
            assigned_olt_ids=assigned_olt_ids
        ))

    return UserListResponse(users=response_users, total=len(response_users))


@app.post("/api/users", response_model=UserResponse, status_code=201)
def create_user(user_data: UserCreate, user: User = Depends(require_admin), db: Session = Depends(get_db)):
    """Create new user (admin only)"""
    # Check license user limit
    from license_manager import license_manager
    license_info = license_manager.get_license_info()
    max_users = license_info.get('max_users', 5)
    current_user_count = db.query(User).count()

    if current_user_count >= max_users:
        package = license_info.get('package_type', 'trial')
        raise HTTPException(
            status_code=403,
            detail=f"User limit reached ({max_users}). Upgrade your package to add more users. Current package: {package}"
        )

    # Check for duplicate username (Phase 1: column is now `email`)
    existing = db.query(User).filter(User.email == user_data.username).first()
    if existing:
        raise HTTPException(status_code=400, detail="Username already exists")

    new_user = User(
        username=user_data.username,
        password_hash=get_password_hash(user_data.password),
        role=user_data.role,
        full_name=user_data.full_name
    )
    db.add(new_user)
    db.commit()
    db.refresh(new_user)

    # Handle OLT assignments (for operators)
    assigned_olt_ids = user_data.assigned_olt_ids or []
    if assigned_olt_ids and user_data.role == "operator":
        for olt_id in assigned_olt_ids:
            # Verify OLT exists
            olt = db.query(OLT).filter(OLT.id == olt_id).first()
            if olt:
                db.execute(user_olts.insert().values(user_id=new_user.id, olt_id=olt_id))
        db.commit()

    # Log event
    log_event(db, 'user_created', 'user', new_user.id, None,
              f"User '{new_user.username}' created with role '{new_user.role}' by {user.username}")

    return UserResponse(
        id=new_user.id,
        username=new_user.username,
        role=new_user.role,
        full_name=new_user.full_name,
        is_active=new_user.is_active,
        created_at=new_user.created_at,
        last_login=new_user.last_login,
        assigned_olt_ids=assigned_olt_ids if user_data.role == "operator" else []
    )


@app.put("/api/users/{user_id}", response_model=UserResponse)
def update_user(user_id: int, user_data: UserUpdate, user: User = Depends(require_admin), db: Session = Depends(get_db)):
    """Update user (admin only)"""
    target_user = db.query(User).filter(User.id == user_id).first()
    if not target_user:
        raise HTTPException(status_code=404, detail="User not found")

    if user_data.password is not None:
        target_user.password_hash = get_password_hash(user_data.password)
    if user_data.role is not None:
        target_user.role = user_data.role
    if user_data.full_name is not None:
        target_user.full_name = user_data.full_name
    if user_data.is_active is not None:
        target_user.is_active = user_data.is_active

    # Handle OLT assignments update
    if user_data.assigned_olt_ids is not None:
        # Clear existing assignments
        db.execute(user_olts.delete().where(user_olts.c.user_id == user_id))

        # Add new assignments (only for operators)
        role = user_data.role if user_data.role is not None else target_user.role
        if role == "operator":
            for olt_id in user_data.assigned_olt_ids:
                olt = db.query(OLT).filter(OLT.id == olt_id).first()
                if olt:
                    db.execute(user_olts.insert().values(user_id=user_id, olt_id=olt_id))

    db.commit()
    db.refresh(target_user)

    assigned_olt_ids = get_user_olt_ids_list(target_user, db)

    return UserResponse(
        id=target_user.id,
        username=target_user.username,
        role=target_user.role,
        full_name=target_user.full_name,
        is_active=target_user.is_active,
        created_at=target_user.created_at,
        last_login=target_user.last_login,
        assigned_olt_ids=assigned_olt_ids
    )


@app.delete("/api/users/{user_id}", status_code=204)
def delete_user(user_id: int, user: User = Depends(require_admin), db: Session = Depends(get_db)):
    """Delete user (admin only)"""
    if user_id == user.id:
        raise HTTPException(status_code=400, detail="Cannot delete yourself")

    target_user = db.query(User).filter(User.id == user_id).first()
    if not target_user:
        raise HTTPException(status_code=404, detail="User not found")

    username = target_user.username
    db.delete(target_user)
    db.commit()

    # Log event
    log_event(db, 'user_deleted', 'user', user_id, None,
              f"User '{username}' deleted by {user.username}")

    return None


# ============ Update Check API ============

@app.get("/api/update-check")
def check_for_updates(current_user: dict = Depends(get_current_user)):
    """Check if software update is available"""
    from license_manager import license_manager, SOFTWARE_VERSION

    update_info = license_manager.update_info

    return {
        "current_version": SOFTWARE_VERSION,
        "update_available": update_info is not None,
        "update": update_info
    }


# Update status tracking
update_status = {
    "in_progress": False,
    "stage": "",
    "progress": 0,
    "error": None,
    "completed": False
}


@app.get("/api/update/status")
def get_update_status(current_user: dict = Depends(get_current_user)):
    """Get current update status"""
    return update_status


@app.post("/api/update/download")
async def download_update(
    background_tasks: BackgroundTasks,
    current_user: User = Depends(require_admin)
):
    """Download update from license server"""
    global update_status
    from license_manager import license_manager, LICENSE_SERVER_URL, SOFTWARE_VERSION

    if update_status["in_progress"]:
        raise HTTPException(status_code=400, detail="Update already in progress")

    # Reset status
    update_status = {
        "in_progress": True,
        "stage": "preparing",
        "progress": 0,
        "error": None,
        "completed": False
    }

    try:
        # Get update info
        update_info = license_manager.update_info
        if not update_info:
            update_status["error"] = "No update available"
            update_status["in_progress"] = False
            raise HTTPException(status_code=400, detail="No update available")

        latest_version = update_info.get("latest_version", "1.0.0")

        # Create updates directory
        updates_dir = Path("/tmp/olt-manager-updates")
        updates_dir.mkdir(exist_ok=True)

        update_status["stage"] = "downloading"
        update_status["progress"] = 10

        # Download from license server
        try:
            response = requests.post(
                f"{LICENSE_SERVER_URL}/api/download-update",
                json={
                    "license_key": license_manager.license_key,
                    "hardware_id": license_manager.hardware_id
                },
                timeout=300,  # 5 minute timeout for download
                stream=True
            )

            if response.status_code != 200:
                error_msg = response.json().get("error", "Download failed")
                update_status["error"] = error_msg
                update_status["in_progress"] = False
                raise HTTPException(status_code=400, detail=error_msg)

            # Save the package
            package_path = updates_dir / f"olt-manager-v{latest_version}.tar.gz"
            total_size = int(response.headers.get('content-length', 0))
            downloaded = 0

            with open(package_path, 'wb') as f:
                for chunk in response.iter_content(chunk_size=8192):
                    if chunk:
                        f.write(chunk)
                        downloaded += len(chunk)
                        if total_size > 0:
                            update_status["progress"] = 10 + int((downloaded / total_size) * 40)

            update_status["stage"] = "downloading_frontend"
            update_status["progress"] = 50

            # Also download frontend
            try:
                frontend_response = requests.get(
                    f"{LICENSE_SERVER_URL}/downloads/frontend.tar.gz",
                    timeout=120,
                    stream=True
                )
                if frontend_response.status_code == 200:
                    frontend_path = updates_dir / "frontend.tar.gz"
                    with open(frontend_path, 'wb') as f:
                        for chunk in frontend_response.iter_content(chunk_size=8192):
                            if chunk:
                                f.write(chunk)
                    update_status["frontend_path"] = str(frontend_path)
                    logger.info("Frontend package downloaded successfully")
            except Exception as fe:
                logger.warning(f"Could not download frontend: {fe}")

            update_status["stage"] = "downloaded"
            update_status["progress"] = 55
            update_status["package_path"] = str(package_path)
            update_status["new_version"] = latest_version  # Save version for install step

            return {
                "success": True,
                "message": f"Update v{latest_version} downloaded successfully",
                "package_path": str(package_path),
                "version": latest_version
            }

        except requests.exceptions.RequestException as e:
            update_status["error"] = f"Download failed: {str(e)}"
            update_status["in_progress"] = False
            raise HTTPException(status_code=500, detail=f"Download failed: {str(e)}")

    except HTTPException:
        raise
    except Exception as e:
        update_status["error"] = str(e)
        update_status["in_progress"] = False
        raise HTTPException(status_code=500, detail=str(e))


@app.post("/api/update/install")
async def install_update(current_user: User = Depends(require_admin)):
    """Install downloaded update"""
    global update_status
    import tarfile
    import subprocess
    from license_manager import license_manager

    if not update_status.get("package_path"):
        raise HTTPException(status_code=400, detail="No update downloaded")

    package_path = Path(update_status["package_path"])
    if not package_path.exists():
        raise HTTPException(status_code=400, detail="Update package not found")

    try:
        update_status["stage"] = "installing"
        update_status["progress"] = 55

        # Extract to temp directory
        extract_dir = Path("/tmp/olt-manager-update-extract")
        if extract_dir.exists():
            shutil.rmtree(extract_dir)
        extract_dir.mkdir()

        update_status["progress"] = 60

        # Extract the tarball
        with tarfile.open(package_path, 'r:gz') as tar:
            _safe_extract_tar(tar, extract_dir)

        update_status["stage"] = "backing_up"
        update_status["progress"] = 70

        new_version = update_status.get("new_version", "1.1.0")

        # Check if package includes install script (new self-contained updates)
        package_install_script = extract_dir / "install.sh"
        if package_install_script.exists():
            logger.info("Using package's install script for update")
            update_status["stage"] = "applying"
            update_status["progress"] = 75

            # Make script executable
            package_install_script.chmod(0o755)

            # Run install script using systemd-run to ensure it survives service restart
            # systemd-run creates a transient service that runs independently
            install_cmd = f"/bin/bash {package_install_script} {extract_dir} {new_version}"

            # Use systemd-run to launch as a transient service
            systemd_result = subprocess.run(
                [
                    "systemd-run",
                    "--unit=olt-update-install",
                    "--description=OLT Manager Update Install",
                    "--no-block",
                    "/bin/bash", "-c", f"sleep 5 && {install_cmd}"
                ],
                capture_output=True,
                text=True
            )

            if systemd_result.returncode != 0:
                # Fallback: direct execution with nohup
                logger.warning(f"systemd-run failed: {systemd_result.stderr}, using direct fallback")
                subprocess.Popen(
                    ["/bin/bash", "-c",
                     f"nohup /bin/bash -c 'sleep 5 && {install_cmd}' > /tmp/olt-update.log 2>&1 &"],
                    start_new_session=True,
                    close_fds=True
                )
            else:
                logger.info(f"systemd-run success: {systemd_result.stdout}")

            logger.info("Install script scheduled to run in background")
            update_status["progress"] = 85
            update_status["stage"] = "restarting"

            # Return success - the install will happen in background
            return {"status": "success", "message": "Update installing... Service will restart automatically."}

        else:
            # Fallback: Legacy install logic for old packages without install.sh
            logger.info("Using legacy install logic (no install.sh in package)")

            # Detect installation type: compiled binary or source
            is_compiled = Path("/opt/olt-manager/olt-manager").exists()

            if is_compiled:
                install_dir = Path("/opt/olt-manager")
                backend_dir = None
                logger.info("Detected compiled binary installation")
            elif Path("/opt/olt-manager/backend").exists():
                install_dir = Path("/opt/olt-manager")
                backend_dir = install_dir / "backend"
            else:
                install_dir = Path("/root/olt-manager")
                backend_dir = install_dir / "backend"

            # Create backup
            backup_dir = Path("/tmp/olt-manager-backup")
            if backup_dir.exists():
                shutil.rmtree(backup_dir)
            backup_dir.mkdir(parents=True)

            if backend_dir and backend_dir.exists():
                shutil.copytree(str(backend_dir), str(backup_dir / "backend"))

            update_status["stage"] = "applying"
            update_status["progress"] = 80

            # Apply update based on installation type
            if is_compiled:
                # Compiled binary installation - update the binary
                new_binary = extract_dir / "olt-manager"
                if new_binary.exists():
                    # Copy new binary to staging location (can't overwrite running binary)
                    staged_binary = install_dir / "olt-manager.new"
                    shutil.copy2(new_binary, staged_binary)
                    staged_binary.chmod(0o755)
                    logger.info("Staged new binary as olt-manager.new")
                else:
                    logger.warning("No olt-manager binary in update package")

                # Update static folder for compiled installation
                new_static = extract_dir / "static"
                if new_static.exists():
                    # Update /opt/olt-manager/static
                    target_static = install_dir / "static"
                    if target_static.exists():
                        shutil.rmtree(target_static)
                    shutil.copytree(new_static, target_static)
                    logger.info("Updated static folder in /opt/olt-manager")

                    # Also update nginx folder if it exists
                    nginx_html = Path("/var/www/html")
                    if nginx_html.exists():
                        # Clear and copy new frontend
                        for item in nginx_html.iterdir():
                            if item.is_dir():
                                shutil.rmtree(item)
                            else:
                                item.unlink()
                        for item in new_static.iterdir():
                            dest = nginx_html / item.name
                            if item.is_dir():
                                shutil.copytree(item, dest)
                            else:
                                shutil.copy2(item, dest)
                        logger.info("Updated nginx frontend in /var/www/html")

            elif backend_dir and backend_dir.exists():
                # Source installation - update backend files
                extracted_backend = extract_dir / "backend"
                if extracted_backend.exists():
                    for item in extracted_backend.iterdir():
                        if item.name not in ["venv", "__pycache__", "data"]:
                            dest = backend_dir / item.name
                            if item.is_dir():
                                if dest.exists():
                                    shutil.rmtree(dest)
                                shutil.copytree(item, dest)
                            else:
                                shutil.copy2(item, dest)

            # Apply frontend update from package
            extracted_frontend = extract_dir / "frontend" / "build"
            if extracted_frontend.exists():
                for nginx_dir in [Path("/var/www/olt-manager"), Path("/var/www/html")]:
                    if nginx_dir.exists():
                        for item in extracted_frontend.iterdir():
                            dest = nginx_dir / item.name
                            if item.is_dir():
                                if dest.exists():
                                    shutil.rmtree(dest)
                                shutil.copytree(item, dest)
                            else:
                                shutil.copy2(item, dest)

            # Apply separately downloaded frontend (frontend.tar.gz)
            frontend_path = update_status.get("frontend_path")
            if frontend_path and Path(frontend_path).exists():
                logger.info("Installing frontend from separate package...")
                frontend_extract = Path("/tmp/frontend-extract")
                if frontend_extract.exists():
                    shutil.rmtree(frontend_extract)
                frontend_extract.mkdir()

                with tarfile.open(frontend_path, 'r:gz') as tar:
                    _safe_extract_tar(tar, frontend_extract)

                # Install to nginx folder
                nginx_html = Path("/var/www/html")
                if nginx_html.exists():
                    # Clear old files
                    for item in nginx_html.iterdir():
                        if item.is_dir():
                            shutil.rmtree(item)
                        else:
                            item.unlink()
                    # Copy new files
                    for item in frontend_extract.iterdir():
                        dest = nginx_html / item.name
                        if item.is_dir():
                            shutil.copytree(item, dest)
                        else:
                            shutil.copy2(item, dest)
                    logger.info("Frontend updated in /var/www/html")

                # Cleanup
                shutil.rmtree(frontend_extract)

            # Update version file
            update_status["progress"] = 85
            if is_compiled:
                version_file = install_dir / "VERSION"
            else:
                version_file = backend_dir / "VERSION"
            version_file.write_text(new_version)
            logger.info(f"Updated VERSION file to {new_version}")

        # Create symlink for database path compatibility (fixes backup/restore)
        # This ensures backup/restore works correctly for compiled installations
        symlink_dir = Path("/root/olt-manager/backend")
        symlink_path = symlink_dir / "olt_manager.db"
        db_target = Path("/opt/olt-manager/olt_manager.db")

        if db_target.exists() and not symlink_path.exists():
            symlink_dir.mkdir(parents=True, exist_ok=True)
            try:
                symlink_path.symlink_to(db_target)
                logger.info("Created database symlink for backup compatibility")
            except Exception as sym_err:
                logger.warning(f"Could not create symlink: {sym_err}")

        update_status["stage"] = "restarting"
        update_status["progress"] = 90

        # Create a script to restart the service after response is sent
        # With health check and auto-rollback if service fails to start
        restart_script = Path("/tmp/restart-olt-manager.sh")
        restart_script.write_text("""#!/bin/bash
sleep 2

LOG_FILE="/tmp/olt-update-restart.log"
echo "$(date): Starting update restart script" > $LOG_FILE

# Function to check if service is healthy
check_health() {
    sleep 5
    for i in {1..6}; do
        if curl -s --connect-timeout 3 http://127.0.0.1:8000/api/system/info > /dev/null 2>&1; then
            echo "$(date): Health check passed on attempt $i" >> $LOG_FILE
            return 0
        fi
        echo "$(date): Health check attempt $i failed, waiting..." >> $LOG_FILE
        sleep 3
    done
    return 1
}

# For compiled installations, swap the binary before restart
if [ -f /opt/olt-manager/olt-manager.new ]; then
    echo "$(date): Compiled installation detected" >> $LOG_FILE

    # Stop service first
    systemctl stop olt-manager 2>/dev/null || systemctl stop olt-backend 2>/dev/null
    sleep 1

    # Backup and swap binary
    cd /opt/olt-manager
    if [ -f olt-manager ]; then
        mv olt-manager olt-manager.old
        echo "$(date): Backed up old binary" >> $LOG_FILE
    fi
    mv olt-manager.new olt-manager
    chmod +x olt-manager

    # Start service
    systemctl start olt-manager 2>/dev/null || systemctl start olt-backend 2>/dev/null
    echo "$(date): Service started, checking health..." >> $LOG_FILE

    # Check if service started successfully
    if check_health; then
        echo "$(date): Update successful!" >> $LOG_FILE
        rm -f olt-manager.old  # Clean up old binary
        exit 0
    else
        # ROLLBACK: Service failed to start, restore old binary
        echo "$(date): Service failed to start! Rolling back..." >> $LOG_FILE
        systemctl stop olt-manager 2>/dev/null || systemctl stop olt-backend 2>/dev/null
        sleep 1
        if [ -f olt-manager.old ]; then
            mv olt-manager olt-manager.failed
            mv olt-manager.old olt-manager
            chmod +x olt-manager
            systemctl start olt-manager 2>/dev/null || systemctl start olt-backend 2>/dev/null
            echo "$(date): Rollback completed" >> $LOG_FILE
        fi
        exit 1
    fi
fi

# Try to restart service - check if service EXISTS (not just active)
if systemctl list-unit-files | grep -q olt-manager.service; then
    systemctl daemon-reload
    systemctl start olt-manager
    echo "$(date): Started olt-manager service" >> $LOG_FILE
elif systemctl list-unit-files | grep -q olt-backend.service; then
    systemctl daemon-reload
    systemctl start olt-backend
    echo "$(date): Started olt-backend service" >> $LOG_FILE
else
    # Fallback: manual restart (for source installations without systemd)
    cd /opt/olt-manager/backend 2>/dev/null || cd /root/olt-manager/backend
    pkill -f "uvicorn main:app" 2>/dev/null
    sleep 1
    source venv/bin/activate
    nohup python -m uvicorn main:app --host 127.0.0.1 --port 8000 > /tmp/olt-manager.log 2>&1 &
    echo "$(date): Started via fallback method" >> $LOG_FILE
fi

# Verify service started
if check_health; then
    echo "$(date): Service restart successful" >> $LOG_FILE
else
    echo "$(date): WARNING: Service may not be running properly" >> $LOG_FILE
fi
""")
        restart_script.chmod(0o755)

        # Run restart script in background
        subprocess.Popen(["/bin/bash", str(restart_script)],
                        stdout=subprocess.DEVNULL,
                        stderr=subprocess.DEVNULL,
                        start_new_session=True)

        update_status["stage"] = "completed"
        update_status["progress"] = 100
        update_status["completed"] = True
        update_status["in_progress"] = False

        # Cleanup
        if extract_dir.exists():
            shutil.rmtree(extract_dir)

        return {
            "success": True,
            "message": "Update installed successfully. Service is restarting...",
            "restart_in_seconds": 3,
            "require_logout": True
        }

    except Exception as e:
        update_status["error"] = str(e)
        update_status["in_progress"] = False
        logger.error(f"Update installation failed: {e}")
        raise HTTPException(status_code=500, detail=f"Installation failed: {str(e)}")


@app.post("/api/update/rollback")
async def rollback_update(current_user: User = Depends(require_admin)):
    """Rollback to previous version"""
    import subprocess

    backup_dir = Path("/root/olt-manager-backup")
    if not backup_dir.exists():
        raise HTTPException(status_code=400, detail="No backup available for rollback")

    try:
        # Restore backend
        backend_backup = backup_dir / "backend"
        if backend_backup.exists():
            backend_dest = Path("/root/olt-manager/backend")
            for item in backend_backup.iterdir():
                if item.name != "venv" and item.name != "__pycache__":
                    dest = backend_dest / item.name
                    if item.is_dir():
                        if dest.exists():
                            shutil.rmtree(dest)
                        shutil.copytree(item, dest)
                    else:
                        shutil.copy2(item, dest)

        # Restart service
        restart_script = Path("/tmp/restart-olt-manager.sh")
        subprocess.Popen(["/bin/bash", str(restart_script)],
                        stdout=subprocess.DEVNULL,
                        stderr=subprocess.DEVNULL,
                        start_new_session=True)

        return {
            "success": True,
            "message": "Rollback completed. Service is restarting..."
        }

    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Rollback failed: {str(e)}")


# ============ Dev Server / Publisher API ============

def is_dev_server():
    """Check if this is the development server"""
    dev_marker = Path("/root/olt-manager/backend/.dev_server")
    return dev_marker.exists()

@app.get("/api/dev/status")
async def get_dev_status():
    """Check if this is a development server - for showing publish UI"""
    from license_manager import SOFTWARE_VERSION
    return {
        "is_dev_server": is_dev_server(),
        "current_version": SOFTWARE_VERSION
    }

from pydantic import BaseModel as PydanticBaseModel

class PublishRequest(PydanticBaseModel):
    version: str
    changelog: str

@app.get("/api/dev/build-status")
async def get_build_status(current_user: User = Depends(require_admin)):
    """Check Nuitka build status"""
    if not is_dev_server():
        raise HTTPException(status_code=403, detail="This feature is only available on development server")

    build_dir = Path("/root/olt-manager/nuitka_build")
    binary_path = build_dir / "olt-manager"
    log_path = Path("/tmp/nuitka_build.log")

    # Check if build is running
    import subprocess
    ps_result = subprocess.run(["pgrep", "-f", "nuitka"], capture_output=True)
    is_building = ps_result.returncode == 0

    # Get log tail
    log_tail = ""
    if log_path.exists():
        log_tail = log_path.read_text()[-2000:]

    # Check binary
    binary_exists = binary_path.exists()
    binary_size = 0
    binary_date = None
    if binary_exists:
        stat = binary_path.stat()
        binary_size = stat.st_size / (1024 * 1024)
        from datetime import datetime
        binary_date = datetime.fromtimestamp(stat.st_mtime).strftime("%Y-%m-%d %H:%M")

    return {
        "is_building": is_building,
        "binary_ready": binary_exists and not is_building,
        "binary_size_mb": round(binary_size, 1),
        "binary_date": binary_date,
        "log": log_tail
    }

@app.post("/api/dev/build")
async def start_build(current_user: User = Depends(require_admin), background_tasks: BackgroundTasks = None):
    """Start Nuitka build in background"""
    import subprocess

    if not is_dev_server():
        raise HTTPException(status_code=403, detail="This feature is only available on development server")

    # Check if already building
    ps_result = subprocess.run(["pgrep", "-f", "nuitka"], capture_output=True)
    if ps_result.returncode == 0:
        raise HTTPException(status_code=400, detail="Build already in progress")

    # Start build in background
    build_script = '''
source /root/olt-manager/backend/venv/bin/activate
cd /root/olt-manager/backend
echo "Starting Nuitka build at $(date)"
BUILD_DIR="/root/olt-manager/nuitka_build"
rm -rf "$BUILD_DIR" main.build main.dist main.onefile-build 2>/dev/null
mkdir -p "$BUILD_DIR"
python -m nuitka --standalone --onefile --output-dir="$BUILD_DIR" --output-filename=olt-manager --follow-imports --prefer-source-code main.py
echo "Build finished at $(date)"
ls -lh "$BUILD_DIR/olt-manager" 2>/dev/null || echo "Build failed"
'''
    subprocess.Popen(
        ["bash", "-c", build_script],
        stdout=open("/tmp/nuitka_build.log", "w"),
        stderr=subprocess.STDOUT,
        start_new_session=True
    )

    return {"success": True, "message": "Nuitka build started in background. Check /api/dev/build-status for progress."}

@app.post("/api/dev/publish")
async def publish_update(
    request: PublishRequest,
    current_user: User = Depends(require_admin)
):
    """Publish PROTECTED update using pre-built binary to license server (dev server only)"""
    import subprocess
    import tarfile
    import shutil

    version = request.version
    changelog = request.changelog

    if not is_dev_server():
        raise HTTPException(status_code=403, detail="This feature is only available on development server")

    # Validate version before it's interpolated into a generated shell script
    # (avoids command injection via a crafted version string).
    if not re.match(r'^\d+\.\d+\.\d+([.-][A-Za-z0-9]+)?$', str(version or "")):
        raise HTTPException(status_code=400, detail="Invalid version format (expected semver)")

    try:
        result_steps = []
        backend_dir = Path("/root/olt-manager/backend")
        build_dir = Path("/root/olt-manager/nuitka_build")
        binary_path = build_dir / "olt-manager"

        # Check if binary exists
        if not binary_path.exists():
            raise HTTPException(
                status_code=400,
                detail="No pre-built binary found. Click 'Build Binary' first and wait for completion."
            )

        # Check if build is still running
        ps_result = subprocess.run(["pgrep", "-f", "nuitka"], capture_output=True)
        if ps_result.returncode == 0:
            raise HTTPException(
                status_code=400,
                detail="Nuitka build still in progress. Please wait for it to complete."
            )

        binary_size = binary_path.stat().st_size / (1024 * 1024)
        result_steps.append(f"Using pre-built binary: {binary_size:.1f} MB")

        # Step 1: Update VERSION file
        version_file = backend_dir / "VERSION"
        version_file.write_text(version)
        result_steps.append(f"Updated VERSION to {version}")

        # Step 2: Create deployment package
        package_dir = build_dir / "package"
        if package_dir.exists():
            shutil.rmtree(package_dir)
        package_dir.mkdir(exist_ok=True)

        shutil.copy(binary_path, package_dir / "olt-manager")
        shutil.copy(version_file, package_dir / "VERSION")
        # Note: data/ and uploads/ folders are NOT included in package
        # to prevent overwriting customer database during updates

        # Create start script
        start_script = package_dir / "start.sh"
        start_script.write_text("#!/bin/bash\ncd \"$(dirname \"$0\")\"\n./olt-manager\n")
        start_script.chmod(0o755)

        # Copy static folder (frontend files)
        static_source = Path("/opt/olt-manager/static")
        if not static_source.exists():
            static_source = Path(__file__).parent / "static"
        if static_source.exists():
            shutil.copytree(static_source, package_dir / "static")
            result_steps.append("Included frontend static files")

        # Create tarball
        package_path = Path("/tmp/olt-manager.tar.gz")
        with tarfile.open(package_path, "w:gz") as tar:
            for item in package_dir.iterdir():
                tar.add(item, arcname=item.name)

        package_size = package_path.stat().st_size / (1024 * 1024)
        result_steps.append(f"Package created: {package_size:.1f} MB (protected binary)")

        # Step 3: Upload to license server via SCP
        license_server = os.environ.get("LICENSE_SERVER", "109.110.185.101")
        license_user = os.environ.get("LICENSE_USER", "testuser")

        # Read password directly from file to avoid shell escaping issues
        license_pass_file = Path("/opt/olt-manager/.license_pass")
        if license_pass_file.exists():
            license_pass = license_pass_file.read_text().strip()
        else:
            license_pass = os.environ.get("LICENSE_PASS")
        if not license_pass:
            raise Exception("LICENSE_PASS not found in file or environment")

        import subprocess
        import tempfile

        # Create a bash script that handles the upload
        # This avoids subprocess escaping issues with special characters
        import json as json_module
        changelog_escaped = json_module.dumps(changelog)

        upload_script = f'''#!/bin/bash
set -e

# First try SSH key authentication (no password needed)
if ssh -o StrictHostKeyChecking=no -o BatchMode=yes -o ConnectTimeout=10 {license_user}@{license_server} "echo OK" >/dev/null 2>&1; then
    echo "Using SSH key authentication"
    scp -o StrictHostKeyChecking=no "{package_path}" {license_user}@{license_server}:/tmp/olt-manager.tar.gz
else
    # Fall back to password authentication
    echo "Using password authentication"
    PASS=$(cat /opt/olt-manager/.license_pass)
    sshpass -p "$PASS" scp -o StrictHostKeyChecking=no "{package_path}" {license_user}@{license_server}:/tmp/olt-manager.tar.gz
fi

# SSH commands for post-upload steps
ssh_cmd() {{
    if ssh -o StrictHostKeyChecking=no -o BatchMode=yes -o ConnectTimeout=10 {license_user}@{license_server} "echo OK" >/dev/null 2>&1; then
        ssh -o StrictHostKeyChecking=no {license_user}@{license_server} "$1"
    else
        PASS=$(cat /opt/olt-manager/.license_pass)
        sshpass -p "$PASS" ssh -o StrictHostKeyChecking=no {license_user}@{license_server} "$1"
    fi
}}

ssh_cmd "sudo cp /tmp/olt-manager.tar.gz /var/www/html/downloads/olt-manager.tar.gz && sudo chmod 644 /var/www/html/downloads/olt-manager.tar.gz"
ssh_cmd "sudo cp /tmp/olt-manager.tar.gz /opt/license-server/updates/olt-manager-{version}.tar.gz && sudo chmod 644 /opt/license-server/updates/olt-manager-{version}.tar.gz"
ssh_cmd "sudo cp /tmp/olt-manager.tar.gz /opt/license-server/updates/olt-manager.tar.gz && sudo chmod 644 /opt/license-server/updates/olt-manager.tar.gz"

# Update JSON
ssh_cmd 'sudo python3 << PYEOF
import json
from datetime import datetime
with open("/opt/license-server/updates.json", "r") as f:
    data = json.load(f)

data["latest"] = "{version}"
data["latest_version"] = "{version}"
data["download_url"] = "https://lic.proxpanel.com/downloads/olt-manager.tar.gz"
data["changelog"] = {changelog_escaped}
data["release_date"] = datetime.now().strftime("%Y-%m-%d")

new_version = {{
    "version": "{version}",
    "changelog": {changelog_escaped},
    "filename": "olt-manager-{version}.tar.gz",
    "uploaded_at": datetime.now().strftime("%Y-%m-%d")
}}
if "versions" not in data:
    data["versions"] = []
data["versions"] = [v for v in data["versions"] if v.get("version") != "{version}"]
data["versions"].insert(0, new_version)

with open("/opt/license-server/updates.json", "w") as f:
    json.dump(data, f, indent=2)
print("OK")
PYEOF'

echo "Upload completed successfully"
'''

        try:
            # Write the script to a temp file and execute it
            with tempfile.NamedTemporaryFile(mode='w', suffix='.sh', delete=False) as f:
                f.write(upload_script)
                script_path = f.name

            os.chmod(script_path, 0o755)
            result = subprocess.run(['bash', script_path], capture_output=True, text=True, timeout=300)
            os.unlink(script_path)

            if result.returncode != 0:
                error_msg = result.stderr.strip() or result.stdout.strip()
                raise Exception(f"Upload failed: {error_msg}")

            result_steps.append("Package uploaded via SCP")
            result_steps.append("Uploaded to license server successfully")

        except subprocess.TimeoutExpired:
            try:
                os.unlink(script_path)
            except:
                pass
            raise Exception("Upload timed out after 5 minutes")
        except Exception as e:
            raise Exception(f"Upload failed: {str(e)}")

        # Cleanup
        package_path.unlink()

        # Step 4: Reload SOFTWARE_VERSION
        import license_manager
        license_manager.SOFTWARE_VERSION = version
        license_manager.license_manager.update_info = None
        result_steps.append(f"Reloaded version to {version}")

        return {
            "success": True,
            "version": version,
            "message": f"Version {version} published successfully!",
            "steps": result_steps
        }

    except HTTPException:
        raise
    except Exception as e:
        import traceback
        logger.error(f"Publish failed: {e}")
        logger.error(traceback.format_exc())
        raise HTTPException(status_code=500, detail=str(e))


# Build and Publish status file
BUILD_PUBLISH_STATUS_FILE = Path("/tmp/build_publish_status.json")

@app.post("/api/dev/build-and-publish")
async def build_and_publish(
    request: PublishRequest,
    current_user: User = Depends(require_admin)
):
    """One-click: Build binary AND publish to license server (dev server only)"""
    import subprocess
    import json as json_module

    version = request.version
    changelog = request.changelog

    if not is_dev_server():
        raise HTTPException(status_code=403, detail="This feature is only available on development server")

    # Validate version before it's interpolated into a generated shell script
    # (avoids command injection via a crafted version string).
    if not re.match(r'^\d+\.\d+\.\d+([.-][A-Za-z0-9]+)?$', str(version or "")):
        raise HTTPException(status_code=400, detail="Invalid version format (expected semver)")

    # Check if already building
    ps_result = subprocess.run(["pgrep", "-f", "nuitka"], capture_output=True)
    if ps_result.returncode == 0:
        raise HTTPException(status_code=400, detail="Build already in progress. Please wait.")

    # Check syntax first
    syntax_check = subprocess.run(
        ["python3", "-m", "py_compile", "/root/olt-manager/backend/main.py"],
        capture_output=True
    )
    if syntax_check.returncode != 0:
        raise HTTPException(status_code=400, detail=f"Syntax error in code: {syntax_check.stderr.decode()}")

    # Initialize status
    status = {
        "stage": "starting",
        "progress": 0,
        "version": version,
        "changelog": changelog,
        "message": "Starting build and publish...",
        "error": None,
        "completed": False
    }
    BUILD_PUBLISH_STATUS_FILE.write_text(json_module.dumps(status))

    # Create the build and publish script
    changelog_escaped = changelog.replace("'", "'\\''")

    build_publish_script = f'''#!/bin/bash
set -e

STATUS_FILE="/tmp/build_publish_status.json"
VERSION="{version}"
CHANGELOG='{changelog_escaped}'

update_status() {{
    echo '{{"stage": "'$1'", "progress": '$2', "version": "'$VERSION'", "changelog": "'"$CHANGELOG"'", "message": "'$3'", "error": null, "completed": false}}' > "$STATUS_FILE"
}}

error_status() {{
    echo '{{"stage": "error", "progress": '$2', "version": "'$VERSION'", "changelog": "'"$CHANGELOG"'", "message": "'$3'", "error": "'$1'", "completed": true}}' > "$STATUS_FILE"
    exit 1
}}

# Step 1: Update version
update_status "version" 5 "Updating version to $VERSION..."
echo "$VERSION" > /root/olt-manager/backend/VERSION

# Step 2: Build binary
update_status "building" 10 "Building binary (this takes 30-60 minutes)..."
cd /root/olt-manager/backend
source venv/bin/activate

BUILD_DIR="/root/olt-manager/nuitka_build"
rm -rf "$BUILD_DIR" main.build main.dist main.onefile-build 2>/dev/null || true
mkdir -p "$BUILD_DIR"

python -m nuitka --standalone --onefile --output-dir="$BUILD_DIR" --output-filename=olt-manager --follow-imports --prefer-source-code main.py 2>&1 || {{
    error_status "Build failed" 50 "Nuitka build failed"
}}

if [ ! -f "$BUILD_DIR/olt-manager" ]; then
    error_status "Binary not created" 50 "Build completed but binary not found"
fi

update_status "packaging" 70 "Creating package..."

# Step 3: Create package
PACKAGE_DIR="$BUILD_DIR/package"
rm -rf "$PACKAGE_DIR" 2>/dev/null || true
mkdir -p "$PACKAGE_DIR"
mkdir -p "$PACKAGE_DIR/data"
mkdir -p "$PACKAGE_DIR/uploads"

cp "$BUILD_DIR/olt-manager" "$PACKAGE_DIR/"
cp /root/olt-manager/backend/VERSION "$PACKAGE_DIR/"

# Copy install script
cat > "$PACKAGE_DIR/install.sh" << 'INSTALLEOF'
#!/bin/bash
set -e
EXTRACT_DIR="${{1:-.}}"
if [ "$EXTRACT_DIR" != "." ] && [ -d "$EXTRACT_DIR" ]; then cd "$EXTRACT_DIR"; fi
echo "Installing OLT Manager..."
systemctl stop olt-manager 2>/dev/null || true
sleep 2
mkdir -p /opt/olt-manager/data /opt/olt-manager/uploads

# Database migration: Application uses /opt/olt-manager/data/olt_manager.db (absolute path)
DB_BACKED_UP=false
OLD_DB="/opt/olt-manager/olt_manager.db"
NEW_DB="/opt/olt-manager/data/olt_manager.db"

echo "Checking database locations..."

# If old location is a symlink, resolve it
if [ -L "$OLD_DB" ]; then
    OLD_DB_REAL=$(readlink -f "$OLD_DB")
    echo "  Old location is symlink -> $OLD_DB_REAL"
else
    OLD_DB_REAL="$OLD_DB"
fi

# Check what exists
OLD_EXISTS=false
NEW_EXISTS=false
[ -f "$OLD_DB_REAL" ] && [ ! -L "$OLD_DB" ] && OLD_EXISTS=true
[ -f "$NEW_DB" ] && NEW_EXISTS=true

if [ "$OLD_EXISTS" = true ] && [ "$NEW_EXISTS" = true ]; then
    OLD_SIZE=$(stat -c%s "$OLD_DB_REAL" 2>/dev/null || echo 0)
    NEW_SIZE=$(stat -c%s "$NEW_DB" 2>/dev/null || echo 0)
    if [ "$OLD_SIZE" -gt "$NEW_SIZE" ]; then
        cp "$OLD_DB_REAL" /tmp/olt_manager.db.backup
        echo "  Using larger database from old location"
    else
        cp "$NEW_DB" /tmp/olt_manager.db.backup
        echo "  Using larger database from new location"
    fi
    DB_BACKED_UP=true
elif [ "$OLD_EXISTS" = true ]; then
    cp "$OLD_DB_REAL" /tmp/olt_manager.db.backup
    echo "  Backed up database from old location (will migrate)"
    DB_BACKED_UP=true
elif [ "$NEW_EXISTS" = true ]; then
    cp "$NEW_DB" /tmp/olt_manager.db.backup
    echo "  Backed up database from new location"
    DB_BACKED_UP=true
fi

[ -f "/opt/olt-manager/.license_pass" ] && cp /opt/olt-manager/.license_pass /tmp/.license_pass.backup

# Install binary - ALWAYS use package binary if available
if [ -f "olt-manager" ]; then
    rm -f /opt/olt-manager/olt-manager.new 2>/dev/null || true
    cp olt-manager /opt/olt-manager/
    echo "Installed binary from package"
elif [ -f "/opt/olt-manager/olt-manager.new" ]; then
    mv /opt/olt-manager/olt-manager.new /opt/olt-manager/olt-manager
    echo "Installed binary from .new file (fallback)"
fi
chmod +x /opt/olt-manager/olt-manager
[ -f "VERSION" ] && cp VERSION /opt/olt-manager/

if [ "$DB_BACKED_UP" = true ]; then
    cp /tmp/olt_manager.db.backup /opt/olt-manager/data/olt_manager.db
    echo "Database restored to /opt/olt-manager/data/olt_manager.db"
fi

# CRITICAL: Remove old database and create symlink for compatibility
rm -f /opt/olt-manager/olt_manager.db 2>/dev/null || true
ln -sf /opt/olt-manager/data/olt_manager.db /opt/olt-manager/olt_manager.db

[ -f "/tmp/.license_pass.backup" ] && cp /tmp/.license_pass.backup /opt/olt-manager/.license_pass
systemctl daemon-reload
systemctl enable olt-manager
systemctl start olt-manager
sleep 3
echo "OLT Manager installed successfully!"
echo "Database: /opt/olt-manager/data/olt_manager.db"
INSTALLEOF
chmod +x "$PACKAGE_DIR/install.sh"

# Copy static files if exist
[ -d "/root/olt-manager/nuitka_build/main.dist/static" ] && cp -r /root/olt-manager/nuitka_build/main.dist/static "$PACKAGE_DIR/"

# Create service file
cat > "$PACKAGE_DIR/olt-manager.service" << 'SVCEOF'
[Unit]
Description=OLT Manager Backend
After=network.target
[Service]
Type=simple
WorkingDirectory=/opt/olt-manager
ExecStart=/opt/olt-manager/run.sh
Restart=always
RestartSec=5
[Install]
WantedBy=multi-user.target
SVCEOF

cd "$PACKAGE_DIR"
tar -czf /tmp/olt-manager-$VERSION.tar.gz .

update_status "uploading" 85 "Uploading to license server..."

# Step 4: Upload to license server
scp -o StrictHostKeyChecking=no /tmp/olt-manager-$VERSION.tar.gz testuser@109.110.185.101:/tmp/ || {{
    error_status "SCP upload failed" 85 "Failed to upload to license server"
}}

ssh -o StrictHostKeyChecking=no testuser@109.110.185.101 "
sudo cp /tmp/olt-manager-$VERSION.tar.gz /var/www/html/downloads/olt-manager.tar.gz
sudo cp /tmp/olt-manager-$VERSION.tar.gz /opt/license-server/updates/olt-manager-$VERSION.tar.gz
sudo cp /tmp/olt-manager-$VERSION.tar.gz /opt/license-server/updates/olt-manager.tar.gz
sudo chmod 644 /var/www/html/downloads/olt-manager.tar.gz
sudo chmod 644 /opt/license-server/updates/*.tar.gz
" || {{
    error_status "SSH commands failed" 90 "Failed to copy files on license server"
}}

update_status "updating" 95 "Updating version info..."

# Step 5: Update updates.json
ssh -o StrictHostKeyChecking=no testuser@109.110.185.101 "
sudo python3 << PYEOF
import json
from datetime import datetime
with open('/opt/license-server/updates.json', 'r') as f:
    data = json.load(f)
data['latest'] = '$VERSION'
data['latest_version'] = '$VERSION'
data['download_url'] = 'https://lic.proxpanel.com/downloads/olt-manager.tar.gz'
data['changelog'] = \"$CHANGELOG\"
data['release_date'] = datetime.now().strftime('%Y-%m-%d')
new_version = {{'version': '$VERSION', 'changelog': \"$CHANGELOG\", 'filename': 'olt-manager-$VERSION.tar.gz', 'uploaded_at': datetime.now().strftime('%Y-%m-%d')}}
if 'versions' not in data: data['versions'] = []
data['versions'] = [v for v in data['versions'] if v.get('version') != '$VERSION']
data['versions'].insert(0, new_version)
with open('/opt/license-server/updates.json', 'w') as f:
    json.dump(data, f, indent=2)
print('OK')
PYEOF
" || {{
    error_status "Failed to update version info" 95 "Could not update updates.json"
}}

# Done!
echo '{{"stage": "completed", "progress": 100, "version": "'$VERSION'", "changelog": "'"$CHANGELOG"'", "message": "Version '$VERSION' published successfully!", "error": null, "completed": true}}' > "$STATUS_FILE"
'''

    # Write script to file and execute in background
    script_path = Path("/tmp/build_and_publish.sh")
    script_path.write_text(build_publish_script)
    script_path.chmod(0o755)

    subprocess.Popen(
        ["bash", str(script_path)],
        stdout=open("/tmp/build_publish.log", "w"),
        stderr=subprocess.STDOUT,
        start_new_session=True
    )

    return {
        "success": True,
        "message": f"Build and publish started for version {version}. Check progress with /api/dev/build-publish-status"
    }


@app.get("/api/dev/build-publish-status")
async def get_build_publish_status(current_user: User = Depends(require_admin)):
    """Get status of build and publish process"""
    import json as json_module

    if not BUILD_PUBLISH_STATUS_FILE.exists():
        return {
            "stage": "idle",
            "progress": 0,
            "message": "No build in progress",
            "completed": True
        }

    try:
        status = json_module.loads(BUILD_PUBLISH_STATUS_FILE.read_text())
        return status
    except:
        return {
            "stage": "unknown",
            "progress": 0,
            "message": "Could not read status",
            "completed": True
        }


# ============ Settings API ============

@app.get("/api/settings")
def get_settings(user: Optional[User] = Depends(get_current_user), db: Session = Depends(get_db)):
    """Get all settings (public - for page name and refresh time).

    Non-sensitive settings stay readable pre-login (page name/refresh). Encrypted
    secrets (whatsapp_secret, trap_community) are NEVER returned to non-admins.
    """
    settings = db.query(Settings).all()
    # Keys that are stored encrypted
    sensitive_keys = ["whatsapp_secret", "trap_community"]
    is_admin = bool(user and getattr(user, "role", None) in ("admin", "owner"))
    result = {}
    for s in settings:
        if s.key in sensitive_keys:
            # Only admins get the decrypted secret; everyone else gets ""
            result[s.key] = decrypt_sensitive(s.value) if is_admin else ""
        else:
            result[s.key] = s.value
    # Return defaults if not set
    if "system_name" not in result:
        result["system_name"] = "OLT Manager"
    if "page_name" not in result:
        result["page_name"] = "OLT Manager Pro"
    if "refresh_interval" not in result:
        result["refresh_interval"] = "30"
    if "polling_interval" not in result:
        result["polling_interval"] = "60"
    # WhatsApp defaults
    if "whatsapp_enabled" not in result:
        result["whatsapp_enabled"] = "false"
    if "whatsapp_api_url" not in result:
        result["whatsapp_api_url"] = ""
    if "whatsapp_secret" not in result:
        result["whatsapp_secret"] = ""
    if "whatsapp_account" not in result:
        result["whatsapp_account"] = ""
    if "whatsapp_recipients" not in result:
        result["whatsapp_recipients"] = "[]"
    # SNMP Trap defaults
    if "trap_enabled" not in result:
        result["trap_enabled"] = "true"
    if "trap_port" not in result:
        result["trap_port"] = "162"
    # Timezone default
    if "timezone" not in result:
        result["timezone"] = "UTC"
    return result


@app.put("/api/settings")
def update_settings(
    data: dict,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Update settings (admin only)"""
    if current_user.role != "admin":
        raise HTTPException(status_code=403, detail="Admin access required")

    allowed_keys = ["system_name", "page_name", "refresh_interval", "polling_interval", "whatsapp_enabled",
                    "whatsapp_api_url", "whatsapp_secret", "whatsapp_account", "whatsapp_recipients",
                    "trap_enabled", "trap_port", "trap_community", "timezone"]
    # Keys that should be encrypted when stored
    sensitive_keys = ["whatsapp_secret", "trap_community"]

    for key, value in data.items():
        if key not in allowed_keys:
            continue
        # Encrypt sensitive values before storing
        store_value = encrypt_sensitive(str(value)) if key in sensitive_keys else str(value)
        setting = db.query(Settings).filter(Settings.key == key).first()
        if setting:
            setting.value = store_value
        else:
            setting = Settings(key=key, value=store_value)
            db.add(setting)
    db.commit()

    # Log event
    changed_keys = [k for k in data.keys() if k in allowed_keys]
    if changed_keys:
        log_event(db, 'settings_changed', 'system', 0, None,
                  f"Settings updated by {current_user.username}: {', '.join(changed_keys)}")

    return {"message": "Settings updated successfully"}


# ============ Alarm Settings API ============

@app.get("/api/alarm-settings")
def get_alarm_settings(db: Session = Depends(get_db)):
    """Get alarm settings (public - for alarm configuration)"""
    settings = db.query(Settings).all()
    result = {}
    for s in settings:
        if s.key.startswith("alarm_"):
            result[s.key.replace("alarm_", "")] = s.value

    # Return defaults if not set
    defaults = {
        "new_onu_registration": "true",
        "onu_offline": "true",
        "onu_back_online": "true",
        "olt_offline": "true",
        "olt_back_online": "true",
        "weak_signal": "false",
        "weak_signal_threshold": "-25",
        "weak_signal_lower_threshold": "-30",
        "high_temperature": "false",
        "high_temperature_threshold": "60",
        "selected_onus": "[]",
        "selected_regions": "[]",
        "quiet_hours_enabled": "false",
        "quiet_hours_start": "22:00",
        "quiet_hours_end": "07:00"
    }

    for key, default_value in defaults.items():
        if key not in result:
            result[key] = default_value

    # Parse JSON arrays
    try:
        result["selected_onus"] = json.loads(result["selected_onus"])
    except:
        result["selected_onus"] = []
    try:
        result["selected_regions"] = json.loads(result["selected_regions"])
    except:
        result["selected_regions"] = []

    # Convert string booleans to actual booleans
    bool_keys = ["new_onu_registration", "onu_offline", "onu_back_online",
                 "olt_offline", "olt_back_online", "weak_signal",
                 "high_temperature", "quiet_hours_enabled"]
    for key in bool_keys:
        if key in result:
            result[key] = result[key].lower() == "true"

    # Convert numeric strings to numbers
    if "weak_signal_threshold" in result:
        try:
            result["weak_signal_threshold"] = int(result["weak_signal_threshold"])
        except:
            result["weak_signal_threshold"] = -25
    if "weak_signal_lower_threshold" in result:
        try:
            result["weak_signal_lower_threshold"] = int(result["weak_signal_lower_threshold"])
        except:
            result["weak_signal_lower_threshold"] = -30
    if "high_temperature_threshold" in result:
        try:
            result["high_temperature_threshold"] = int(result["high_temperature_threshold"])
        except:
            result["high_temperature_threshold"] = 60

    return result


@app.put("/api/alarm-settings")
def update_alarm_settings(
    data: dict,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Update alarm settings (admin only)"""
    if current_user.role != "admin":
        raise HTTPException(status_code=403, detail="Admin access required")

    allowed_keys = ["new_onu_registration", "onu_offline", "onu_back_online",
                    "olt_offline", "olt_back_online", "weak_signal",
                    "weak_signal_threshold", "weak_signal_lower_threshold",
                    "high_temperature", "high_temperature_threshold",
                    "selected_onus", "selected_regions", "quiet_hours_enabled",
                    "quiet_hours_start", "quiet_hours_end"]

    for key, value in data.items():
        if key not in allowed_keys:
            continue
        # Convert lists to JSON strings for storage
        if isinstance(value, list):
            store_value = json.dumps(value)
        elif isinstance(value, bool):
            store_value = "true" if value else "false"
        else:
            store_value = str(value)

        db_key = f"alarm_{key}"
        setting = db.query(Settings).filter(Settings.key == db_key).first()
        if setting:
            setting.value = store_value
        else:
            setting = Settings(key=db_key, value=store_value)
            db.add(setting)
    db.commit()
    return {"message": "Alarm settings updated successfully"}


@app.post("/api/auth/change-password")
def change_password(
    data: dict,
    current_user: User = Depends(require_auth),
    db: Session = Depends(get_db)
):
    """Change current user's password"""
    current_password = data.get("current_password")
    new_password = data.get("new_password")

    if not new_password:
        raise HTTPException(status_code=400, detail="New password is required")

    if len(new_password) < 4:
        raise HTTPException(status_code=400, detail="New password must be at least 4 characters")

    # Skip current password check if must_change_password is set (first login)
    if not current_user.must_change_password:
        if not current_password:
            raise HTTPException(status_code=400, detail="Current password is required")
        # Verify current password
        if not bcrypt.checkpw(current_password.encode('utf-8'), current_user.password_hash.encode('utf-8')):
            raise HTTPException(status_code=400, detail="Current password is incorrect")

    # Update password and reset must_change_password flag
    new_hash = bcrypt.hashpw(new_password.encode('utf-8'), bcrypt.gensalt()).decode('utf-8')
    current_user.password_hash = new_hash
    current_user.must_change_password = False
    db.commit()

    return {"message": "Password changed successfully"}


# ============ WhatsApp Test Endpoint ============

@app.post("/api/whatsapp/test")
def test_whatsapp(
    data: dict,
    current_user: User = Depends(require_admin),
    db: Session = Depends(get_db)
):
    """Test WhatsApp notification (admin only)"""

    api_url = data.get("api_url", "").strip()
    secret = data.get("secret", "").strip()
    account = data.get("account", "").strip()
    recipient = data.get("recipient", "").strip()

    if not all([api_url, secret, account, recipient]):
        raise HTTPException(status_code=400, detail="All WhatsApp settings are required")

    try:
        # Send test message
        test_message = f"🔔 *OLT Manager Test*\n\nThis is a test notification from OLT Manager.\n\nTime: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"

        response = requests.post(
            api_url,
            data={
                'secret': secret,
                'account': account,
                'recipient': recipient,
                'type': 'text',
                'message': test_message,
                'priority': 1
            },
            timeout=30
        )

        if response.status_code == 200:
            try:
                result = response.json()
                if result.get("status") == 200:
                    return {"success": True, "message": "Test message sent successfully!"}
                else:
                    return {"success": False, "message": f"API returned: {result.get('message', 'Unknown error')}"}
            except:
                return {"success": True, "message": "Message sent (response received)"}
        else:
            return {"success": False, "message": f"HTTP Error {response.status_code}: {response.text[:200]}"}

    except requests.exceptions.Timeout:
        raise HTTPException(status_code=504, detail="Request timed out - check API URL")
    except requests.exceptions.ConnectionError:
        raise HTTPException(status_code=502, detail="Connection failed - check API URL")
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error: {str(e)}")


# ============ OLT Port Endpoints ============


def get_pon_port_count(model: str) -> int:
    """Return the PON port count for an OLT model.

    Resolved from the driver registry — adding a new OLT model only requires
    creating a driver class. Falls back to 8 for unknown models.
    """
    try:
        return get_driver_class(model).PON_COUNT
    except ValueError:
        return 8  # Safe default for unknown models


# Store previous port counters for rate calculation
_port_counters_cache = {}

def poll_port_traffic_snmp(ip: str, community: str = 'public') -> dict:
    """Poll port traffic counters from OLT via SNMP.
    Returns dict with interface index -> {'rx_bytes': int, 'tx_bytes': int}

    Uses 64-bit OIDs (HC = High Capacity) to avoid counter wrap on high-traffic ports:
    - 1.3.6.1.2.1.31.1.1.1.6 (ifHCInOctets) - 64-bit input bytes
    - 1.3.6.1.2.1.31.1.1.1.10 (ifHCOutOctets) - 64-bit output bytes

    Falls back to 32-bit counters if 64-bit not available:
    - 1.3.6.1.2.1.2.2.1.10 (ifInOctets) - 32-bit input bytes
    - 1.3.6.1.2.1.2.2.1.16 (ifOutOctets) - 32-bit output bytes
    """
    import subprocess
    import time
    import re

    port_traffic = {}

    def _walk(oid: str) -> str:
        # snmpbulkwalk is far faster than snmpwalk on an OLT with 100+ ifIndexes
        # (the plain snmpwalk kept hitting the 10s timeout -> empty result ->
        # broken rate baseline). -On keeps the numeric OID so we key by the REAL
        # ifIndex instead of fragile line-order.
        r = subprocess.run(
            ['snmpbulkwalk', '-v2c', '-c', community, '-On', ip, oid, '-Cr50', '-t', '10'],
            capture_output=True, text=True, timeout=60
        )
        return r.stdout if r.returncode == 0 else ''

    def _parse(raw: str, col: str) -> dict:
        out = {}
        for line in raw.split('\n'):
            m = re.search(re.escape(col) + r'(\d+)\s*=\s*Counter\d+:\s*(\d+)', line)
            if m:
                out[int(m.group(1))] = int(m.group(2))
        return out

    try:
        from concurrent.futures import ThreadPoolExecutor
        # 64-bit first (ifHCInOctets/ifHCOutOctets), else 32-bit (ifIn/OutOctets).
        # Run the in+out walks in parallel to roughly halve the poll time.
        with ThreadPoolExecutor(max_workers=2) as ex:
            fin = ex.submit(_walk, '1.3.6.1.2.1.31.1.1.1.6')
            fout = ex.submit(_walk, '1.3.6.1.2.1.31.1.1.1.10')
            in_raw, out_raw = fin.result(), fout.result()
        col_in, col_out = '.31.1.1.1.6.', '.31.1.1.1.10.'
        if 'Counter64' not in in_raw:
            with ThreadPoolExecutor(max_workers=2) as ex:
                fin = ex.submit(_walk, '1.3.6.1.2.1.2.2.1.10')
                fout = ex.submit(_walk, '1.3.6.1.2.1.2.2.1.16')
                in_raw, out_raw = fin.result(), fout.result()
            col_in, col_out = '.2.2.1.10.', '.2.2.1.16.'

        in_by_idx = _parse(in_raw, col_in)
        out_by_idx = _parse(out_raw, col_out)
        timestamp = time.time()

        # rx_bytes = ifInOctets (customer download from the uplink's view),
        # tx_bytes = ifOutOctets (upload). Keyed by the real ifIndex.
        for idx, in_bytes in in_by_idx.items():
            if idx in out_by_idx:
                port_traffic[idx] = {
                    'rx_bytes': in_bytes,
                    'tx_bytes': out_by_idx[idx],
                    'timestamp': timestamp,
                }
    except Exception as e:
        logger.warning(f"Failed to poll port traffic from {ip}: {e}")

    return port_traffic


def calculate_port_rates(olt_id: int, ip: str, current_counters: dict) -> dict:
    """Calculate port traffic rates (kbps) from counter differences."""
    global _port_counters_cache

    cache_key = f"{olt_id}_{ip}"
    rates = {}

    if cache_key in _port_counters_cache:
        prev = _port_counters_cache[cache_key]
        prev_time = prev.get('timestamp', 0)
        curr_time = current_counters.get(1, {}).get('timestamp', 0)

        time_diff = curr_time - prev_time
        # Require a real poll interval. Sub-second gaps happen when two poll
        # cycles overlap and share this global cache — dividing a byte delta by a
        # near-zero time produced the absurd (multi-Tbps) / spiky port rates.
        if time_diff >= 2:
            for if_idx, curr in current_counters.items():
                if if_idx in prev and isinstance(prev.get(if_idx), dict):
                    prev_rx = prev[if_idx].get('rx_bytes', 0)
                    prev_tx = prev[if_idx].get('tx_bytes', 0)
                    curr_rx = curr.get('rx_bytes', 0)
                    curr_tx = curr.get('tx_bytes', 0)

                    # Counter reset/wrap — discard sample (prevents negatives)
                    if curr_rx < prev_rx or curr_tx < prev_tx:
                        continue

                    # Calculate rates in kbps (bytes * 8 / 1000 / seconds)
                    rx_kbps = ((curr_rx - prev_rx) * 8) / (1000 * time_diff)
                    tx_kbps = ((curr_tx - prev_tx) * 8) / (1000 * time_diff)

                    # 10 Gbps cap for uplink ports (skip implausible spikes)
                    if rx_kbps < 0 or tx_kbps < 0 or rx_kbps > 10_000_000 or tx_kbps > 10_000_000:
                        continue

                    rates[if_idx] = {
                        'rx_kbps': round(rx_kbps, 2),
                        'tx_kbps': round(tx_kbps, 2)
                    }

    # Store current counters for next calculation — but NEVER overwrite a good
    # baseline with an empty/failed poll (that would break the next cycle's rate).
    if current_counters:
        ts = 0
        for _v in current_counters.values():
            if isinstance(_v, dict) and _v.get('timestamp'):
                ts = _v['timestamp']
                break
        _port_counters_cache[cache_key] = current_counters
        _port_counters_cache[cache_key]['timestamp'] = ts

    return rates


def poll_port_status_snmp(ip: str, community: str = 'public', model: str = None) -> dict:
    """Poll port status from OLT via SNMP.
    Returns dict with interface index -> {'status': 'up'/'down', 'descr': '...', 'name': '...'}

    V1600D8 interface mapping (from SNMP):
    - ifIndex 1-16: GE uplink ports (GE0/1 to GE0/16)
    - ifIndex 17-24: PON ports (EPON0/1 to EPON0/8)
    - ifIndex 25: Management interface
    - ifIndex 26+: ONU virtual interfaces

    V1600G2-B interface mapping (from SNMP):
    - ifIndex 1-8: GE uplink ports (GE0/1 to GE0/8)
    - ifIndex 9-24: PON ports (GPON0/1 to GPON0/16)
    - ifIndex 25+: ONU virtual interfaces
    """
    import subprocess
    import re

    port_info = {}

    # The OID used to read interface labels comes from the model's driver:
    # V1600D8 uses ifName, V1600G2-B uses ifDescr. Adding a new model is just
    # a matter of declaring ``PORT_NAME_OID`` on its driver class.
    try:
        name_oid = get_driver_class(model).PORT_NAME_OID
    except ValueError:
        name_oid = '1.3.6.1.2.1.2.2.1.2'  # ifDescr default

    try:
        from concurrent.futures import ThreadPoolExecutor, as_completed
        names = {}
        statuses = {}

        def get_port_info(idx):
            """Get ifDescr/ifName and ifOperStatus for a single port"""
            name = None
            status = '2'  # Default down
            try:
                # Get port name (ifName for V1600D8, ifDescr for others)
                result = subprocess.run(
                    ['snmpget', '-v2c', '-c', community, '-t', '3', ip, f'{name_oid}.{idx}'],
                    capture_output=True, text=True, timeout=5
                )
                if result.stdout and 'STRING:' in result.stdout:
                    match = re.search(r'STRING:\s*"?([^"]*)"?', result.stdout)
                    if match:
                        name = match.group(1).strip()
                # Get ifOperStatus
                result = subprocess.run(
                    ['snmpget', '-v2c', '-c', community, '-t', '3', ip, f'1.3.6.1.2.1.2.2.1.8.{idx}'],
                    capture_output=True, text=True, timeout=5
                )
                if result.stdout and 'INTEGER:' in result.stdout:
                    match = re.search(r'INTEGER:\s*(\d+)', result.stdout)
                    if match:
                        status = match.group(1)
            except:
                pass
            return idx, name, status

        # Poll first 24 ports in parallel (8 uplinks + 16 PON ports)
        with ThreadPoolExecutor(max_workers=12) as executor:
            futures = {executor.submit(get_port_info, i): i for i in range(1, 25)}
            try:
                for future in as_completed(futures, timeout=30):
                    try:
                        idx, name, status = future.result(timeout=5)
                        if name:
                            names[idx] = name
                        statuses[idx] = status
                    except:
                        pass
            except TimeoutError:
                # Collect any completed results even if some timed out
                for future, idx in futures.items():
                    if future.done():
                        try:
                            idx, name, status = future.result(timeout=0)
                            if name:
                                names[idx] = name
                            statuses[idx] = status
                        except:
                            pass

        # Combine results - only process first 30 interfaces (uplink + PON ports)
        for i in range(1, 31):
            name = names.get(i, f"IF{i}")
            status_val = statuses.get(i, '2')  # Default to down if not found

            # Parse interface name - format is "GE0/7 MIKRO" or "GE0/1 "
            parts = name.split(' ', 1)
            if_name = parts[0] if parts else f"IF{i}"
            descr = parts[1].strip() if len(parts) > 1 and parts[1].strip() else None

            port_info[i] = {
                'status': 'up' if status_val == '1' else 'down',
                'name': if_name,
                'descr': descr
            }
    except Exception as e:
        logger.warning(f"Failed to poll port status from {ip}: {e}")

    return port_info


# Cache for port status to avoid repeated SNMP timeouts
_port_status_cache: Dict[str, Dict] = {}  # ip -> {'data': {...}, 'timestamp': datetime}
_port_status_updating: Dict[str, bool] = {}  # ip -> True if update in progress
PORT_STATUS_CACHE_SECONDS = 300  # Cache for 5 minutes (background polling updates it)

def _update_port_status_background(ip: str, model: str = None):
    """Background task to update port status cache"""
    try:
        if _port_status_updating.get(ip):
            return  # Already updating
        _port_status_updating[ip] = True
        data = poll_port_status_snmp(ip, model=model)
        if data:
            _port_status_cache[ip] = {'data': data, 'timestamp': datetime.now()}
    except Exception as e:
        logger.warning(f"Background port status update failed for {ip}: {e}")
    finally:
        _port_status_updating[ip] = False

@app.get("/api/olts/{olt_id}/ports")
def get_olt_ports(olt_id: int, user: User = Depends(require_auth), db: Session = Depends(get_db)):
    """
    Get all ports for an OLT with their status.
    Returns PON ports with ONU counts and SFP uplink ports.
    Returns cached data instantly, updates in background if stale.
    """
    from models import OLTPort

    olt = db.query(OLT).filter(OLT.id == olt_id).first()
    if not olt:
        raise HTTPException(status_code=404, detail="OLT not found")

    snmp_ports = {}
    model_upper = (olt.model or '').upper()
    cache_key = olt.ip_address
    now = datetime.now()

    # Always use cached data if available (instant response)
    if cache_key in _port_status_cache:
        cached = _port_status_cache[cache_key]
        snmp_ports = cached['data']
        age = (now - cached['timestamp']).total_seconds()
        # Trigger background update if cache is stale (>60s) but still return cached data
        if age > 60 and not _port_status_updating.get(cache_key):
            thread_executor.submit(_update_port_status_background, cache_key, olt.model)

    # If no cache at all, do a quick poll (only for first request)
    if not snmp_ports:
        snmp_ports = poll_port_status_snmp(olt.ip_address, model=olt.model)
        _port_status_cache[cache_key] = {'data': snmp_ports, 'timestamp': now}

    # Get PON port count based on model
    pon_count = get_pon_port_count(olt.model)

    # Get existing port data from database
    port_data = db.query(OLTPort).filter(OLTPort.olt_id == olt_id).all()
    port_map = {(p.port_type, p.port_number): p for p in port_data}

    # Get ONU counts per PON port from actual ONUs
    onu_counts = {}
    onus = db.query(ONU).filter(ONU.olt_id == olt_id).all()
    for onu in onus:
        if onu.pon_port not in onu_counts:
            onu_counts[onu.pon_port] = {'total': 0, 'online': 0}
        onu_counts[onu.pon_port]['total'] += 1
        if onu.is_online:
            onu_counts[onu.pon_port]['online'] += 1

    # Build PON ports list
    pon_ports = []
    for i in range(1, pon_count + 1):
        port = port_map.get(('pon', i))
        counts = onu_counts.get(i, {'total': 0, 'online': 0})

        # Determine status based on ONUs or port data
        if counts['online'] > 0:
            status = 'up'
        elif port and port.status:
            status = port.status
        else:
            status = 'unknown'

        pon_ports.append({
            "port_number": i,
            "status": status,
            "onu_count": counts['total'],
            "onu_online": counts['online'],
            "tx_power": port.tx_power if port else None,
            "rx_power": port.rx_power if port else None,
            "temperature": port.temperature if port else None  # PON transceiver temperature
        })

    # Uplink port layout comes from the OLT driver. Adding a new model = one
    # new driver class with its ``get_port_layout()`` implementation; nothing
    # in this file needs to change.
    try:
        layout = get_driver(olt).get_port_layout()
        ge_config = list(layout.ge_ports)
        sfp_config = list(layout.sfp_ports)
        xge_config = list(layout.sfp_plus_ports)
        qsfp_config = list(layout.qsfp_ports)
    except ValueError:
        # Unknown model — fall back to a generic 2 GE + 2 SFP layout so the
        # dashboard still renders something useful.
        ge_config = [(1, 'GE1', '1G'), (2, 'GE2', '1G')]
        sfp_config = [(3, 'SFP1', '1G'), (4, 'SFP2', '1G')]
        xge_config = []
        qsfp_config = []

    # Build GE RJ45 ports with live SNMP status (fallback to database if no SNMP data)
    ge_ports = []
    for if_idx, default_label, default_speed in ge_config:
        port = port_map.get(('ge', if_idx))
        snmp_info = snmp_ports.get(if_idx, {})
        descr = snmp_info.get('descr')
        # Use database status as fallback when no SNMP data available
        db_status = port.status if port else 'down'
        snmp_status = snmp_info.get('status') if snmp_info else None
        # Port status from SNMP takes priority, fallback to database if no SNMP data
        if snmp_status == 'up':
            status = 'up'
        elif snmp_status == 'down':
            status = 'down'
        elif snmp_status is None and db_status == 'up':
            status = 'up'  # Keep last known state when SNMP unavailable
        else:
            status = 'down'
        ge_ports.append({
            "port_number": if_idx,
            "type": "ge",
            "status": status,
            "speed": port.speed if port else default_speed,
            "label": descr if descr else default_label
        })

    # Build SFP ports with live SNMP status (fallback to database if no SNMP data)
    sfp_ports = []
    for if_idx, default_label, default_speed in sfp_config:
        port = port_map.get(('sfp', if_idx))
        snmp_info = snmp_ports.get(if_idx, {})
        descr = snmp_info.get('descr')
        db_status = port.status if port else 'down'
        snmp_status = snmp_info.get('status') if snmp_info else None
        # Port status from SNMP takes priority, fallback to database if no SNMP data
        if snmp_status == 'up':
            status = 'up'
        elif snmp_status == 'down':
            status = 'down'
        elif snmp_status is None and db_status == 'up':
            status = 'up'
        else:
            status = 'down'
        sfp_ports.append({
            "port_number": if_idx,
            "type": "sfp",
            "status": status,
            "speed": port.speed if port else default_speed,
            "label": descr if descr else default_label
        })

    # Build 10G SFP+ ports with live SNMP status (fallback to database if no SNMP data)
    xge_ports = []
    for if_idx, default_label, default_speed in xge_config:
        port = port_map.get(('xge', if_idx))
        snmp_info = snmp_ports.get(if_idx, {})
        descr = snmp_info.get('descr')
        db_status = port.status if port else 'down'
        snmp_status = snmp_info.get('status') if snmp_info else None
        # Port status from SNMP takes priority, fallback to database if no SNMP data
        if snmp_status == 'up':
            status = 'up'
        elif snmp_status == 'down':
            status = 'down'
        elif snmp_status is None and db_status == 'up':
            status = 'up'
        else:
            status = 'down'
        xge_ports.append({
            "port_number": if_idx,
            "type": "xge",
            "status": status,
            "speed": port.speed if port else default_speed,
            "label": descr if descr else default_label
        })

    # Build QSFP28 40G/100G ports with live SNMP status (fallback to database if no SNMP data)
    qsfp_ports = []
    for if_idx, default_label, default_speed in qsfp_config:
        port = port_map.get(('qsfp', if_idx))
        snmp_info = snmp_ports.get(if_idx, {})
        descr = snmp_info.get('descr')
        db_status = port.status if port else 'down'
        snmp_status = snmp_info.get('status') if snmp_info else None
        # Port status from SNMP takes priority, fallback to database if no SNMP data
        if snmp_status == 'up':
            status = 'up'
        elif snmp_status == 'down':
            status = 'down'
        elif snmp_status is None and db_status == 'up':
            status = 'up'
        else:
            status = 'down'
        qsfp_ports.append({
            "port_number": if_idx,
            "type": "qsfp",
            "status": status,
            "speed": port.speed if port else default_speed,
            "label": descr if descr else default_label
        })

    # Save port status to database for persistence (only when SNMP data is available)
    if snmp_ports:
        try:
            all_ports = ge_ports + sfp_ports + xge_ports + qsfp_ports
            for p in all_ports:
                port_key = (p['type'], p['port_number'])
                existing = port_map.get(port_key)
                if existing:
                    if existing.status != p['status']:
                        existing.status = p['status']
                        existing.last_updated = datetime.utcnow()
                else:
                    new_port = OLTPort(
                        olt_id=olt_id,
                        port_type=p['type'],
                        port_number=p['port_number'],
                        status=p['status'],
                        speed=p['speed'],
                        last_updated=datetime.utcnow()
                    )
                    db.add(new_port)
            db.commit()
        except Exception as e:
            logger.warning(f"Failed to save port status to DB: {e}")
            db.rollback()

    return {
        "olt_id": olt_id,
        "olt_name": olt.name,
        "model": olt.model,
        "ip_address": olt.ip_address,
        "is_online": olt.is_online,
        "total_pon": pon_count,
        "pon_ports": pon_ports,
        "ge_ports": ge_ports,
        "sfp_ports": sfp_ports,
        "xge_ports": xge_ports,
        "qsfp_ports": qsfp_ports,  # QSFP28 40G/100G ports for V3600 series
        "total_onus": sum(c['total'] for c in onu_counts.values()),
        "online_onus": sum(c['online'] for c in onu_counts.values())
    }


def set_port_description_snmp(ip: str, if_index: int, description: str, community: str = 'private') -> bool:
    """Set interface description on OLT via SNMP.

    Uses OID 1.3.6.1.2.1.31.1.1.1.18 (ifAlias) for interface description.
    """
    import subprocess

    try:
        # ifAlias OID for setting interface description
        oid = f'1.3.6.1.2.1.31.1.1.1.18.{if_index}'

        result = subprocess.run(
            ['snmpset', '-v2c', '-c', community, ip, oid, 's', description],
            capture_output=True, text=True, timeout=10
        )

        if result.returncode == 0:
            logger.info(f"Set interface {if_index} description to '{description}' on {ip}")
            return True
        else:
            logger.error(f"Failed to set interface description: {result.stderr}")
            return False

    except Exception as e:
        logger.error(f"SNMP set failed for {ip}: {e}")
        return False


@app.put("/api/olts/{olt_id}/ports/{port_number}/description")
async def set_port_description(
    olt_id: int,
    port_number: int,
    description: str = Query(..., description="New port description"),
    user: User = Depends(require_auth),
    db: Session = Depends(get_db)
):
    """Set port description on OLT via web interface.

    Uses web scraping to set port description on gecfg.html page.
    This updates the ifDescr which is read by the dashboard.
    """
    from olt_web_scraper import set_port_description_web

    olt = db.query(OLT).filter(OLT.id == olt_id).first()
    if not olt:
        raise HTTPException(status_code=404, detail="OLT not found")

    # Set description via web interface
    # port_number is the GE port number (1-8 for V1600G2-B, 1-16 for V1600D8)
    loop = asyncio.get_event_loop()
    success = await loop.run_in_executor(
        thread_executor,
        set_port_description_web,
        olt.ip_address,
        port_number,
        description,
        "admin",  # username
        "admin",  # password
        olt.model
    )

    if success:
        # Clear port status cache so next poll picks up new description
        if olt.ip_address in _port_status_cache:
            del _port_status_cache[olt.ip_address]
        return {"success": True, "message": f"Port {port_number} description set to '{description}'"}
    else:
        raise HTTPException(status_code=500, detail="Failed to set port description")


@app.get("/api/olts/{olt_id}/ports/{port_type}/{port_number}/traffic")
def get_port_traffic_history(
    olt_id: int,
    port_type: str,
    port_number: int,
    range: str = Query('1h', description="Time range: 5m, 15m, 30m, 1h, 6h, 24h, 1w, 1M"),
    user: User = Depends(require_auth),
    db: Session = Depends(get_db)
):
    """
    Get traffic history for a specific port.
    Used for per-port traffic graphs.
    """
    from models import PortTraffic
    from datetime import timedelta

    # Time range mapping
    time_ranges = {
        '5m': timedelta(minutes=5),
        '15m': timedelta(minutes=15),
        '30m': timedelta(minutes=30),
        '1h': timedelta(hours=1),
        '6h': timedelta(hours=6),
        '24h': timedelta(hours=24),
        '1w': timedelta(weeks=1),
        '1M': timedelta(days=30)
    }

    if range not in time_ranges:
        raise HTTPException(status_code=400, detail=f"Invalid range. Use: {', '.join(time_ranges.keys())}")

    olt = db.query(OLT).filter(OLT.id == olt_id).first()
    if not olt:
        raise HTTPException(status_code=404, detail="OLT not found")

    time_delta = time_ranges[range]
    since = datetime.utcnow() - time_delta

    traffic = db.query(PortTraffic).filter(
        PortTraffic.olt_id == olt_id,
        PortTraffic.port_type == port_type,
        PortTraffic.port_number == port_number,
        PortTraffic.timestamp > since
    ).order_by(PortTraffic.timestamp).all()

    # If no per-port traffic data, fall back to TrafficHistory
    if not traffic:
        # For PON, GE, XGE ports - check TrafficHistory
        if port_type in ('pon', 'ge', 'xge'):
            if port_type == 'pon':
                # PON uses pon_port field
                history_traffic = db.query(TrafficHistory).filter(
                    TrafficHistory.olt_id == olt_id,
                    TrafficHistory.entity_type == 'pon',
                    TrafficHistory.pon_port == port_number,
                    TrafficHistory.timestamp > since
                ).order_by(TrafficHistory.timestamp).all()
            else:
                # GE/XGE uses entity_id format "olt_id:port_type:port_num"
                entity_id = f"{olt_id}:{port_type}:{port_number}"
                history_traffic = db.query(TrafficHistory).filter(
                    TrafficHistory.olt_id == olt_id,
                    TrafficHistory.entity_type == port_type,
                    TrafficHistory.entity_id == entity_id,
                    TrafficHistory.timestamp > since
                ).order_by(TrafficHistory.timestamp).all()

            if history_traffic:
                return {
                    "olt_id": olt_id,
                    "port_type": port_type,
                    "port_number": port_number,
                    "range": range,
                    "data": [
                        {
                            "timestamp": t.timestamp.isoformat() + "Z",
                            "rx_kbps": t.rx_kbps,
                            "tx_kbps": t.tx_kbps
                        }
                        for t in history_traffic
                    ]
                }

    return {
        "olt_id": olt_id,
        "port_type": port_type,
        "port_number": port_number,
        "range": range,
        "data": [
            {
                "timestamp": t.timestamp.isoformat() + "Z",
                "rx_kbps": t.rx_kbps,
                "tx_kbps": t.tx_kbps
            }
            for t in traffic
        ]
    }


# ============ Traffic Monitoring Endpoints ============

@app.get("/api/olts/{olt_id}/traffic")
async def get_olt_traffic(olt_id: int, user: User = Depends(require_auth), db: Session = Depends(get_db)):
    """
    Get live traffic data for all ONUs on an OLT.

    Returns current bandwidth in Kbps for each ONU.
    Polls SNMP counters and calculates rate based on previous snapshot.
    OLT updates counters approximately every 30 seconds.
    """
    olt = db.query(OLT).filter(OLT.id == olt_id).first()
    if not olt:
        raise HTTPException(status_code=404, detail="OLT not found")

    try:
        # Serve the CACHED per-ONU rates that the background poll cycle maintains
        # (updated every ~30s), instead of doing a slow synchronous SNMP re-poll
        # on every request (which hung 30s+). Same data the live WebSocket serves.
        current_time = datetime.utcnow()
        snapshots = db.query(TrafficSnapshot).filter(TrafficSnapshot.olt_id == olt_id).all()
        onus = {o.mac_address: o for o in db.query(ONU).filter(ONU.olt_id == olt_id).all()}
        traffic_data = []
        for s in snapshots:
            onu = onus.get(s.mac_address)
            onu_online = bool(onu and onu.is_online)
            rx = (s.last_rx_kbps or 0) if onu_online else 0   # offline ONUs read 0
            tx = (s.last_tx_kbps or 0) if onu_online else 0
            traffic_data.append({
                "mac_address": s.mac_address,
                "pon_port": onu.pon_port if onu else 0,
                "onu_id": onu.onu_id if onu else 0,
                "description": onu.description if onu else None,
                "is_online": onu_online,
                "rx_kbps": rx,
                "tx_kbps": tx,
                "rx_mbps": round(rx / 1000, 2),
                "tx_mbps": round(tx / 1000, 2),
            })
        traffic_data.sort(key=lambda x: (not x.get('is_online', False), -(x['rx_kbps'] + x['tx_kbps'])))
        return {
            "olt_id": olt_id,
            "olt_name": olt.name,
            "timestamp": current_time.isoformat(),
            "onu_count": len(traffic_data),
            "traffic": traffic_data
        }
    except Exception as e:
        logger.error(f"Traffic read failed for OLT {olt_id}: {e}")
        raise HTTPException(status_code=500, detail=f"Traffic read failed: {str(e)}")


@app.get("/api/traffic/all")
async def get_all_traffic(user: User = Depends(require_auth), db: Session = Depends(get_db)):
    """
    Get live traffic for all ONUs across all OLTs the user can access.
    Returns aggregated bandwidth data.
    """
    allowed_olt_ids = get_user_allowed_olt_ids(user, db)

    # Build OLT query
    olt_query = db.query(OLT)
    if allowed_olt_ids is not None:
        olt_query = olt_query.filter(OLT.id.in_(allowed_olt_ids))

    olts = olt_query.filter(OLT.is_online == True).all()

    all_traffic = []
    current_time = datetime.utcnow()

    # Serve the CACHED per-ONU rates the background poll maintains (updated every
    # ~30s) instead of live-polling every OLT synchronously (which hung 30s+).
    olt_ids = [o.id for o in olts]
    olt_names = {o.id: o.name for o in olts}
    if olt_ids:
        snaps = db.query(TrafficSnapshot).filter(TrafficSnapshot.olt_id.in_(olt_ids)).all()
        onus = {(o.olt_id, o.mac_address): o
                for o in db.query(ONU).filter(ONU.olt_id.in_(olt_ids)).all()}
        for sn in snaps:
            onu = onus.get((sn.olt_id, sn.mac_address))
            onu_online = bool(onu and onu.is_online)
            rx = (sn.last_rx_kbps or 0) if onu_online else 0   # offline ONUs read 0
            tx = (sn.last_tx_kbps or 0) if onu_online else 0
            all_traffic.append({
                "olt_id": sn.olt_id,
                "olt_name": olt_names.get(sn.olt_id, ""),
                "mac_address": sn.mac_address,
                "pon_port": onu.pon_port if onu else 0,
                "onu_id": onu.onu_id if onu else 0,
                "description": onu.description if onu else None,
                "is_online": onu_online,
                "rx_kbps": rx,
                "tx_kbps": tx,
                "rx_mbps": round(rx / 1000, 2),
                "tx_mbps": round(tx / 1000, 2),
            })

    # Sort: online first (by traffic descending), then offline at bottom
    all_traffic.sort(key=lambda x: (not x.get('is_online', False), -(x['rx_kbps'] + x['tx_kbps'])))

    return {
        "timestamp": current_time.isoformat(),
        "onu_count": len(all_traffic),
        "traffic": all_traffic
    }


# ============ WebSocket Live Traffic ============

async def traffic_polling_loop(olt_id: int, olt_ip: str, db_session_factory):
    """Background loop to poll traffic and broadcast to WebSocket clients"""
    logger.info(f"Started traffic polling loop for OLT {olt_id}")

    # Small delay to let WebSocket connection stabilize before sending large payload
    # This prevents disconnection on OLTs with many ONUs (large initial payload)
    await asyncio.sleep(0.5)

    # Send immediate cached traffic data from database (updated by background polling)
    # This gives instant feedback while SNMP poll runs in background
    db = db_session_factory()
    try:
        olt = db.query(OLT).filter(OLT.id == olt_id).first()
        olt_name = olt.name if olt else f"OLT-{olt_id}"

        # Get cached traffic data from snapshots
        snapshots = db.query(TrafficSnapshot).filter(TrafficSnapshot.olt_id == olt_id).all()
        cached_traffic = []
        for s in snapshots:
            onu = db.query(ONU).filter(ONU.olt_id == olt_id, ONU.mac_address == s.mac_address).first()
            # Only online ONUs — offline ones with a stale snapshot must not appear active.
            if onu and onu.is_online:
                cached_traffic.append({
                    "mac_address": s.mac_address,
                    "pon_port": onu.pon_port if onu else 0,
                    "onu_id": onu.onu_id if onu else 0,
                    "description": onu.description if onu else None,
                    "is_online": onu.is_online if onu else False,
                    "rx_kbps": s.last_rx_kbps or 0,
                    "tx_kbps": s.last_tx_kbps or 0,
                    "rx_mbps": round((s.last_rx_kbps or 0) / 1000, 2),
                    "tx_mbps": round((s.last_tx_kbps or 0) / 1000, 2)
                })
        # Sort: online first (by traffic descending), then offline at bottom
        cached_traffic.sort(key=lambda x: (not x.get('is_online', False), -(x['rx_kbps'] + x['tx_kbps'])))

        await traffic_manager.broadcast(olt_id, {
            "olt_id": olt_id,
            "olt_name": olt_name,
            "timestamp": datetime.now().isoformat(),
            "onu_count": len(cached_traffic),
            "traffic": cached_traffic,
            "message": "Cached data (updating...)"
        })
    except Exception as e:
        logger.warning(f"Failed to send cached traffic: {e}")
    finally:
        db.close()

    while olt_id in traffic_manager.active_connections and traffic_manager.active_connections[olt_id]:
        try:
            poll_start = datetime.now()
            db = db_session_factory()
            try:
                # Get current traffic counters from SNMP with keep-alive pings
                loop = asyncio.get_event_loop()
                snmp_task = loop.run_in_executor(
                    thread_executor,
                    get_traffic_counters_snmp,
                    olt_ip,
                    "public"
                )

                # Send keep-alive pings every 5 seconds while SNMP poll runs
                while not snmp_task.done():
                    try:
                        current_counters = await asyncio.wait_for(asyncio.shield(snmp_task), timeout=5.0)
                        break  # Poll completed
                    except asyncio.TimeoutError:
                        # Still polling - send keep-alive to prevent WebSocket timeout
                        elapsed = int((datetime.now() - poll_start).total_seconds())
                        await traffic_manager.broadcast(olt_id, {
                            "olt_id": olt_id,
                            "timestamp": datetime.now().isoformat(),
                            "polling": True,
                            "elapsed_seconds": elapsed,
                            "message": f"Polling SNMP ({elapsed}s)..."
                        })

                current_counters = await snmp_task

                if not current_counters:
                    await traffic_manager.broadcast(olt_id, {
                        "olt_id": olt_id,
                        "timestamp": datetime.now().isoformat(),
                        "traffic": [],
                        "message": "No traffic data"
                    })
                    await asyncio.sleep(1.5)
                    continue

                current_time = datetime.utcnow()

                # Get previous snapshots for this OLT
                prev_snapshots = {
                    s.mac_address: s
                    for s in db.query(TrafficSnapshot).filter(TrafficSnapshot.olt_id == olt_id).all()
                }

                # Get OLT name
                olt = db.query(OLT).filter(OLT.id == olt_id).first()
                olt_name = olt.name if olt else f"OLT-{olt_id}"

                # Get Mikrotik traffic rates if configured
                mk_rates = {}
                if olt and getattr(olt, 'mk_enabled', False) and getattr(olt, 'mk_ip', None):
                    try:
                        from mikrotik_traffic import get_mikrotik_traffic
                        onu_db_map = {}
                        for onu in db.query(ONU).filter(ONU.olt_id == olt_id).all():
                            onu_db_map[(onu.pon_port, onu.onu_id)] = onu.mac_address
                        mk_rates = await loop.run_in_executor(
                            thread_executor,
                            lambda: get_mikrotik_traffic(
                                mk_ip=olt.mk_ip,
                                mk_user=olt.mk_username or 'admin',
                                mk_pass=olt.mk_password or '',
                                mk_port=olt.mk_port or 8728,
                                olt_ip=olt.ip_address,
                                snmp_community=olt.snmp_community or 'public',
                                onu_db_map=onu_db_map,
                            )
                        )
                    except Exception as exc:
                        logger.warning(f"Mikrotik traffic (WS) failed: {exc}")

                traffic_data = []

                # No EMA smoothing — use exact instant rate.
                # OLT counters update every ~30s so the rate is already a
                # 30s average; smoothing just makes the display lag behind.

                for key, counters in current_counters.items():
                    rx_bytes = counters['rx_bytes']
                    tx_bytes = counters['tx_bytes']
                    pon_port = counters.get('pon_port', 0)
                    onu_id = counters.get('onu_id', 0)

                    # Handle both MAC and pon:onu key formats
                    if ':' in key and len(key) < 10:  # pon:onu format
                        onu_for_mac = db.query(ONU).filter(
                            ONU.olt_id == olt_id,
                            ONU.pon_port == pon_port,
                            ONU.onu_id == onu_id
                        ).first()
                        mac = onu_for_mac.mac_address if onu_for_mac else None
                        if not mac:
                            continue
                    else:
                        mac = key

                    rx_kbps = 0
                    tx_kbps = 0

                    # READ-ONLY: the background poll cycle (collect_traffic_history)
                    # is the sole writer of TrafficSnapshot rows. This WS loop only
                    # COMPUTES a live display rate from the latest snapshot and never
                    # writes it back — this removes the dual-writer race that caused
                    # jittery/oscillating rates when a dashboard was open during a poll.
                    if mac in prev_snapshots:
                        prev = prev_snapshots[mac]
                        prev_rx = getattr(prev, 'last_rx_kbps', 0) or 0
                        prev_tx = getattr(prev, 'last_tx_kbps', 0) or 0
                        time_diff = (current_time - prev.timestamp).total_seconds()

                        if time_diff <= 0 or time_diff > 300:
                            # No usable window — show the poll's last known rate.
                            rx_kbps = prev_rx
                            tx_kbps = prev_tx
                        else:
                            rx_diff = rx_bytes - prev.rx_bytes
                            tx_diff = tx_bytes - prev.tx_bytes
                            if rx_diff < 0 or tx_diff < 0 or (rx_diff == 0 and tx_diff == 0):
                                # Counter reset or not yet refreshed — hold last rate.
                                rx_kbps = prev_rx
                                tx_kbps = prev_tx
                            else:
                                instant_rx = (rx_diff * 8) / time_diff / 1000
                                instant_tx = (tx_diff * 8) / time_diff / 1000
                                MAX_VALID_KBPS = 1_500_000  # 1.5 Gbps per ONU
                                if instant_rx > MAX_VALID_KBPS or instant_tx > MAX_VALID_KBPS:
                                    rx_kbps = prev_rx  # implausible spike — hold
                                    tx_kbps = prev_tx
                                else:
                                    rx_kbps = round(instant_rx, 2)
                                    tx_kbps = round(instant_tx, 2)
                    # else: no baseline yet -> rx_kbps/tx_kbps stay 0; the poll
                    # cycle will create the snapshot on its next run.

                    # Get ONU description from database
                    onu = db.query(ONU).filter(
                        ONU.olt_id == olt_id,
                        ONU.mac_address == mac
                    ).first()

                    onu_online = bool(onu and onu.is_online)

                    # Override with Mikrotik rates if available (more accurate) —
                    # but never for an offline ONU.
                    if onu_online and mac in mk_rates:
                        rx_kbps = mk_rates[mac]['rx_kbps']
                        tx_kbps = mk_rates[mac]['tx_kbps']

                    # Offline ONUs must never show live traffic.
                    if not onu_online:
                        rx_kbps = 0
                        tx_kbps = 0

                    traffic_data.append({
                        "mac_address": mac,
                        "pon_port": pon_port,
                        "onu_id": onu_id,
                        "description": onu.description if onu else None,
                        "is_online": onu_online,
                        "rx_kbps": rx_kbps,
                        "tx_kbps": tx_kbps,
                        "rx_mbps": round(rx_kbps / 1000, 2),
                        "tx_mbps": round(tx_kbps / 1000, 2)
                    })

                db.commit()

                # Sort: online first (by traffic descending), then offline at bottom
                traffic_data.sort(key=lambda x: (not x.get('is_online', False), -(x['rx_kbps'] + x['tx_kbps'])))

                # Calculate poll duration for live indicator
                poll_ms = int((datetime.now() - poll_start).total_seconds() * 1000)

                # Broadcast to all connected clients
                await traffic_manager.broadcast(olt_id, {
                    "olt_id": olt_id,
                    "olt_name": olt_name,
                    "timestamp": current_time.isoformat(),
                    "onu_count": len(traffic_data),
                    "poll_ms": poll_ms,  # Shows how fast the poll was
                    "traffic": traffic_data
                })

            finally:
                db.close()

            # Wait 1.5 seconds before next poll (faster refresh)
            await asyncio.sleep(1.5)

        except asyncio.CancelledError:
            logger.info(f"Traffic polling loop cancelled for OLT {olt_id}")
            break
        except Exception as e:
            logger.error(f"Traffic polling error for OLT {olt_id}: {e}")
            await asyncio.sleep(5)  # Wait longer on error

    logger.info(f"Stopped traffic polling loop for OLT {olt_id}")


async def saas_traffic_polling_loop(olt_id: int, db_session_factory):
    """SaaS-mode traffic loop: reads latest agent-pushed data from TrafficHistory."""
    logger.info(f"Started SaaS traffic polling loop for OLT {olt_id}")
    await asyncio.sleep(0.5)

    while olt_id in traffic_manager.active_connections and traffic_manager.active_connections[olt_id]:
        try:
            db = db_session_factory()
            try:
                olt = db.query(OLT).filter(OLT.id == olt_id).first()
                if not olt:
                    break
                if olt.tenant_id:
                    set_session_tenant(db, olt.tenant_id)
                    from models import is_postgres
                    if is_postgres():
                        from sqlalchemy import text as _sql_text
                        safe_tid = str(olt.tenant_id).replace("'", "")
                        db.execute(_sql_text(f"SET LOCAL app.current_tenant_id = '{safe_tid}'"))
                olt_name = olt.name or f"OLT-{olt_id}"

                # Get latest ONU traffic from the most recent TrafficHistory rows
                cutoff = datetime.utcnow() - timedelta(seconds=90)
                from sqlalchemy import func as sqlfunc
                # Subquery: latest timestamp per ONU entity
                latest = db.query(
                    TrafficHistory.onu_db_id,
                    sqlfunc.max(TrafficHistory.timestamp).label("max_ts"),
                ).filter(
                    TrafficHistory.olt_id == olt_id,
                    TrafficHistory.entity_type == "onu",
                    TrafficHistory.onu_db_id != None,
                    TrafficHistory.timestamp >= cutoff,
                ).group_by(TrafficHistory.onu_db_id).subquery()

                rows = db.query(TrafficHistory).join(
                    latest,
                    (TrafficHistory.onu_db_id == latest.c.onu_db_id) &
                    (TrafficHistory.timestamp == latest.c.max_ts),
                ).all()

                traffic_data = []
                for th in rows:
                    onu = db.query(ONU).filter(ONU.id == th.onu_db_id).first()
                    if onu:
                        traffic_data.append({
                            "mac_address": onu.mac_address,
                            "pon_port": onu.pon_port,
                            "onu_id": onu.onu_id,
                            "description": onu.description,
                            "is_online": onu.is_online,
                            "rx_kbps": th.rx_kbps or 0,
                            "tx_kbps": th.tx_kbps or 0,
                            "rx_mbps": round((th.rx_kbps or 0) / 1000, 2),
                            "tx_mbps": round((th.tx_kbps or 0) / 1000, 2),
                        })

                # Fallback: if no per-ONU TrafficHistory (fallback-polled OLTs),
                # read rates from TrafficSnapshot instead.
                if not traffic_data:
                    all_onus = db.query(ONU).filter(ONU.olt_id == olt_id).all()
                    snap_map = {
                        s.mac_address: s
                        for s in db.query(TrafficSnapshot).filter(TrafficSnapshot.olt_id == olt_id).all()
                    }
                    for onu in all_onus:
                        s = snap_map.get(onu.mac_address)
                        rx = (s.last_rx_kbps or 0) if s else 0
                        tx = (s.last_tx_kbps or 0) if s else 0
                        traffic_data.append({
                            "mac_address": onu.mac_address,
                            "pon_port": onu.pon_port,
                            "onu_id": onu.onu_id,
                            "description": onu.description,
                            "is_online": onu.is_online,
                            "rx_kbps": rx, "tx_kbps": tx,
                            "rx_mbps": round(rx / 1000, 2),
                            "tx_mbps": round(tx / 1000, 2),
                        })
                else:
                    # Include offline ONUs with zero traffic
                    seen_onu_ids = {th.onu_db_id for th in rows}
                    for onu in db.query(ONU).filter(ONU.olt_id == olt_id).all():
                        if onu.id not in seen_onu_ids:
                            traffic_data.append({
                                "mac_address": onu.mac_address,
                                "pon_port": onu.pon_port,
                                "onu_id": onu.onu_id,
                                "description": onu.description,
                                "is_online": onu.is_online,
                                "rx_kbps": 0, "tx_kbps": 0,
                                "rx_mbps": 0, "tx_mbps": 0,
                            })

                traffic_data.sort(key=lambda x: (not x.get("is_online", False), -(x["rx_kbps"] + x["tx_kbps"])))

                await traffic_manager.broadcast(olt_id, {
                    "olt_id": olt_id,
                    "olt_name": olt_name,
                    "timestamp": datetime.utcnow().isoformat(),
                    "onu_count": len(traffic_data),
                    "traffic": traffic_data,
                })
            finally:
                db.close()

            await asyncio.sleep(5)

        except asyncio.CancelledError:
            break
        except Exception as e:
            logger.error(f"SaaS traffic polling error for OLT {olt_id}: {e}")
            await asyncio.sleep(5)

    logger.info(f"Stopped SaaS traffic polling loop for OLT {olt_id}")


@app.websocket("/ws/traffic/{olt_id}")
async def websocket_traffic(websocket: WebSocket, olt_id: int):
    """
    WebSocket endpoint for live traffic updates.
    Connects to an OLT and streams traffic data every 3 seconds.
    In SAAS_MODE, reads from agent-pushed TrafficHistory instead of SNMP.
    """
    from models import SessionLocal

    # Verify OLT exists
    db = SessionLocal()
    try:
        olt = db.query(OLT).filter(OLT.id == olt_id).first()
        if not olt:
            await websocket.close(code=4004, reason="OLT not found")
            return
        olt_ip = olt.ip_address
    finally:
        db.close()

    await traffic_manager.connect(websocket, olt_id)

    try:
        # Start traffic polling task if not already running
        if olt_id not in traffic_manager.traffic_tasks or traffic_manager.traffic_tasks[olt_id].done():
            if os.getenv("SAAS_MODE"):
                traffic_manager.traffic_tasks[olt_id] = asyncio.create_task(
                    saas_traffic_polling_loop(olt_id, SessionLocal)
                )
            else:
                traffic_manager.traffic_tasks[olt_id] = asyncio.create_task(
                    traffic_polling_loop(olt_id, olt_ip, SessionLocal)
                )

        # Keep connection alive and wait for client messages
        while True:
            try:
                # Wait for any message from client (ping/pong or disconnect)
                data = await websocket.receive_text()
                # Client can send "ping" to keep alive
                if data == "ping":
                    await websocket.send_text("pong")
            except WebSocketDisconnect:
                break
    except Exception as e:
        logger.error(f"WebSocket error for OLT {olt_id}: {e}")
    finally:
        traffic_manager.disconnect(websocket, olt_id)


# ============ Traffic History/Graph Endpoints ============

TIME_RANGES = {
    '5m': timedelta(minutes=5),
    '15m': timedelta(minutes=15),
    '30m': timedelta(minutes=30),
    '1h': timedelta(hours=1),
    '6h': timedelta(hours=6),
    '24h': timedelta(hours=24),
    '1w': timedelta(weeks=1),
    '1M': timedelta(days=30)
}


@app.get("/api/traffic/history/onu/{onu_id}")
async def get_onu_traffic_history(
    onu_id: int,
    range: str = Query('1h', description="Time range: 5m, 15m, 30m, 1h, 6h, 24h, 1w, 1M"),
    user: User = Depends(require_auth),
    db: Session = Depends(get_db)
):
    """
    Get historical traffic data for a specific ONU.
    Returns time-series data for graphing.
    """
    if range not in TIME_RANGES:
        raise HTTPException(status_code=400, detail=f"Invalid range. Use: {', '.join(TIME_RANGES.keys())}")

    onu = db.query(ONU).filter(ONU.id == onu_id).first()
    if not onu:
        raise HTTPException(status_code=404, detail="ONU not found")

    # Check user access to this ONU's OLT
    allowed_olt_ids = get_user_allowed_olt_ids(user, db)
    if allowed_olt_ids is not None and onu.olt_id not in allowed_olt_ids:
        raise HTTPException(status_code=403, detail="Access denied")

    time_delta = TIME_RANGES[range]
    start_time = datetime.utcnow() - time_delta
    # Format timestamp to match SQLite storage format (space instead of T)
    start_time_str = start_time.strftime('%Y-%m-%d %H:%M:%S')

    history = db.query(TrafficHistory).filter(
        TrafficHistory.entity_type == 'onu',
        TrafficHistory.onu_db_id == onu_id,
        TrafficHistory.timestamp >= start_time_str
    ).order_by(TrafficHistory.timestamp.asc()).all()

    return {
        "onu_id": onu_id,
        "onu_description": onu.description,
        "olt_id": onu.olt_id,
        "pon_port": onu.pon_port,
        "range": range,
        "start_time": start_time.isoformat(),
        "end_time": datetime.utcnow().isoformat(),
        "data_points": len(history),
        "data": [
            {
                "timestamp": h.timestamp.isoformat() + "Z",  # mark UTC so the browser doesn't shift it
                "rx_kbps": h.rx_kbps,
                "tx_kbps": h.tx_kbps,
                "rx_mbps": round(h.rx_kbps / 1000, 2),
                "tx_mbps": round(h.tx_kbps / 1000, 2)
            }
            for h in history
        ]
    }


@app.get("/api/traffic/history/pon/{olt_id}/{pon_port}")
async def get_pon_traffic_history(
    olt_id: int,
    pon_port: int,
    range: str = Query('1h', description="Time range: 5m, 15m, 30m, 1h, 6h, 24h, 1w, 1M"),
    user: User = Depends(require_auth),
    db: Session = Depends(get_db)
):
    """
    Get historical traffic data for a specific PON port on an OLT.
    Returns aggregated time-series data for graphing.
    """
    if range not in TIME_RANGES:
        raise HTTPException(status_code=400, detail=f"Invalid range. Use: {', '.join(TIME_RANGES.keys())}")

    olt = db.query(OLT).filter(OLT.id == olt_id).first()
    if not olt:
        raise HTTPException(status_code=404, detail="OLT not found")

    # Check user access
    allowed_olt_ids = get_user_allowed_olt_ids(user, db)
    if allowed_olt_ids is not None and olt_id not in allowed_olt_ids:
        raise HTTPException(status_code=403, detail="Access denied")

    time_delta = TIME_RANGES[range]
    start_time = datetime.utcnow() - time_delta
    start_time_str = start_time.strftime('%Y-%m-%d %H:%M:%S')
    entity_id = f"{olt_id}:{pon_port}"

    history = db.query(TrafficHistory).filter(
        TrafficHistory.entity_type == 'pon',
        TrafficHistory.entity_id == entity_id,
        TrafficHistory.timestamp >= start_time_str
    ).order_by(TrafficHistory.timestamp.asc()).all()

    return {
        "olt_id": olt_id,
        "olt_name": olt.name,
        "pon_port": pon_port,
        "range": range,
        "start_time": start_time.isoformat(),
        "end_time": datetime.utcnow().isoformat(),
        "data_points": len(history),
        "data": [
            {
                "timestamp": h.timestamp.isoformat() + "Z",  # mark UTC so the browser doesn't shift it
                "rx_kbps": h.rx_kbps,
                "tx_kbps": h.tx_kbps,
                "rx_mbps": round(h.rx_kbps / 1000, 2),
                "tx_mbps": round(h.tx_kbps / 1000, 2)
            }
            for h in history
        ]
    }


@app.get("/api/traffic/history/olt/{olt_id}")
async def get_olt_traffic_history(
    olt_id: int,
    range: str = Query('1h', description="Time range: 5m, 15m, 30m, 1h, 6h, 24h, 1w, 1M"),
    user: User = Depends(require_auth),
    db: Session = Depends(get_db)
):
    """
    Get historical traffic data for a specific OLT (total traffic).
    Returns aggregated time-series data for graphing.
    """
    if range not in TIME_RANGES:
        raise HTTPException(status_code=400, detail=f"Invalid range. Use: {', '.join(TIME_RANGES.keys())}")

    olt = db.query(OLT).filter(OLT.id == olt_id).first()
    if not olt:
        raise HTTPException(status_code=404, detail="OLT not found")

    # Check user access
    allowed_olt_ids = get_user_allowed_olt_ids(user, db)
    if allowed_olt_ids is not None and olt_id not in allowed_olt_ids:
        raise HTTPException(status_code=403, detail="Access denied")

    time_delta = TIME_RANGES[range]
    start_time = datetime.utcnow() - time_delta
    start_time_str = start_time.strftime('%Y-%m-%d %H:%M:%S')

    history = db.query(TrafficHistory).filter(
        TrafficHistory.entity_type == 'olt',
        TrafficHistory.olt_id == olt_id,
        TrafficHistory.timestamp >= start_time_str
    ).order_by(TrafficHistory.timestamp.asc()).all()

    return {
        "olt_id": olt_id,
        "olt_name": olt.name,
        "range": range,
        "start_time": start_time.isoformat(),
        "end_time": datetime.utcnow().isoformat(),
        "data_points": len(history),
        "data": [
            {
                "timestamp": h.timestamp.isoformat() + "Z",  # mark UTC so the browser doesn't shift it
                "rx_kbps": h.rx_kbps,
                "tx_kbps": h.tx_kbps,
                "rx_mbps": round(h.rx_kbps / 1000, 2),
                "tx_mbps": round(h.tx_kbps / 1000, 2)
            }
            for h in history
        ]
    }


@app.delete("/api/traffic/history/cleanup")
async def cleanup_traffic_history(
    days: int = Query(30, description="Delete history older than X days"),
    user: User = Depends(require_admin),
    db: Session = Depends(get_db)
):
    """
    Clean up old traffic history data.
    Admin only. Deletes records older than specified days.
    """
    cutoff_time = datetime.utcnow() - timedelta(days=days)

    deleted = db.query(TrafficHistory).filter(
        TrafficHistory.timestamp < cutoff_time
    ).delete()

    db.commit()

    return {
        "message": f"Deleted {deleted} records older than {days} days",
        "cutoff_time": cutoff_time.isoformat()
    }


# ============ Diagram Endpoints ============

@app.get("/api/diagrams", response_model=DiagramListResponse)
async def list_diagrams(
    user: User = Depends(require_auth),
    db: Session = Depends(get_db)
):
    """
    List all diagrams accessible by the current user.
    Returns user's own diagrams plus shared diagrams.
    """
    # Get user's own diagrams and shared diagrams
    diagrams = db.query(Diagram).filter(
        or_(
            Diagram.owner_id == user.id,
            Diagram.is_shared == True
        )
    ).order_by(Diagram.updated_at.desc()).all()

    result = []
    for d in diagrams:
        owner = db.query(User).filter(User.id == d.owner_id).first()
        result.append(DiagramResponse(
            id=d.id,
            owner_id=d.owner_id,
            owner_name=owner.username if owner else None,
            name=d.name,
            nodes=d.nodes,
            connections=d.connections,
            settings=d.settings,
            is_shared=d.is_shared,
            created_at=d.created_at,
            updated_at=d.updated_at
        ))

    return DiagramListResponse(diagrams=result, total=len(result))


@app.post("/api/diagrams", response_model=DiagramResponse)
async def create_diagram(
    diagram: DiagramCreate,
    user: User = Depends(require_auth),
    db: Session = Depends(get_db)
):
    """Create a new diagram"""
    db_diagram = Diagram(
        owner_id=user.id,
        name=diagram.name,
        nodes=diagram.nodes,
        connections=diagram.connections,
        settings=diagram.settings,
        is_shared=diagram.is_shared
    )
    db.add(db_diagram)
    db.commit()
    db.refresh(db_diagram)

    return DiagramResponse(
        id=db_diagram.id,
        owner_id=db_diagram.owner_id,
        owner_name=user.username,
        name=db_diagram.name,
        nodes=db_diagram.nodes,
        connections=db_diagram.connections,
        settings=db_diagram.settings,
        is_shared=db_diagram.is_shared,
        created_at=db_diagram.created_at,
        updated_at=db_diagram.updated_at
    )


@app.get("/api/diagrams/{diagram_id}", response_model=DiagramResponse)
async def get_diagram(
    diagram_id: int,
    user: User = Depends(require_auth),
    db: Session = Depends(get_db)
):
    """Get a specific diagram by ID"""
    diagram = db.query(Diagram).filter(Diagram.id == diagram_id).first()
    if not diagram:
        raise HTTPException(status_code=404, detail="Diagram not found")

    # Check access: owner or shared
    if diagram.owner_id != user.id and not diagram.is_shared:
        raise HTTPException(status_code=403, detail="Access denied")

    owner = db.query(User).filter(User.id == diagram.owner_id).first()

    return DiagramResponse(
        id=diagram.id,
        owner_id=diagram.owner_id,
        owner_name=owner.username if owner else None,
        name=diagram.name,
        nodes=diagram.nodes,
        connections=diagram.connections,
        settings=diagram.settings,
        is_shared=diagram.is_shared,
        created_at=diagram.created_at,
        updated_at=diagram.updated_at
    )


@app.put("/api/diagrams/{diagram_id}", response_model=DiagramResponse)
async def update_diagram(
    diagram_id: int,
    update: DiagramUpdate,
    user: User = Depends(require_auth),
    db: Session = Depends(get_db)
):
    """Update a diagram (owner only)"""
    diagram = db.query(Diagram).filter(Diagram.id == diagram_id).first()
    if not diagram:
        raise HTTPException(status_code=404, detail="Diagram not found")

    # Only owner can update
    if diagram.owner_id != user.id:
        raise HTTPException(status_code=403, detail="Only the owner can update this diagram")

    # Update fields if provided
    if update.name is not None:
        diagram.name = update.name
    if update.nodes is not None:
        diagram.nodes = update.nodes
    if update.connections is not None:
        diagram.connections = update.connections
    if update.settings is not None:
        diagram.settings = update.settings
    if update.is_shared is not None:
        diagram.is_shared = update.is_shared

    diagram.updated_at = datetime.utcnow()
    db.commit()
    db.refresh(diagram)

    owner = db.query(User).filter(User.id == diagram.owner_id).first()

    return DiagramResponse(
        id=diagram.id,
        owner_id=diagram.owner_id,
        owner_name=owner.username if owner else None,
        name=diagram.name,
        nodes=diagram.nodes,
        connections=diagram.connections,
        settings=diagram.settings,
        is_shared=diagram.is_shared,
        created_at=diagram.created_at,
        updated_at=diagram.updated_at
    )


@app.delete("/api/diagrams/{diagram_id}")
async def delete_diagram(
    diagram_id: int,
    user: User = Depends(require_auth),
    db: Session = Depends(get_db)
):
    """Delete a diagram (owner only)"""
    diagram = db.query(Diagram).filter(Diagram.id == diagram_id).first()
    if not diagram:
        raise HTTPException(status_code=404, detail="Diagram not found")

    # Only owner can delete
    if diagram.owner_id != user.id:
        raise HTTPException(status_code=403, detail="Only the owner can delete this diagram")

    db.delete(diagram)
    db.commit()

    return {"message": "Diagram deleted successfully"}


# ============ License Info ============

@app.get("/api/license")
async def get_license_info(user: User = Depends(require_auth)):
    """Get current license information"""
    from license_manager import license_manager
    info = license_manager.get_license_info()

    # Add license key from file
    license_key = None
    try:
        license_file = Path('/etc/olt-manager/license.key')
        if license_file.exists():
            license_key = license_file.read_text().strip()
    except:
        pass
    info['license_key'] = license_key or os.getenv('OLT_LICENSE_KEY', 'N/A')

    # Calculate days remaining
    expires_at = info.get('expires_at')
    if expires_at:
        try:
            exp_date = datetime.fromisoformat(expires_at.replace('Z', '+00:00'))
            if exp_date.tzinfo:
                exp_date = exp_date.replace(tzinfo=None)
            days_remaining = (exp_date - datetime.now()).days
            info['days_remaining'] = max(0, days_remaining)
        except:
            info['days_remaining'] = None
    else:
        info['days_remaining'] = None

    # Add status for frontend
    if not info.get('valid'):
        error_msg = info.get('error_message', '').lower()
        if 'suspended' in error_msg:
            info['status'] = 'suspended'
            info['status_message'] = 'License has been suspended. Please contact support.'
        elif 'expired' in error_msg:
            info['status'] = 'expired'
            info['status_message'] = 'License has expired. Please renew your subscription.'
        elif 'revoked' in error_msg:
            info['status'] = 'revoked'
            info['status_message'] = 'License has been revoked. Please contact support.'
        else:
            info['status'] = 'invalid'
            info['status_message'] = info.get('error_message', 'License is invalid')
    else:
        info['status'] = 'active'
        info['status_message'] = None

    # Add current usage counts
    db = next(get_db())
    try:
        info['current_olts'] = db.query(OLT).count()
        info['current_onus'] = db.query(ONU).count()
        info['current_users'] = db.query(User).count()
    except:
        info['current_olts'] = 0
        info['current_onus'] = 0
        info['current_users'] = 0
    finally:
        db.close()

    return info


@app.post("/api/license/refresh")
async def refresh_license(user: User = Depends(require_auth)):
    """Force refresh license from server"""
    from license_manager import license_manager
    try:
        # Force revalidation from server
        license_manager.validate()
        return {"success": True, "message": "License refreshed successfully"}
    except Exception as e:
        return {"success": False, "message": str(e)}


@app.post("/api/system/restart-service")
async def restart_service(user: User = Depends(require_auth)):
    """Restart the OLT Manager service"""
    import subprocess
    try:
        # Try different service names
        service_names = ['olt-backend', 'olt-manager', 'oltmanager']
        for name in service_names:
            result = subprocess.run(['systemctl', 'is-active', name], capture_output=True)
            if result.returncode == 0:
                subprocess.Popen(['systemctl', 'restart', name])
                return {"success": True, "message": f"Service {name} is restarting..."}

        # If no systemd service, try to restart the process
        subprocess.Popen(['pkill', '-f', 'uvicorn|main.bin|main.py'])
        return {"success": True, "message": "Service is restarting..."}
    except Exception as e:
        return {"success": False, "message": str(e)}


@app.post("/api/system/reboot")
async def reboot_server(user: User = Depends(require_auth)):
    """Reboot the server"""
    import subprocess
    try:
        # Schedule reboot in 3 seconds to allow response to be sent
        subprocess.Popen(['shutdown', '-r', '+0'])
        return {"success": True, "message": "Server is rebooting..."}
    except Exception as e:
        return {"success": False, "message": str(e)}


# ============ Remote Access Tunnel API (REMOVED in Phase 3) ============
#
# The Cloudflare Tunnel feature was the single-tenant on-prem story.
# In the SaaS world it's replaced by per-workspace WireGuard provisioning
# (see backend/wireguard/). These routes return HTTP 410 Gone for one
# release with a migration message, then will be deleted entirely.

_TUNNEL_GONE_MSG = (
    "The Cloudflare Tunnel feature has been removed. Use the WireGuard "
    "workspace connectivity at /api/workspaces/{id}/wireguard/provision instead."
)


@app.get("/api/tunnel/status")
def _tunnel_status_gone():
    raise HTTPException(status_code=410, detail=_TUNNEL_GONE_MSG)


@app.post("/api/tunnel/enable")
def _tunnel_enable_gone():
    raise HTTPException(status_code=410, detail=_TUNNEL_GONE_MSG)


@app.post("/api/tunnel/disable")
def _tunnel_disable_gone():
    raise HTTPException(status_code=410, detail=_TUNNEL_GONE_MSG)


@app.delete("/api/tunnel")
def _tunnel_delete_gone():
    raise HTTPException(status_code=410, detail=_TUNNEL_GONE_MSG)


# ============ Health Check ============

@app.get("/api/health")
def health_check():
    """Health check endpoint"""
    return {"status": "healthy", "timestamp": datetime.now().isoformat()}


@app.get("/api/trap/status")
def get_trap_status():
    """Get SNMP trap receiver status"""
    global trap_receiver
    return {
        "running": trap_receiver is not None and trap_receiver.running,
        "port": trap_receiver.port if trap_receiver else 162
    }


# ============ FEATURE 1: GPS Map View ============

@app.get("/api/map/onus")
def get_map_onus(
    olt_id: Optional[int] = None,
    region_id: Optional[int] = None,
    online_only: bool = False,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_auth)
):
    """Get all ONUs with GPS coordinates for map display"""
    query = db.query(ONU).filter(
        ONU.latitude.isnot(None),
        ONU.longitude.isnot(None)
    )

    if olt_id:
        query = query.filter(ONU.olt_id == olt_id)
    if region_id:
        query = query.filter(ONU.region_id == region_id)
    if online_only:
        query = query.filter(ONU.is_online == True)

    onus = query.all()

    return {
        "total": len(onus),
        "onus": [
            {
                "id": onu.id,
                "mac_address": onu.mac_address,
                "description": onu.description,
                "latitude": onu.latitude,
                "longitude": onu.longitude,
                "address": onu.address,
                "is_online": onu.is_online,
                "rx_power": onu.rx_power,
                "olt_id": onu.olt_id,
                "olt_name": onu.olt.name if onu.olt else None,
                "region_id": onu.region_id,
                "region_name": onu.region.name if onu.region else None,
                "region_color": onu.region.color if onu.region else "#3B82F6"
            }
            for onu in onus
        ]
    }


@app.get("/api/map/regions")
def get_map_regions(
    db: Session = Depends(get_db),
    current_user: User = Depends(require_auth)
):
    """Get all regions with GPS coordinates for map display"""
    regions = db.query(Region).filter(
        Region.latitude.isnot(None),
        Region.longitude.isnot(None)
    ).all()

    return {
        "total": len(regions),
        "regions": [
            {
                "id": region.id,
                "name": region.name,
                "latitude": region.latitude,
                "longitude": region.longitude,
                "address": region.address,
                "color": region.color,
                "onu_count": len(region.onus)
            }
            for region in regions
        ]
    }


# ============ FEATURE 2: Email Alerts ============

@app.post("/api/email/test")
async def test_email(
    db: Session = Depends(get_db),
    current_user: User = Depends(require_admin)
):
    """Test email notification"""
    import smtplib
    from email.mime.text import MIMEText
    from email.mime.multipart import MIMEMultipart

    # Get settings
    settings = {s.key: s.value for s in db.query(Settings).all()}

    smtp_server = settings.get('smtp_server', '')
    smtp_port = int(settings.get('smtp_port', 587))
    smtp_user = settings.get('smtp_user', '')
    smtp_password = settings.get('smtp_password', '')
    email_recipient = settings.get('email_recipient', '')

    if not all([smtp_server, smtp_user, smtp_password, email_recipient]):
        raise HTTPException(status_code=400, detail="Email settings not configured")

    try:
        msg = MIMEMultipart()
        msg['From'] = smtp_user
        msg['To'] = email_recipient
        msg['Subject'] = "OLT Manager - Test Email"

        body = f"""
        This is a test email from OLT Manager.

        Time: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}
        Server: {os.uname().nodename}

        If you received this, email notifications are working correctly!
        """
        msg.attach(MIMEText(body, 'plain'))

        server = smtplib.SMTP(smtp_server, smtp_port)
        server.starttls()
        server.login(smtp_user, smtp_password)
        server.send_message(msg)
        server.quit()

        return {"success": True, "message": f"Test email sent to {email_recipient}"}
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Failed to send email: {str(e)}")


@app.put("/api/email/settings")
async def update_email_settings(
    smtp_server: str = Body(...),
    smtp_port: int = Body(587),
    smtp_user: str = Body(...),
    smtp_password: str = Body(...),
    email_recipient: str = Body(...),
    email_enabled: bool = Body(True),
    db: Session = Depends(get_db),
    current_user: User = Depends(require_admin)
):
    """Update email notification settings"""
    settings_map = {
        'smtp_server': smtp_server,
        'smtp_port': str(smtp_port),
        'smtp_user': smtp_user,
        'smtp_password': smtp_password,
        'email_recipient': email_recipient,
        'email_enabled': str(email_enabled)
    }

    for key, value in settings_map.items():
        setting = db.query(Settings).filter(Settings.key == key).first()
        if setting:
            setting.value = value
        else:
            db.add(Settings(key=key, value=value))

    db.commit()
    return {"success": True, "message": "Email settings updated"}


# ============ FEATURE 3: Report Generation ============

@app.get("/api/reports/onus")
async def generate_onu_report(
    format: str = Query("json", regex="^(json|csv|excel)$"),
    olt_id: Optional[int] = None,
    region_id: Optional[int] = None,
    online_only: bool = False,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_auth)
):
    """Generate ONU report in various formats"""
    from fastapi.responses import StreamingResponse
    import csv
    import io

    query = db.query(ONU).join(OLT)

    if olt_id:
        query = query.filter(ONU.olt_id == olt_id)
    if region_id:
        query = query.filter(ONU.region_id == region_id)
    if online_only:
        query = query.filter(ONU.is_online == True)

    onus = query.all()

    if format == "json":
        return {
            "generated_at": datetime.now().isoformat(),
            "total": len(onus),
            "onus": [
                {
                    "id": onu.id,
                    "olt_name": onu.olt.name,
                    "pon_port": onu.pon_port,
                    "onu_id": onu.onu_id,
                    "mac_address": onu.mac_address,
                    "description": onu.description,
                    "is_online": onu.is_online,
                    "rx_power": onu.rx_power,
                    "distance": onu.distance,
                    "region": onu.region.name if onu.region else None,
                    "address": onu.address,
                    "last_seen": onu.last_seen.isoformat() if onu.last_seen else None
                }
                for onu in onus
            ]
        }

    elif format == "csv":
        output = io.StringIO()
        writer = csv.writer(output)
        writer.writerow(['ID', 'OLT', 'PON Port', 'ONU ID', 'MAC Address', 'Description',
                        'Online', 'RX Power (dBm)', 'Distance (m)', 'Region', 'Address', 'Last Seen'])

        for onu in onus:
            writer.writerow([
                onu.id, onu.olt.name, onu.pon_port, onu.onu_id, onu.mac_address,
                onu.description or '', onu.is_online, onu.rx_power or '',
                onu.distance or '', onu.region.name if onu.region else '',
                onu.address or '', onu.last_seen.strftime('%Y-%m-%d %H:%M') if onu.last_seen else ''
            ])

        output.seek(0)
        return StreamingResponse(
            iter([output.getvalue()]),
            media_type="text/csv",
            headers={"Content-Disposition": f"attachment; filename=onu_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"}
        )

    elif format == "excel":
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill

            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "ONU Report"

            # Header
            headers = ['ID', 'OLT', 'PON Port', 'ONU ID', 'MAC Address', 'Description',
                      'Online', 'RX Power (dBm)', 'Distance (m)', 'Region', 'Address', 'Last Seen']
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                cell.font = Font(bold=True, color="FFFFFF")

            # Data
            for row, onu in enumerate(onus, 2):
                ws.cell(row=row, column=1, value=onu.id)
                ws.cell(row=row, column=2, value=onu.olt.name)
                ws.cell(row=row, column=3, value=onu.pon_port)
                ws.cell(row=row, column=4, value=onu.onu_id)
                ws.cell(row=row, column=5, value=onu.mac_address)
                ws.cell(row=row, column=6, value=onu.description or '')
                ws.cell(row=row, column=7, value='Online' if onu.is_online else 'Offline')
                ws.cell(row=row, column=8, value=onu.rx_power)
                ws.cell(row=row, column=9, value=onu.distance)
                ws.cell(row=row, column=10, value=onu.region.name if onu.region else '')
                ws.cell(row=row, column=11, value=onu.address or '')
                ws.cell(row=row, column=12, value=onu.last_seen.strftime('%Y-%m-%d %H:%M') if onu.last_seen else '')

            output = io.BytesIO()
            wb.save(output)
            output.seek(0)

            return StreamingResponse(
                iter([output.getvalue()]),
                media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                headers={"Content-Disposition": f"attachment; filename=onu_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"}
            )
        except ImportError:
            raise HTTPException(status_code=500, detail="openpyxl not installed. Use CSV format instead.")


@app.get("/api/reports/signal-quality")
async def generate_signal_report(
    threshold: float = Query(None, description="Signal threshold in dBm (uses alarm setting if not specified)"),
    db: Session = Depends(get_db),
    current_user: User = Depends(require_auth)
):
    """Generate signal quality report - ONUs with low signal"""
    from sqlalchemy import or_

    # Use alarm threshold from settings if not specified
    if threshold is None:
        alarm_settings = get_alarm_settings(db)
        threshold = float(alarm_settings.get("weak_signal_threshold", -25))
        lower_threshold = float(alarm_settings.get("weak_signal_lower_threshold", -30))
    else:
        lower_threshold = threshold - 5  # Default 5 dBm below upper threshold

    # Query ONUs with weak signal - check BOTH rx_power (OLT measured) AND onu_rx_power (ONU self-reported)
    onus = db.query(ONU).filter(
        ONU.is_online == True,
        or_(
            (ONU.rx_power.isnot(None)) & (ONU.rx_power < threshold),
            (ONU.onu_rx_power.isnot(None)) & (ONU.onu_rx_power < threshold)
        )
    ).all()

    # Sort by the worse signal value
    def get_signal(onu):
        # Use onu_rx_power if available (more accurate), otherwise rx_power
        if onu.onu_rx_power is not None:
            return onu.onu_rx_power
        return onu.rx_power or 0

    onus = sorted(onus, key=get_signal)

    # Calculate severity based on danger zone
    def get_severity(rx_power):
        if rx_power is None:
            return "unknown"
        range_size = threshold - lower_threshold
        position = (rx_power - lower_threshold) / range_size if range_size != 0 else 0.5
        if position < 0.33:
            return "critical"
        elif position < 0.66:
            return "high"
        else:
            return "warning"

    result_onus = []
    for onu in onus:
        # Use best available signal value (prefer ONU self-reported)
        signal = onu.onu_rx_power if onu.onu_rx_power is not None else onu.rx_power
        result_onus.append({
            "id": onu.id,
            "olt_name": onu.olt.name,
            "pon_port": onu.pon_port,
            "onu_id": onu.onu_id,
            "mac_address": onu.mac_address,
            "description": onu.description if onu.description and onu.description.upper() != "NULL" else None,
            "rx_power": signal,
            "olt_rx_power": onu.rx_power,
            "onu_rx_power": onu.onu_rx_power,
            "distance": onu.distance,
            "address": onu.address,
            "region": onu.region.name if onu.region else None,
            "severity": get_severity(signal)
        })

    return {
        "generated_at": datetime.now().isoformat(),
        "threshold": threshold,
        "lower_threshold": lower_threshold,
        "total_low_signal": len(result_onus),
        "onus": result_onus
    }


# ============ FEATURE 4: Signal Quality Alerts ============

@app.get("/api/alerts/rules")
def get_alert_rules(
    db: Session = Depends(get_db),
    current_user: User = Depends(require_auth)
):
    """Get all alert rules"""
    rules = db.query(AlertRule).all()
    return {
        "total": len(rules),
        "rules": [
            {
                "id": rule.id,
                "name": rule.name,
                "rule_type": rule.rule_type,
                "threshold": rule.threshold,
                "comparison": rule.comparison,
                "notify_email": rule.notify_email,
                "notify_sms": rule.notify_sms,
                "notify_whatsapp": rule.notify_whatsapp,
                "is_enabled": rule.is_enabled,
                "cooldown_minutes": rule.cooldown_minutes
            }
            for rule in rules
        ]
    }


@app.post("/api/alerts/rules")
def create_alert_rule(
    name: str = Body(...),
    rule_type: str = Body(...),
    threshold: Optional[float] = Body(None),
    comparison: Optional[str] = Body("lt"),
    notify_email: bool = Body(False),
    notify_sms: bool = Body(False),
    notify_whatsapp: bool = Body(True),
    cooldown_minutes: int = Body(60),
    db: Session = Depends(get_db),
    current_user: User = Depends(require_admin)
):
    """Create a new alert rule"""
    rule = AlertRule(
        name=name,
        rule_type=rule_type,
        threshold=threshold,
        comparison=comparison,
        notify_email=notify_email,
        notify_sms=notify_sms,
        notify_whatsapp=notify_whatsapp,
        cooldown_minutes=cooldown_minutes
    )
    db.add(rule)
    db.commit()
    db.refresh(rule)

    return {"success": True, "id": rule.id, "message": "Alert rule created"}


@app.delete("/api/alerts/rules/{rule_id}")
def delete_alert_rule(
    rule_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_admin)
):
    """Delete an alert rule"""
    rule = db.query(AlertRule).filter(AlertRule.id == rule_id).first()
    if not rule:
        raise HTTPException(status_code=404, detail="Alert rule not found")

    db.delete(rule)
    db.commit()
    return {"success": True, "message": "Alert rule deleted"}


# ============ FEATURE 5: Batch Operations ============

@app.post("/api/batch/onus/update")
async def batch_update_onus(
    onu_ids: List[int] = Body(...),
    region_id: Optional[int] = Body(None),
    description_prefix: Optional[str] = Body(None),
    db: Session = Depends(get_db),
    current_user: User = Depends(require_admin)
):
    """Batch update multiple ONUs"""
    updated = 0
    for onu_id in onu_ids:
        onu = db.query(ONU).filter(ONU.id == onu_id).first()
        if onu:
            if region_id is not None:
                onu.region_id = region_id
            if description_prefix:
                onu.description = f"{description_prefix} - {onu.mac_address[-8:]}"
            updated += 1

    db.commit()
    return {"success": True, "updated": updated, "message": f"Updated {updated} ONUs"}


@app.post("/api/batch/onus/reboot")
async def batch_reboot_onus(
    onu_ids: List[int] = Body(...),
    db: Session = Depends(get_db),
    current_user: User = Depends(require_admin)
):
    """Batch reboot multiple ONUs"""
    results = []
    for onu_id in onu_ids:
        onu = db.query(ONU).filter(ONU.id == onu_id).first()
        if onu:
            try:
                connector = OLTConnector(
                    onu.olt.ip_address,
                    onu.olt.username,
                    decrypt_sensitive(onu.olt.password)
                )
                success = connector.reboot_onu(onu.pon_port, onu.onu_id)
                if success:
                    # Reset online_since and mark offline so polling detects online transition
                    onu.online_since = None
                    onu.is_online = False  # Mark offline so next poll sets online_since when ONU comes back
                results.append({"onu_id": onu_id, "success": success})
            except Exception as e:
                results.append({"onu_id": onu_id, "success": False, "error": str(e)})

    db.commit()  # Save online_since changes
    return {"results": results, "total": len(results)}


@app.post("/api/batch/onus/delete")
async def batch_delete_onus(
    onu_ids: List[int] = Body(...),
    db: Session = Depends(get_db),
    current_user: User = Depends(require_admin)
):
    """Batch delete multiple ONUs from database"""
    deleted = db.query(ONU).filter(ONU.id.in_(onu_ids)).delete(synchronize_session=False)
    db.commit()
    return {"success": True, "deleted": deleted, "message": f"Deleted {deleted} ONUs"}


# ============ FEATURE 6: Configuration Backup/Restore ============

@app.get("/api/backups")
def get_backups(
    olt_id: Optional[int] = None,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_auth)
):
    """Get list of configuration backups"""
    query = db.query(ConfigBackup)
    if olt_id:
        query = query.filter(ConfigBackup.olt_id == olt_id)

    backups = query.order_by(ConfigBackup.created_at.desc()).all()

    return {
        "total": len(backups),
        "backups": [
            {
                "id": backup.id,
                "olt_id": backup.olt_id,
                "filename": backup.filename,
                "file_size": backup.file_size,
                "backup_type": backup.backup_type,
                "notes": backup.notes,
                "created_at": backup.created_at.isoformat()
            }
            for backup in backups
        ]
    }


@app.post("/api/backups/{olt_id}")
async def create_backup(
    olt_id: int,
    notes: Optional[str] = Body(None),
    db: Session = Depends(get_db),
    current_user: User = Depends(require_admin)
):
    """Create a configuration backup for an OLT"""
    olt = db.query(OLT).filter(OLT.id == olt_id).first()
    if not olt:
        raise HTTPException(status_code=404, detail="OLT not found")

    try:
        connector = OLTConnector(olt.ip_address, olt.username, decrypt_sensitive(olt.password))
        config = connector.get_running_config()

        # Save to file
        backup_dir = Path("/opt/olt-manager/backend/backups")
        backup_dir.mkdir(exist_ok=True)

        filename = f"{olt.name}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.cfg"
        filepath = backup_dir / filename

        with open(filepath, 'w') as f:
            f.write(config)

        # Save to database
        backup = ConfigBackup(
            olt_id=olt_id,
            filename=filename,
            file_size=len(config),
            backup_type='manual',
            notes=notes,
            created_by=current_user.id
        )
        db.add(backup)
        db.commit()

        # Log event
        log_event(db, 'backup_created', 'olt', olt_id, olt_id,
                  f"Config backup created for OLT '{olt.name}' by {current_user.username}")

        return {"success": True, "backup_id": backup.id, "filename": filename}
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Backup failed: {str(e)}")


@app.get("/api/backups/{backup_id}/download")
async def download_backup(
    backup_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_auth)
):
    """Download a configuration backup file"""
    from fastapi.responses import FileResponse

    backup = db.query(ConfigBackup).filter(ConfigBackup.id == backup_id).first()
    if not backup:
        raise HTTPException(status_code=404, detail="Backup not found")

    filepath = Path(f"/opt/olt-manager/backend/backups/{backup.filename}")
    if not filepath.exists():
        raise HTTPException(status_code=404, detail="Backup file not found")

    return FileResponse(filepath, filename=backup.filename)


@app.delete("/api/backups/{backup_id}")
async def delete_backup(
    backup_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_admin)
):
    """Delete a configuration backup"""
    backup = db.query(ConfigBackup).filter(ConfigBackup.id == backup_id).first()
    if not backup:
        raise HTTPException(status_code=404, detail="Backup not found")

    # Store info before delete
    filename = backup.filename
    olt_id = backup.olt_id

    # Delete file
    filepath = Path(f"/opt/olt-manager/backend/backups/{backup.filename}")
    if filepath.exists():
        filepath.unlink()

    db.delete(backup)
    db.commit()

    # Log event
    log_event(db, 'backup_deleted', 'olt', backup_id, olt_id,
              f"Config backup '{filename}' deleted by {current_user.username}")

    return {"success": True, "message": "Backup deleted"}


# ============ SYSTEM BACKUP - Full Database Backup/Restore ============

def upload_to_ftp(filepath: Path, settings: BackupSettings) -> tuple:
    """Upload backup file to FTP/SFTP server"""
    import ftplib
    try:
        if settings.ftp_use_sftp:
            # SFTP using paramiko
            import paramiko
            transport = paramiko.Transport((settings.ftp_host, settings.ftp_port or 22))
            transport.connect(username=settings.ftp_username, password=settings.ftp_password)
            sftp = paramiko.SFTPClient.from_transport(transport)
            remote_path = f"{settings.ftp_path}/{filepath.name}"
            sftp.put(str(filepath), remote_path)
            sftp.close()
            transport.close()
            return True, remote_path
        else:
            # FTP
            ftp = ftplib.FTP()
            ftp.connect(settings.ftp_host, settings.ftp_port or 21)
            ftp.login(settings.ftp_username, settings.ftp_password)
            try:
                ftp.cwd(settings.ftp_path)
            except:
                ftp.mkd(settings.ftp_path)
                ftp.cwd(settings.ftp_path)
            with open(filepath, 'rb') as f:
                ftp.storbinary(f'STOR {filepath.name}', f)
            ftp.quit()
            return True, f"{settings.ftp_path}/{filepath.name}"
    except Exception as e:
        return False, str(e)


def upload_to_s3(filepath: Path, settings: BackupSettings) -> tuple:
    """Upload backup file to AWS S3"""
    try:
        import boto3
        s3 = boto3.client(
            's3',
            region_name=settings.s3_region,
            aws_access_key_id=settings.s3_access_key,
            aws_secret_access_key=settings.s3_secret_key
        )
        s3_key = f"{settings.s3_path.strip('/')}/{filepath.name}"
        s3.upload_file(str(filepath), settings.s3_bucket, s3_key)
        return True, f"s3://{settings.s3_bucket}/{s3_key}"
    except Exception as e:
        return False, str(e)


# Backup encryption key (embedded in compiled binary - not visible to customers)
BACKUP_ENCRYPTION_KEY = b'OLT_M@n@g3r_S3cur3_B@ckup_K3y_2024!'  # 32 bytes for AES-256

def encrypt_backup(data: bytes) -> bytes:
    """Encrypt backup data using AES-256"""
    from cryptography.hazmat.primitives.ciphers import Cipher, algorithms, modes
    from cryptography.hazmat.backends import default_backend
    import os

    # Generate random IV
    iv = os.urandom(16)

    # Pad data to 16-byte boundary
    pad_len = 16 - (len(data) % 16)
    padded_data = data + bytes([pad_len] * pad_len)

    # Encrypt
    cipher = Cipher(algorithms.AES(BACKUP_ENCRYPTION_KEY[:32]), modes.CBC(iv), backend=default_backend())
    encryptor = cipher.encryptor()
    encrypted = encryptor.update(padded_data) + encryptor.finalize()

    # Return IV + encrypted data
    return iv + encrypted

def decrypt_backup(encrypted_data: bytes) -> bytes:
    """Decrypt backup data using AES-256"""
    from cryptography.hazmat.primitives.ciphers import Cipher, algorithms, modes
    from cryptography.hazmat.backends import default_backend

    # Extract IV and encrypted data
    iv = encrypted_data[:16]
    encrypted = encrypted_data[16:]

    # Decrypt
    cipher = Cipher(algorithms.AES(BACKUP_ENCRYPTION_KEY[:32]), modes.CBC(iv), backend=default_backend())
    decryptor = cipher.decryptor()
    padded_data = decryptor.update(encrypted) + decryptor.finalize()

    # Remove padding
    pad_len = padded_data[-1]
    return padded_data[:-pad_len]


def _is_within_dir(base: Path, target: Path) -> bool:
    """True if `target` resolves to a path inside `base` (zip/tar-slip guard)."""
    try:
        base_r = base.resolve()
        target_r = target.resolve()
        return base_r == target_r or base_r in target_r.parents
    except Exception:
        return False


def _safe_extract_zip(zf, dest) -> None:
    """extractall() but reject members that escape `dest` (zip-slip)."""
    dest = Path(dest)
    for name in zf.namelist():
        if not _is_within_dir(dest, dest / name):
            raise ValueError(f"Unsafe path in archive (zip-slip): {name!r}")
    zf.extractall(dest)


def _safe_extract_tar(tar, dest) -> None:
    """extractall() but reject members that escape `dest` (tar-slip)."""
    dest = Path(dest)
    for member in tar.getmembers():
        if not _is_within_dir(dest, dest / member.name):
            raise ValueError(f"Unsafe path in archive (tar-slip): {member.name!r}")
    tar.extractall(dest)


def create_system_backup_file(db: Session, include_uploads: bool = False) -> tuple:
    """Create a full system backup file (encrypted)"""
    import zipfile
    import sqlite3
    import shutil

    backup_dir = Path("/opt/olt-manager/backups")
    backup_dir.mkdir(exist_ok=True)

    # Use user's configured timezone for timestamp
    timestamp = format_timestamp_for_filename(db)
    backup_filename = f"olt_manager_backup_{timestamp}.zip"
    backup_path = backup_dir / backup_filename

    try:
        with zipfile.ZipFile(backup_path, 'w', zipfile.ZIP_DEFLATED) as zipf:
            # Backup database - check multiple possible locations
            db_paths = [
                Path("/opt/olt-manager/olt_manager.db"),  # Compiled binary installation
                Path("/opt/olt-manager/data/olt_manager.db"),  # Compiled with data folder
                Path("/root/olt-manager/backend/olt_manager.db"),
                Path("/opt/olt-manager/backend/olt_manager.db"),
                Path("/opt/olt-manager/backend/data/olt_manager.db"),
                Path("./olt_manager.db"),
                Path("./data/olt_manager.db")
            ]
            for db_path in db_paths:
                if db_path.exists():
                    # Create a copy to avoid locking issues
                    temp_db = backup_dir / f"temp_db_{timestamp}.db"
                    shutil.copy2(db_path, temp_db)
                    zipf.write(temp_db, "database/olt_manager.db")
                    temp_db.unlink()
                    break

            # Backup license files
            license_dir = Path("/etc/olt-manager")
            if license_dir.exists():
                for f in license_dir.iterdir():
                    if f.is_file():
                        zipf.write(f, f"config/{f.name}")

            # Backup uploads if requested - check multiple locations
            if include_uploads:
                upload_paths = [
                    Path("/opt/olt-manager/uploads"),  # Compiled binary installation
                    Path("/root/olt-manager/backend/uploads"),
                    Path("/opt/olt-manager/backend/uploads")
                ]
                for uploads_dir in upload_paths:
                    if uploads_dir.exists():
                        for f in uploads_dir.rglob("*"):
                            if f.is_file():
                                arcname = f"uploads/{f.relative_to(uploads_dir)}"
                                zipf.write(f, arcname)
                        break  # Only backup from first found location

            # Add backup metadata
            metadata = {
                "created_at": datetime.now().isoformat(),
                "version": "1.0",
                "includes_uploads": include_uploads
            }
            zipf.writestr("metadata.json", json.dumps(metadata, indent=2))

        # Encrypt the backup file
        with open(backup_path, 'rb') as f:
            zip_data = f.read()

        encrypted_data = encrypt_backup(zip_data)

        # Save encrypted backup with .bak extension (hides that it's a zip)
        encrypted_path = backup_path.with_suffix('.bak')
        with open(encrypted_path, 'wb') as f:
            # Add magic header to identify encrypted backups
            f.write(b'OLTBAK01')  # 8-byte magic header
            f.write(encrypted_data)

        # Remove unencrypted zip
        backup_path.unlink()

        file_size = encrypted_path.stat().st_size
        return True, encrypted_path, file_size
    except Exception as e:
        if backup_path.exists():
            backup_path.unlink()
        return False, None, str(e)


@app.get("/api/system-backups")
def get_system_backups(
    db: Session = Depends(get_db),
    current_user: User = Depends(require_auth)
):
    """Get list of system backups"""
    backups = db.query(SystemBackup).order_by(SystemBackup.created_at.desc()).all()

    return {
        "total": len(backups),
        "backups": [
            {
                "id": backup.id,
                "filename": backup.filename,
                "file_size": backup.file_size,
                "file_size_mb": round(backup.file_size / 1024 / 1024, 2) if backup.file_size else 0,
                "backup_type": backup.backup_type,
                "storage_type": backup.storage_type,
                "storage_path": backup.storage_path,
                "includes_db": backup.includes_db,
                "includes_config": backup.includes_config,
                "includes_uploads": backup.includes_uploads,
                "status": backup.status,
                "error_message": backup.error_message,
                "notes": backup.notes,
                "created_at": backup.created_at.isoformat()
            }
            for backup in backups
        ]
    }


@app.post("/api/system-backups")
async def create_system_backup(
    include_uploads: bool = Body(False),
    upload_to: Optional[str] = Body(None),  # 'ftp', 's3', or None for local only
    notes: Optional[str] = Body(None),
    db: Session = Depends(get_db),
    current_user: User = Depends(require_admin)
):
    """Create a full system backup"""
    # Get backup settings
    settings = db.query(BackupSettings).first()

    # Create backup file
    success, backup_path, file_size = create_system_backup_file(db, include_uploads)
    if not success:
        raise HTTPException(status_code=500, detail=f"Backup failed: {file_size}")

    storage_type = 'local'
    storage_path = str(backup_path)
    status = 'completed'
    error_message = None

    # Upload to remote storage if requested
    if upload_to and settings:
        if upload_to == 'ftp' and settings.ftp_host:
            storage_type = 'sftp' if settings.ftp_use_sftp else 'ftp'
            success, result = upload_to_ftp(backup_path, settings)
            if success:
                storage_path = result
            else:
                status = 'failed'
                error_message = f"FTP upload failed: {result}"
        elif upload_to == 's3' and settings.s3_bucket:
            storage_type = 's3'
            success, result = upload_to_s3(backup_path, settings)
            if success:
                storage_path = result
            else:
                status = 'failed'
                error_message = f"S3 upload failed: {result}"

    # Save backup record
    backup = SystemBackup(
        filename=backup_path.name,
        file_size=file_size if isinstance(file_size, int) else 0,
        backup_type='manual',
        storage_type=storage_type,
        storage_path=storage_path,
        includes_db=True,
        includes_config=True,
        includes_uploads=include_uploads,
        status=status,
        error_message=error_message,
        notes=notes,
        created_by=current_user.id
    )
    db.add(backup)
    db.commit()

    # Log event
    log_event(db, 'system_backup_created', 'system', backup.id, None,
              f"System backup created by {current_user.username} ({storage_type})")

    # Update backup settings with last backup info
    if settings:
        settings.last_backup_at = datetime.now()
        settings.last_backup_status = status
        db.commit()

    return {
        "success": status == 'completed',
        "backup_id": backup.id,
        "filename": backup.filename,
        "file_size_mb": round(backup.file_size / 1024 / 1024, 2) if backup.file_size else 0,
        "storage_type": storage_type,
        "storage_path": storage_path,
        "error": error_message
    }


@app.get("/api/system-backups/{backup_id}/download")
async def download_system_backup(
    backup_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_auth)
):
    """Download a system backup file"""
    from fastapi.responses import FileResponse

    backup = db.query(SystemBackup).filter(SystemBackup.id == backup_id).first()
    if not backup:
        raise HTTPException(status_code=404, detail="Backup not found")

    if backup.storage_type != 'local':
        raise HTTPException(status_code=400, detail="Cannot download remote backups directly. Use the storage path to access.")

    filepath = Path(f"/opt/olt-manager/backups/{backup.filename}")
    if not filepath.exists():
        raise HTTPException(status_code=404, detail="Backup file not found on disk")

    return FileResponse(filepath, filename=backup.filename, media_type='application/zip')


@app.post("/api/system-backups/{backup_id}/restore")
async def restore_system_backup(
    backup_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_admin)
):
    """Restore from a system backup"""
    import zipfile
    import shutil
    import sqlite3

    backup = db.query(SystemBackup).filter(SystemBackup.id == backup_id).first()
    if not backup:
        raise HTTPException(status_code=404, detail="Backup not found")

    backup_path = Path(f"/opt/olt-manager/backups/{backup.filename}")
    if not backup_path.exists():
        raise HTTPException(status_code=404, detail="Backup file not found")

    # Log event before restore (database will be replaced)
    log_event(db, 'system_restore_started', 'system', backup_id, None,
              f"System restore from '{backup.filename}' initiated by {current_user.username}")

    # Save all current backup records BEFORE restoring (they will be lost when DB is replaced)
    all_backups = db.query(SystemBackup).all()
    backup_records = []
    for b in all_backups:
        backup_records.append({
            'id': b.id,
            'filename': b.filename,
            'file_size': b.file_size,
            'backup_type': b.backup_type,
            'storage_type': b.storage_type,
            'storage_path': b.storage_path,
            'includes_db': b.includes_db,
            'includes_config': b.includes_config,
            'includes_uploads': b.includes_uploads,
            'status': b.status,
            'error_message': b.error_message,
            'notes': b.notes,
            'created_by': b.created_by,
            'created_at': b.created_at.isoformat() if b.created_at else None
        })

    try:
        # Check if backup is encrypted
        with open(backup_path, 'rb') as f:
            header = f.read(8)
            is_encrypted = header == b'OLTBAK01'

        temp_dir = Path("/tmp/olt_restore")
        if temp_dir.exists():
            shutil.rmtree(temp_dir)
        temp_dir.mkdir()

        if is_encrypted:
            # Decrypt the backup first
            with open(backup_path, 'rb') as f:
                f.read(8)  # Skip magic header
                encrypted_data = f.read()

            zip_data = decrypt_backup(encrypted_data)

            # Write decrypted zip to temp file
            temp_zip = temp_dir / "backup.zip"
            with open(temp_zip, 'wb') as f:
                f.write(zip_data)

            # Extract from decrypted zip
            with zipfile.ZipFile(temp_zip, 'r') as zipf:
                _safe_extract_zip(zipf, temp_dir)

            # Remove temp zip
            temp_zip.unlink()
        else:
            # Old unencrypted backup - extract directly
            with zipfile.ZipFile(backup_path, 'r') as zipf:
                _safe_extract_zip(zipf, temp_dir)

        # Restore database - find the actual database location
        db_backup = temp_dir / "database" / "olt_manager.db"
        if db_backup.exists():
            # Find actual database location
            db_target = None
            db_paths = [
                Path("/opt/olt-manager/olt_manager.db"),  # Compiled binary installation
                Path("/opt/olt-manager/data/olt_manager.db"),  # Compiled with data folder
                Path("/root/olt-manager/backend/olt_manager.db"),
                Path("/opt/olt-manager/backend/olt_manager.db"),
                Path("/opt/olt-manager/backend/data/olt_manager.db")
            ]
            for p in db_paths:
                if p.exists():
                    db_target = p
                    break

            # If no existing db found, use the first path
            if not db_target:
                db_target = db_paths[0]
                db_target.parent.mkdir(parents=True, exist_ok=True)

            # Close current db connections
            db.close()
            # Backup current db just in case
            if db_target.exists():
                shutil.copy2(db_target, db_target.with_suffix('.db.bak'))
            # Restore
            shutil.copy2(db_backup, db_target)

        # Restore config files
        config_dir = temp_dir / "config"
        if config_dir.exists():
            target_config = Path("/etc/olt-manager")
            target_config.mkdir(exist_ok=True)
            for f in config_dir.iterdir():
                if f.is_file():
                    shutil.copy2(f, target_config / f.name)

        # Restore uploads if present - find actual uploads location
        uploads_backup = temp_dir / "uploads"
        if uploads_backup.exists():
            # Find actual uploads location
            uploads_target = None
            upload_paths = [
                Path("/opt/olt-manager/uploads"),  # Compiled binary installation
                Path("/root/olt-manager/backend/uploads"),
                Path("/opt/olt-manager/backend/uploads")
            ]
            for p in upload_paths:
                if p.exists():
                    uploads_target = p
                    break

            if not uploads_target:
                uploads_target = upload_paths[0]

            uploads_target.mkdir(parents=True, exist_ok=True)
            for f in uploads_backup.rglob("*"):
                if f.is_file():
                    target = uploads_target / f.relative_to(uploads_backup)
                    target.parent.mkdir(parents=True, exist_ok=True)
                    shutil.copy2(f, target)

        # Cleanup temp directory
        shutil.rmtree(temp_dir)

        # Re-insert all backup records into the restored database
        # This ensures backup history is preserved even after restore
        if db_target and backup_records:
            try:
                conn = sqlite3.connect(str(db_target))
                cursor = conn.cursor()

                # Ensure table exists
                cursor.execute('''
                    CREATE TABLE IF NOT EXISTS system_backups (
                        id INTEGER PRIMARY KEY,
                        filename VARCHAR(255) NOT NULL,
                        file_size INTEGER,
                        backup_type VARCHAR(20) DEFAULT 'manual',
                        storage_type VARCHAR(20) DEFAULT 'local',
                        storage_path VARCHAR(500),
                        includes_db BOOLEAN DEFAULT 1,
                        includes_config BOOLEAN DEFAULT 1,
                        includes_uploads BOOLEAN DEFAULT 0,
                        status VARCHAR(20) DEFAULT 'completed',
                        error_message VARCHAR(500),
                        notes VARCHAR(500),
                        created_by INTEGER,
                        created_at DATETIME
                    )
                ''')

                # Insert or replace backup records
                for rec in backup_records:
                    cursor.execute('''
                        INSERT OR REPLACE INTO system_backups
                        (id, filename, file_size, backup_type, storage_type, storage_path,
                         includes_db, includes_config, includes_uploads, status,
                         error_message, notes, created_by, created_at)
                        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                    ''', (
                        rec['id'], rec['filename'], rec['file_size'], rec['backup_type'],
                        rec['storage_type'], rec['storage_path'], rec['includes_db'],
                        rec['includes_config'], rec['includes_uploads'], rec['status'],
                        rec['error_message'], rec['notes'], rec['created_by'], rec['created_at']
                    ))

                conn.commit()
                conn.close()
            except Exception as sql_err:
                print(f"Warning: Could not restore backup records: {sql_err}")

        return {
            "success": True,
            "message": "Backup restored successfully. Please restart the service for changes to take effect.",
            "restart_command": "systemctl restart olt-backend"
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Restore failed: {str(e)}")


@app.delete("/api/system-backups/{backup_id}")
async def delete_system_backup(
    backup_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_admin)
):
    """Delete a system backup"""
    backup = db.query(SystemBackup).filter(SystemBackup.id == backup_id).first()
    if not backup:
        raise HTTPException(status_code=404, detail="Backup not found")

    # Store info for logging
    filename = backup.filename

    # Delete local file if exists
    if backup.storage_type == 'local':
        filepath = Path(f"/opt/olt-manager/backups/{backup.filename}")
        if filepath.exists():
            filepath.unlink()

    db.delete(backup)
    db.commit()

    # Log event
    log_event(db, 'system_backup_deleted', 'system', backup_id, None,
              f"System backup '{filename}' deleted by {current_user.username}")

    return {"success": True, "message": "Backup deleted"}


# ============ BACKUP SETTINGS ============

@app.get("/api/backup-settings")
def get_backup_settings(
    db: Session = Depends(get_db),
    current_user: User = Depends(require_admin)
):
    """Get backup settings"""
    settings = db.query(BackupSettings).first()
    if not settings:
        # Create default settings
        settings = BackupSettings()
        db.add(settings)
        db.commit()
        db.refresh(settings)

    return {
        "auto_backup_enabled": settings.auto_backup_enabled,
        "backup_frequency": settings.backup_frequency,
        "backup_time": settings.backup_time,
        "backup_day": settings.backup_day,
        "retention_days": settings.retention_days,
        "backup_database": settings.backup_database,
        "backup_config": settings.backup_config,
        "backup_uploads": settings.backup_uploads,
        "storage_type": settings.storage_type,
        "local_path": settings.local_path,
        "ftp_host": settings.ftp_host,
        "ftp_port": settings.ftp_port,
        "ftp_username": settings.ftp_username,
        "ftp_password": "***" if settings.ftp_password else None,
        "ftp_path": settings.ftp_path,
        "ftp_use_sftp": settings.ftp_use_sftp,
        "s3_bucket": settings.s3_bucket,
        "s3_region": settings.s3_region,
        "s3_access_key": settings.s3_access_key[:8] + "***" if settings.s3_access_key else None,
        "s3_secret_key": "***" if settings.s3_secret_key else None,
        "s3_path": settings.s3_path,
        "last_backup_at": settings.last_backup_at.isoformat() if settings.last_backup_at else None,
        "last_backup_status": settings.last_backup_status,
        "next_backup_at": settings.next_backup_at.isoformat() if settings.next_backup_at else None
    }


@app.put("/api/backup-settings")
def update_backup_settings(
    auto_backup_enabled: Optional[bool] = Body(None),
    backup_frequency: Optional[str] = Body(None),
    backup_time: Optional[str] = Body(None),
    backup_day: Optional[int] = Body(None),
    retention_days: Optional[int] = Body(None),
    backup_database: Optional[bool] = Body(None),
    backup_config: Optional[bool] = Body(None),
    backup_uploads: Optional[bool] = Body(None),
    storage_type: Optional[str] = Body(None),
    local_path: Optional[str] = Body(None),
    ftp_host: Optional[str] = Body(None),
    ftp_port: Optional[int] = Body(None),
    ftp_username: Optional[str] = Body(None),
    ftp_password: Optional[str] = Body(None),
    ftp_path: Optional[str] = Body(None),
    ftp_use_sftp: Optional[bool] = Body(None),
    s3_bucket: Optional[str] = Body(None),
    s3_region: Optional[str] = Body(None),
    s3_access_key: Optional[str] = Body(None),
    s3_secret_key: Optional[str] = Body(None),
    s3_path: Optional[str] = Body(None),
    db: Session = Depends(get_db),
    current_user: User = Depends(require_admin)
):
    """Update backup settings"""
    settings = db.query(BackupSettings).first()
    if not settings:
        settings = BackupSettings()
        db.add(settings)

    # Update fields if provided
    if auto_backup_enabled is not None:
        settings.auto_backup_enabled = auto_backup_enabled
    if backup_frequency:
        settings.backup_frequency = backup_frequency
    if backup_time:
        settings.backup_time = backup_time
    if backup_day is not None:
        settings.backup_day = backup_day
    if retention_days is not None:
        settings.retention_days = retention_days
    if backup_database is not None:
        settings.backup_database = backup_database
    if backup_config is not None:
        settings.backup_config = backup_config
    if backup_uploads is not None:
        settings.backup_uploads = backup_uploads
    if storage_type:
        settings.storage_type = storage_type
    if local_path:
        settings.local_path = local_path
    if ftp_host is not None:
        settings.ftp_host = ftp_host
    if ftp_port is not None:
        settings.ftp_port = ftp_port
    if ftp_username is not None:
        settings.ftp_username = ftp_username
    if ftp_password and ftp_password != "***":
        settings.ftp_password = ftp_password
    if ftp_path is not None:
        settings.ftp_path = ftp_path
    if ftp_use_sftp is not None:
        settings.ftp_use_sftp = ftp_use_sftp
    if s3_bucket is not None:
        settings.s3_bucket = s3_bucket
    if s3_region is not None:
        settings.s3_region = s3_region
    if s3_access_key and not s3_access_key.endswith("***"):
        settings.s3_access_key = s3_access_key
    if s3_secret_key and s3_secret_key != "***":
        settings.s3_secret_key = s3_secret_key
    if s3_path is not None:
        settings.s3_path = s3_path

    # Calculate next backup time
    if settings.auto_backup_enabled:
        settings.next_backup_at = calculate_next_backup_time(settings)

    db.commit()
    return {"success": True, "message": "Backup settings updated"}


def calculate_next_backup_time(settings: BackupSettings) -> datetime:
    """Calculate the next scheduled backup time"""
    now = datetime.now()
    backup_hour, backup_minute = map(int, settings.backup_time.split(':'))

    if settings.backup_frequency == 'hourly':
        next_time = now.replace(minute=backup_minute, second=0, microsecond=0)
        if next_time <= now:
            next_time += timedelta(hours=1)
    elif settings.backup_frequency == 'daily':
        next_time = now.replace(hour=backup_hour, minute=backup_minute, second=0, microsecond=0)
        if next_time <= now:
            next_time += timedelta(days=1)
    elif settings.backup_frequency == 'weekly':
        next_time = now.replace(hour=backup_hour, minute=backup_minute, second=0, microsecond=0)
        days_ahead = (settings.backup_day or 0) - now.weekday()
        if days_ahead <= 0:
            days_ahead += 7
        next_time += timedelta(days=days_ahead)
    elif settings.backup_frequency == 'monthly':
        import calendar
        day = settings.backup_day or 1
        # Clamp the requested day to the month's last day (e.g. 31 -> 30/28),
        # else datetime.replace(day=31) throws in short months and auto-backup
        # silently stalls forever.
        last_day = calendar.monthrange(now.year, now.month)[1]
        next_time = now.replace(day=min(day, last_day), hour=backup_hour,
                                minute=backup_minute, second=0, microsecond=0)
        if next_time <= now:
            year = now.year + 1 if now.month == 12 else now.year
            month = 1 if now.month == 12 else now.month + 1
            last_day = calendar.monthrange(year, month)[1]
            next_time = next_time.replace(year=year, month=month, day=min(day, last_day))
    else:
        next_time = now + timedelta(days=1)

    return next_time


@app.post("/api/backup-settings/test-ftp")
async def test_ftp_connection(
    ftp_host: str = Body(...),
    ftp_port: int = Body(21),
    ftp_username: str = Body(...),
    ftp_password: str = Body(...),
    ftp_path: str = Body("/"),
    ftp_use_sftp: bool = Body(False),
    db: Session = Depends(get_db),
    current_user: User = Depends(require_admin)
):
    """Test FTP/SFTP connection"""
    try:
        if ftp_use_sftp:
            import paramiko
            transport = paramiko.Transport((ftp_host, ftp_port or 22))
            transport.connect(username=ftp_username, password=ftp_password)
            sftp = paramiko.SFTPClient.from_transport(transport)
            sftp.listdir(ftp_path)
            sftp.close()
            transport.close()
        else:
            import ftplib
            ftp = ftplib.FTP()
            ftp.connect(ftp_host, ftp_port or 21)
            ftp.login(ftp_username, ftp_password)
            ftp.cwd(ftp_path)
            ftp.quit()
        return {"success": True, "message": "Connection successful"}
    except Exception as e:
        return {"success": False, "message": str(e)}


@app.post("/api/backup-settings/test-s3")
async def test_s3_connection(
    s3_bucket: str = Body(...),
    s3_region: str = Body(...),
    s3_access_key: str = Body(...),
    s3_secret_key: str = Body(...),
    s3_path: str = Body("/"),
    db: Session = Depends(get_db),
    current_user: User = Depends(require_admin)
):
    """Test AWS S3 connection"""
    try:
        import boto3
        s3 = boto3.client(
            's3',
            region_name=s3_region,
            aws_access_key_id=s3_access_key,
            aws_secret_access_key=s3_secret_key
        )
        # Test by listing bucket contents
        s3.list_objects_v2(Bucket=s3_bucket, Prefix=s3_path.strip('/'), MaxKeys=1)
        return {"success": True, "message": "Connection successful"}
    except Exception as e:
        return {"success": False, "message": str(e)}


@app.post("/api/system-backups/upload")
async def upload_backup_file(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    current_user: User = Depends(require_admin)
):
    """Upload a backup file for restore"""
    if not file.filename.endswith('.zip'):
        raise HTTPException(status_code=400, detail="Only .zip backup files are supported")

    backup_dir = Path("/opt/olt-manager/backups")
    backup_dir.mkdir(exist_ok=True)

    # Never trust the client filename for the path — strip any directory
    # components so "../../etc/x.zip" can't escape the backups dir.
    safe_name = Path(file.filename).name
    if not safe_name or not safe_name.endswith('.zip'):
        raise HTTPException(status_code=400, detail="Invalid backup filename")

    # Save uploaded file
    filepath = backup_dir / safe_name
    with open(filepath, 'wb') as f:
        content = await file.read()
        f.write(content)

    # Create backup record
    backup = SystemBackup(
        filename=safe_name,
        file_size=len(content),
        backup_type='uploaded',
        storage_type='local',
        storage_path=str(filepath),
        includes_db=True,
        includes_config=True,
        status='completed',
        notes='Uploaded for restore',
        created_by=current_user.id
    )
    db.add(backup)
    db.commit()

    return {
        "success": True,
        "backup_id": backup.id,
        "filename": file.filename,
        "file_size_mb": round(len(content) / 1024 / 1024, 2)
    }


# ============ FEATURE 7: Scheduled Tasks ============

def compute_task_next_run(schedule_type, schedule_time, schedule_day, from_time=None):
    """Next fire time for a ScheduledTask (naive local, matches next_run column)."""
    import calendar
    now = from_time or datetime.now()
    try:
        hh, mm = (int(x) for x in (schedule_time or "00:00").split(":")[:2])
    except Exception:
        hh, mm = 0, 0
    st = (schedule_type or "daily").lower()
    if st == "hourly":
        nt = now.replace(minute=mm, second=0, microsecond=0)
        if nt <= now:
            nt += timedelta(hours=1)
    elif st == "weekly":
        nt = now.replace(hour=hh, minute=mm, second=0, microsecond=0)
        days_ahead = ((schedule_day or 0) - now.weekday()) % 7
        if days_ahead == 0 and nt <= now:
            days_ahead = 7
        nt += timedelta(days=days_ahead)
    elif st == "monthly":
        day = schedule_day or 1
        last = calendar.monthrange(now.year, now.month)[1]
        nt = now.replace(day=min(day, last), hour=hh, minute=mm, second=0, microsecond=0)
        if nt <= now:
            y = now.year + 1 if now.month == 12 else now.year
            m = 1 if now.month == 12 else now.month + 1
            last = calendar.monthrange(y, m)[1]
            nt = nt.replace(year=y, month=m, day=min(day, last))
    else:  # daily
        nt = now.replace(hour=hh, minute=mm, second=0, microsecond=0)
        if nt <= now:
            nt += timedelta(days=1)
    return nt


def _execute_scheduled_task(task, db) -> str:
    """Run a scheduled task's action. Returns a short status string."""
    tt = (task.task_type or "").lower()
    if tt in ("backup", "system_backup"):
        create_system_backup_file(db)
        return "system backup created"
    if tt in ("reboot_onu",) and task.target_id:
        onu = db.query(ONU).filter(ONU.id == task.target_id).first()
        if not onu:
            return f"ONU {task.target_id} not found"
        olt = db.query(OLT).filter(OLT.id == onu.olt_id).first()
        get_driver(olt).reboot_onu(onu.pon_port, onu.onu_id)
        return f"rebooted ONU {onu.mac_address}"
    if tt in ("reboot_olt",) and task.target_id:
        olt = db.query(OLT).filter(OLT.id == task.target_id).first()
        if not olt:
            return f"OLT {task.target_id} not found"
        # Reboot via the driver's OLT-reboot if available.
        drv = get_driver(olt)
        reboot = getattr(drv, "reboot_olt", None)
        if callable(reboot):
            reboot()
            return f"rebooted OLT {olt.name}"
        return "reboot_olt not supported by this driver"
    return f"unsupported task_type '{task.task_type}' (skipped)"


async def run_due_scheduled_tasks(db_session_factory):
    """Dispatcher: run enabled tasks whose next_run has passed, then reschedule.

    NOTE: uses an unscoped session (fine for the single-tenant/SQLite build); a
    SaaS/Postgres deploy should iterate tenants like poll_all_tenants does.
    """
    db = db_session_factory()
    try:
        now = datetime.now()
        tasks = db.query(ScheduledTask).filter(ScheduledTask.is_enabled == True).all()
        changed = False
        for t in tasks:
            if t.next_run is None:
                t.next_run = compute_task_next_run(t.schedule_type, t.schedule_time, t.schedule_day, now)
                changed = True
        if changed:
            db.commit()

        due = [t for t in tasks if t.next_run is not None and t.next_run <= now]
        for t in due:
            try:
                result = _execute_scheduled_task(t, db)
                logger.info(f"[scheduler] Ran '{t.name}' ({t.task_type}): {result}")
            except Exception as e:
                logger.error(f"[scheduler] Task '{t.name}' failed: {e}")
                db.rollback()
            t.last_run = now
            t.next_run = compute_task_next_run(t.schedule_type, t.schedule_time, t.schedule_day, now)
        if due:
            db.commit()
    except Exception as e:
        logger.error(f"[scheduler] dispatcher error: {e}")
    finally:
        db.close()


@app.get("/api/scheduled-tasks")
def get_scheduled_tasks(
    db: Session = Depends(get_db),
    current_user: User = Depends(require_auth)
):
    """Get all scheduled tasks"""
    tasks = db.query(ScheduledTask).all()
    return {
        "total": len(tasks),
        "tasks": [
            {
                "id": task.id,
                "name": task.name,
                "task_type": task.task_type,
                "target_type": task.target_type,
                "target_id": task.target_id,
                "schedule_type": task.schedule_type,
                "schedule_time": task.schedule_time,
                "schedule_day": task.schedule_day,
                "is_enabled": task.is_enabled,
                "last_run": task.last_run.isoformat() if task.last_run else None,
                "next_run": task.next_run.isoformat() if task.next_run else None
            }
            for task in tasks
        ]
    }


@app.post("/api/scheduled-tasks")
def create_scheduled_task(
    name: str = Body(...),
    task_type: str = Body(...),
    target_type: Optional[str] = Body(None),
    target_id: Optional[int] = Body(None),
    schedule_type: str = Body(...),
    schedule_time: str = Body(...),
    schedule_day: Optional[int] = Body(None),
    db: Session = Depends(get_db),
    current_user: User = Depends(require_admin)
):
    """Create a new scheduled task"""
    task = ScheduledTask(
        name=name,
        task_type=task_type,
        target_type=target_type,
        target_id=target_id,
        schedule_type=schedule_type,
        schedule_time=schedule_time,
        schedule_day=schedule_day,
        created_by=current_user.id,
        next_run=compute_task_next_run(schedule_type, schedule_time, schedule_day),
    )
    db.add(task)
    db.commit()
    db.refresh(task)

    return {"success": True, "id": task.id, "message": "Scheduled task created"}


@app.put("/api/scheduled-tasks/{task_id}/toggle")
def toggle_scheduled_task(
    task_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_admin)
):
    """Enable/disable a scheduled task"""
    task = db.query(ScheduledTask).filter(ScheduledTask.id == task_id).first()
    if not task:
        raise HTTPException(status_code=404, detail="Task not found")

    task.is_enabled = not task.is_enabled
    db.commit()

    return {"success": True, "is_enabled": task.is_enabled}


@app.delete("/api/scheduled-tasks/{task_id}")
def delete_scheduled_task(
    task_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_admin)
):
    """Delete a scheduled task"""
    task = db.query(ScheduledTask).filter(ScheduledTask.id == task_id).first()
    if not task:
        raise HTTPException(status_code=404, detail="Task not found")

    db.delete(task)
    db.commit()
    return {"success": True, "message": "Scheduled task deleted"}


# ============ FEATURE 8: Event History Log ============

def log_event(db: Session, event_type: str, entity_type: str, entity_id: int,
              olt_id: int = None, description: str = None, details: dict = None):
    """Helper function to log events"""
    try:
        # Use timezone from settings
        current_time = get_current_time_in_timezone(db)
        event = EventLog(
            event_type=event_type,
            entity_type=entity_type,
            entity_id=entity_id,
            olt_id=olt_id,
            description=description,
            details=json.dumps(details) if details else None,
            created_at=current_time
        )
        db.add(event)
        db.commit()
    except Exception as e:
        # Don't let logging failures break the main operation
        db.rollback()
        logger.warning(f"Failed to log event: {e}")


@app.get("/api/events")
def get_events(
    event_type: Optional[str] = None,
    entity_type: Optional[str] = None,
    olt_id: Optional[int] = None,
    limit: int = Query(100, le=1000),
    offset: int = 0,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_auth)
):
    """Get event history"""
    query = db.query(EventLog)

    if event_type:
        query = query.filter(EventLog.event_type == event_type)
    if entity_type:
        query = query.filter(EventLog.entity_type == entity_type)
    if olt_id:
        query = query.filter(EventLog.olt_id == olt_id)

    total = query.count()
    events = query.order_by(EventLog.created_at.desc()).offset(offset).limit(limit).all()

    return {
        "total": total,
        "events": [
            {
                "id": event.id,
                "event_type": event.event_type,
                "entity_type": event.entity_type,
                "entity_id": event.entity_id,
                "olt_id": event.olt_id,
                "description": event.description,
                "details": json.loads(event.details) if event.details else None,
                "created_at": event.created_at.isoformat()
            }
            for event in events
        ]
    }


@app.get("/api/events/onu/{onu_id}")
def get_onu_events(
    onu_id: int,
    limit: int = Query(50, le=500),
    db: Session = Depends(get_db),
    current_user: User = Depends(require_auth)
):
    """Get event history for a specific ONU"""
    events = db.query(EventLog).filter(
        EventLog.entity_type == 'onu',
        EventLog.entity_id == onu_id
    ).order_by(EventLog.created_at.desc()).limit(limit).all()

    return {
        "onu_id": onu_id,
        "total": len(events),
        "events": [
            {
                "id": event.id,
                "event_type": event.event_type,
                "description": event.description,
                "details": json.loads(event.details) if event.details else None,
                "created_at": event.created_at.isoformat()
            }
            for event in events
        ]
    }


@app.delete("/api/events/cleanup")
def cleanup_old_events(
    days: int = Query(30, ge=1, le=365),
    db: Session = Depends(get_db),
    current_user: User = Depends(require_admin)
):
    """Delete events older than specified days"""
    cutoff = datetime.utcnow() - timedelta(days=days)
    deleted = db.query(EventLog).filter(EventLog.created_at < cutoff).delete()
    db.commit()
    return {"success": True, "deleted": deleted, "message": f"Deleted {deleted} events older than {days} days"}


# ============ FEATURE 9: Customer Self-Portal (API) ============

@app.get("/api/portal/status/{mac_address}")
def get_customer_status(
    mac_address: str,
    db: Session = Depends(get_db)
):
    """Public endpoint for customers to check their ONU status by MAC address"""
    # Normalize MAC address format
    mac = mac_address.upper().replace('-', ':')

    onu = db.query(ONU).filter(ONU.mac_address == mac).first()
    if not onu:
        raise HTTPException(status_code=404, detail="ONU not found")

    return {
        "mac_address": onu.mac_address,
        "description": onu.description,
        "is_online": onu.is_online,
        "signal_quality": "good" if onu.rx_power and onu.rx_power > -25 else "fair" if onu.rx_power and onu.rx_power > -28 else "poor",
        "rx_power": onu.rx_power,
        "last_seen": onu.last_seen.isoformat() if onu.last_seen else None,
        "olt_name": onu.olt.name if onu.olt else None,
        "region": onu.region.name if onu.region else None
    }


@app.get("/api/portal/speed-test/{mac_address}")
def get_customer_traffic(
    mac_address: str,
    db: Session = Depends(get_db)
):
    """Get current traffic for customer's ONU"""
    mac = mac_address.upper().replace('-', ':')

    onu = db.query(ONU).filter(ONU.mac_address == mac).first()
    if not onu:
        raise HTTPException(status_code=404, detail="ONU not found")

    # Get latest traffic history
    traffic = db.query(TrafficHistory).filter(
        TrafficHistory.entity_type == 'onu',
        TrafficHistory.onu_db_id == onu.id
    ).order_by(TrafficHistory.timestamp.desc()).first()

    return {
        "mac_address": onu.mac_address,
        "download_kbps": traffic.rx_kbps if traffic else 0,
        "upload_kbps": traffic.tx_kbps if traffic else 0,
        "download_mbps": round(traffic.rx_kbps / 1000, 2) if traffic else 0,
        "upload_mbps": round(traffic.tx_kbps / 1000, 2) if traffic else 0,
        "timestamp": traffic.timestamp.isoformat() if traffic else None
    }


# ============ FEATURE 10: Mobile App API Endpoints ============

@app.get("/api/mobile/dashboard")
def mobile_dashboard(
    db: Session = Depends(get_db),
    current_user: User = Depends(require_auth)
):
    """Optimized dashboard for mobile app"""
    # Get counts
    total_olts = db.query(func.count(OLT.id)).scalar()
    online_olts = db.query(func.count(OLT.id)).filter(OLT.is_online == True).scalar()
    total_onus = db.query(func.count(ONU.id)).scalar()
    online_onus = db.query(func.count(ONU.id)).filter(ONU.is_online == True).scalar()

    # Get OLTs with issues
    olts_with_issues = db.query(OLT).filter(OLT.is_online == False).all()

    # Get recent events
    recent_events = db.query(EventLog).order_by(EventLog.created_at.desc()).limit(10).all()

    return {
        "summary": {
            "olts": {"total": total_olts, "online": online_olts, "offline": total_olts - online_olts},
            "onus": {"total": total_onus, "online": online_onus, "offline": total_onus - online_onus}
        },
        "issues": [
            {"id": olt.id, "name": olt.name, "ip": olt.ip_address, "error": olt.last_error}
            for olt in olts_with_issues
        ],
        "recent_events": [
            {
                "type": event.event_type,
                "description": event.description,
                "time": event.created_at.isoformat()
            }
            for event in recent_events
        ]
    }


@app.get("/api/mobile/onus")
def mobile_onu_list(
    search: Optional[str] = None,
    online_only: bool = False,
    limit: int = Query(50, le=200),
    db: Session = Depends(get_db),
    current_user: User = Depends(require_auth)
):
    """Simplified ONU list for mobile app"""
    query = db.query(ONU)

    if search:
        query = query.filter(
            or_(
                ONU.mac_address.ilike(f"%{search}%"),
                ONU.description.ilike(f"%{search}%")
            )
        )
    if online_only:
        query = query.filter(ONU.is_online == True)

    onus = query.limit(limit).all()

    return {
        "total": len(onus),
        "onus": [
            {
                "id": onu.id,
                "mac": onu.mac_address,
                "name": onu.description,
                "online": onu.is_online,
                "signal": onu.rx_power,
                "olt": onu.olt.name if onu.olt else None
            }
            for onu in onus
        ]
    }


@app.post("/api/mobile/onu/{onu_id}/reboot")
async def mobile_reboot_onu(
    onu_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_auth)
):
    """Quick ONU reboot for mobile app"""
    onu = db.query(ONU).filter(ONU.id == onu_id).first()
    if not onu:
        raise HTTPException(status_code=404, detail="ONU not found")

    try:
        connector = OLTConnector(
            onu.olt.ip_address,
            onu.olt.username,
            decrypt_sensitive(onu.olt.password)
        )
        success = connector.reboot_onu(onu.pon_port, onu.onu_id)

        if success:
            # Reset online_since and mark offline so polling detects online transition
            onu.online_since = None
            onu.is_online = False  # Mark offline so next poll sets online_since when ONU comes back
            db.commit()
            log_event(db, 'onu_reboot', 'onu', onu.id, onu.olt_id,
                     f"ONU {onu.mac_address} rebooted via mobile app")

        return {"success": success}
    except Exception as e:
        return {"success": False, "error": str(e)}


@app.get("/api/mobile/notifications")
def mobile_notifications(
    db: Session = Depends(get_db),
    current_user: User = Depends(require_auth)
):
    """Get recent notifications for mobile app"""
    # Get recent alerts and events
    recent_events = db.query(EventLog).filter(
        EventLog.event_type.in_(['onu_offline', 'olt_offline', 'signal_low', 'signal_critical'])
    ).order_by(EventLog.created_at.desc()).limit(50).all()

    return {
        "notifications": [
            {
                "id": event.id,
                "type": event.event_type,
                "title": event.event_type.replace('_', ' ').title(),
                "message": event.description,
                "time": event.created_at.isoformat(),
                "read": False
            }
            for event in recent_events
        ]
    }


# ============ SERVICES MANAGEMENT ============

@app.get("/api/services/status")
def get_service_status(
    current_user: User = Depends(require_admin)
):
    """Get OLT Manager service status"""
    import subprocess

    try:
        # Check olt-manager service status
        result = subprocess.run(
            ['systemctl', 'is-active', 'olt-manager'],
            capture_output=True, text=True, timeout=10
        )
        service_status = result.stdout.strip()

        # Get uptime
        uptime_result = subprocess.run(
            ['systemctl', 'show', 'olt-manager', '--property=ActiveEnterTimestamp'],
            capture_output=True, text=True, timeout=10
        )
        uptime_line = uptime_result.stdout.strip()
        service_uptime = uptime_line.split('=')[1] if '=' in uptime_line else None

        # Get system uptime
        with open('/proc/uptime', 'r') as f:
            system_uptime_seconds = float(f.read().split()[0])

        # Get memory info
        mem_result = subprocess.run(
            ['free', '-m'],
            capture_output=True, text=True, timeout=10
        )
        mem_lines = mem_result.stdout.strip().split('\n')
        mem_parts = mem_lines[1].split()
        total_mem = int(mem_parts[1])
        used_mem = int(mem_parts[2])

        # Get disk info
        disk_result = subprocess.run(
            ['df', '-h', '/'],
            capture_output=True, text=True, timeout=10
        )
        disk_lines = disk_result.stdout.strip().split('\n')
        disk_parts = disk_lines[1].split()
        disk_total = disk_parts[1]
        disk_used = disk_parts[2]
        disk_percent = disk_parts[4]

        # Get CPU load averages
        with open('/proc/loadavg', 'r') as f:
            loadavg = f.read().split()
            cpu_load_1 = float(loadavg[0])
            cpu_load_5 = float(loadavg[1])
            cpu_load_15 = float(loadavg[2])

        # Get CPU count
        import os
        cpu_count = os.cpu_count() or 1

        # Get actual CPU usage from /proc/stat
        cpu_percent = 0.0
        try:
            with open('/proc/stat', 'r') as f:
                line = f.readline()  # First line is total CPU
                parts = line.split()
                # cpu user nice system idle iowait irq softirq
                user = int(parts[1])
                nice = int(parts[2])
                system = int(parts[3])
                idle = int(parts[4])
                iowait = int(parts[5]) if len(parts) > 5 else 0

                total = user + nice + system + idle + iowait
                busy = user + nice + system

                # Read again after short delay to get delta
                import time
                time.sleep(0.1)

                with open('/proc/stat', 'r') as f2:
                    line2 = f2.readline()
                    parts2 = line2.split()
                    user2 = int(parts2[1])
                    nice2 = int(parts2[2])
                    system2 = int(parts2[3])
                    idle2 = int(parts2[4])
                    iowait2 = int(parts2[5]) if len(parts2) > 5 else 0

                    total2 = user2 + nice2 + system2 + idle2 + iowait2
                    busy2 = user2 + nice2 + system2

                    total_delta = total2 - total
                    busy_delta = busy2 - busy

                    if total_delta > 0:
                        cpu_percent = round(busy_delta / total_delta * 100, 1)
        except:
            cpu_percent = 0.0

        return {
            "service": {
                "name": "olt-manager",
                "status": service_status,
                "uptime": service_uptime
            },
            "system": {
                "uptime_seconds": int(system_uptime_seconds),
                "uptime_days": round(system_uptime_seconds / 86400, 1),
                "memory_total_mb": total_mem,
                "memory_used_mb": used_mem,
                "memory_percent": round(used_mem / total_mem * 100, 1),
                "disk_total": disk_total,
                "disk_used": disk_used,
                "disk_percent": disk_percent,
                "cpu_load_1": cpu_load_1,
                "cpu_load_5": cpu_load_5,
                "cpu_load_15": cpu_load_15,
                "cpu_count": cpu_count,
                "cpu_percent": cpu_percent
            }
        }
    except Exception as e:
        logger.error(f"Error getting service status: {e}")
        return {"error": str(e)}


@app.post("/api/services/restart")
def restart_service(
    db: Session = Depends(get_db),
    current_user: User = Depends(require_admin)
):
    """Restart OLT Manager service"""
    import subprocess

    try:
        # Log the action (use 0 for system entity_id)
        log_event(db, 'service_restart', 'system', 0, None,
                 f"Service restart initiated by {current_user.username}")

        # Schedule restart in background (give time for response)
        subprocess.Popen(
            ['bash', '-c', 'sleep 2 && systemctl restart olt-manager'],
            stdout=subprocess.DEVNULL,
            stderr=subprocess.DEVNULL
        )

        return {
            "success": True,
            "message": "Service restart initiated. Please wait 10-15 seconds and refresh."
        }
    except Exception as e:
        logger.error(f"Error restarting service: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@app.post("/api/services/reboot-server")
def services_reboot_server(
    delay: int = 5,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_admin)
):
    """Reboot the server"""
    import subprocess

    try:
        # Log the action (use 0 for system entity_id)
        log_event(db, 'server_reboot', 'system', 0, None,
                 f"Server reboot initiated by {current_user.username}")

        # Schedule reboot
        subprocess.Popen(
            ['bash', '-c', f'sleep {delay} && reboot'],
            stdout=subprocess.DEVNULL,
            stderr=subprocess.DEVNULL
        )

        return {
            "success": True,
            "message": f"Server will reboot in {delay} seconds."
        }
    except Exception as e:
        logger.error(f"Error initiating reboot: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@app.post("/api/services/stop")
def stop_service(
    db: Session = Depends(get_db),
    current_user: User = Depends(require_admin)
):
    """Stop OLT Manager service"""
    import subprocess

    try:
        # Log the action (use 0 for system entity_id)
        log_event(db, 'service_stop', 'system', 0, None,
                 f"Service stop initiated by {current_user.username}")

        # Schedule stop in background
        subprocess.Popen(
            ['bash', '-c', 'sleep 2 && systemctl stop olt-manager'],
            stdout=subprocess.DEVNULL,
            stderr=subprocess.DEVNULL
        )

        return {
            "success": True,
            "message": "Service stop initiated."
        }
    except Exception as e:
        logger.error(f"Error stopping service: {e}")
        raise HTTPException(status_code=500, detail=str(e))


# ============ Network Diagnostic Tools ============

@app.post("/api/tools/ping")
async def ping_host(
    data: dict,
    current_user: User = Depends(require_auth)
):
    """Ping a host and return results"""
    import subprocess
    import re

    host = data.get("host", "").strip()
    count = min(int(data.get("count", 4)), 10)  # Max 10 pings

    if not host:
        raise HTTPException(status_code=400, detail="Host is required")

    # Validate host (prevent command injection)
    if not re.match(r'^[a-zA-Z0-9.\-]+$', host):
        raise HTTPException(status_code=400, detail="Invalid host format")

    try:
        result = subprocess.run(
            ['ping', '-c', str(count), '-W', '2', host],
            capture_output=True,
            text=True,
            timeout=30
        )
        return {
            "success": result.returncode == 0,
            "output": result.stdout + result.stderr,
            "host": host
        }
    except subprocess.TimeoutExpired:
        return {"success": False, "output": "Ping timeout", "host": host}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.post("/api/tools/traceroute")
async def traceroute_host(
    data: dict,
    current_user: User = Depends(require_auth)
):
    """Traceroute to a host"""
    import subprocess
    import re

    host = data.get("host", "").strip()
    max_hops = min(int(data.get("max_hops", 20)), 30)  # Max 30 hops

    if not host:
        raise HTTPException(status_code=400, detail="Host is required")

    # Validate host (prevent command injection)
    if not re.match(r'^[a-zA-Z0-9.\-]+$', host):
        raise HTTPException(status_code=400, detail="Invalid host format")

    try:
        result = subprocess.run(
            ['traceroute', '-m', str(max_hops), '-w', '2', host],
            capture_output=True,
            text=True,
            timeout=60
        )
        return {
            "success": True,
            "output": result.stdout + result.stderr,
            "host": host
        }
    except subprocess.TimeoutExpired:
        return {"success": False, "output": "Traceroute timeout", "host": host}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.post("/api/tools/port-check")
async def check_port(
    data: dict,
    current_user: User = Depends(require_auth)
):
    """Check if a port is open on a host"""
    import socket
    import re

    host = data.get("host", "").strip()
    port = int(data.get("port", 161))  # Default SNMP port
    timeout = min(float(data.get("timeout", 3)), 10)  # Max 10 seconds

    if not host:
        raise HTTPException(status_code=400, detail="Host is required")

    # Validate host (prevent command injection)
    if not re.match(r'^[a-zA-Z0-9.\-]+$', host):
        raise HTTPException(status_code=400, detail="Invalid host format")

    if port < 1 or port > 65535:
        raise HTTPException(status_code=400, detail="Port must be between 1 and 65535")

    try:
        sock = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
        sock.settimeout(timeout)
        result = sock.connect_ex((host, port))
        sock.close()

        is_open = result == 0
        return {
            "success": True,
            "host": host,
            "port": port,
            "is_open": is_open,
            "message": f"Port {port} is {'OPEN' if is_open else 'CLOSED/FILTERED'} on {host}"
        }
    except socket.gaierror:
        return {
            "success": False,
            "host": host,
            "port": port,
            "is_open": False,
            "message": f"Could not resolve hostname: {host}"
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.post("/api/tools/snmp-check")
async def check_snmp(
    data: dict,
    current_user: User = Depends(require_auth)
):
    """Check if SNMP is responding on a host using snmpget command"""
    import subprocess
    import re

    host = data.get("host", "").strip()
    community = data.get("community", "public")
    port = int(data.get("port", 161))

    if not host:
        raise HTTPException(status_code=400, detail="Host is required")

    # Validate host format
    if not re.match(r'^[a-zA-Z0-9.\-]+$', host):
        raise HTTPException(status_code=400, detail="Invalid host format")

    # Validate community string (prevent injection)
    if not re.match(r'^[a-zA-Z0-9_\-]+$', community):
        raise HTTPException(status_code=400, detail="Invalid community string format")

    try:
        # Use snmpget command to query sysDescr (OID 1.3.6.1.2.1.1.1.0)
        result = subprocess.run(
            ['snmpget', '-v', '2c', '-c', community, '-t', '3', '-r', '1',
             f'{host}:{port}', '1.3.6.1.2.1.1.1.0'],
            capture_output=True,
            text=True,
            timeout=10
        )

        if result.returncode == 0 and result.stdout:
            # Parse sysDescr from output
            output = result.stdout.strip()
            # Extract value after "STRING:" or "="
            sys_descr = "Unknown"
            if "STRING:" in output:
                sys_descr = output.split("STRING:")[-1].strip().strip('"')
            elif "=" in output:
                sys_descr = output.split("=")[-1].strip().strip('"')

            return {
                "success": True,
                "host": host,
                "port": port,
                "responding": True,
                "sys_descr": sys_descr,
                "message": f"SNMP responding on {host}:{port}"
            }
        else:
            error_msg = result.stderr.strip() if result.stderr else "No response"
            # Clean up error message
            if "Timeout" in error_msg or "No Response" in error_msg:
                error_msg = "SNMP timeout - device not responding"
            elif "Unknown host" in error_msg:
                error_msg = "Unknown host"

            return {
                "success": False,
                "host": host,
                "port": port,
                "responding": False,
                "message": error_msg
            }
    except subprocess.TimeoutExpired:
        return {
            "success": False,
            "host": host,
            "port": port,
            "responding": False,
            "message": "SNMP check timeout"
        }
    except Exception as e:
        return {
            "success": False,
            "host": host,
            "port": port,
            "responding": False,
            "message": f"SNMP check failed: {str(e)}"
        }


# Catch-all route to serve React frontend (must be LAST)
@app.get("/{full_path:path}")
async def serve_frontend(full_path: str):
    """Serve React frontend for non-API routes"""
    from fastapi.responses import FileResponse, HTMLResponse

    # Skip API routes
    if full_path.startswith("api/"):
        raise HTTPException(status_code=404, detail="Not Found")

    # Use the global STATIC_DIR found at startup
    if not STATIC_DIR:
        raise HTTPException(status_code=404, detail="Frontend not configured")

    index_file = os.path.join(STATIC_DIR, "index.html")

    # Try to serve the exact file first (for assets like favicon, manifest, etc.)
    if full_path:
        file_path = os.path.join(STATIC_DIR, full_path)
        # Contain the resolved path inside STATIC_DIR so "../.." can't read
        # arbitrary files (e.g. /etc/passwd) via the SPA catch-all route.
        real_static = os.path.realpath(STATIC_DIR)
        real_target = os.path.realpath(file_path)
        if (real_target == real_static or real_target.startswith(real_static + os.sep)) \
                and os.path.isfile(real_target):
            return FileResponse(real_target)

    # Otherwise serve index.html for SPA routing
    if os.path.exists(index_file):
        return FileResponse(index_file)

    raise HTTPException(status_code=404, detail="Frontend not found")


if __name__ == "__main__":
    import uvicorn
    uvicorn.run(app, host="0.0.0.0", port=8000)
